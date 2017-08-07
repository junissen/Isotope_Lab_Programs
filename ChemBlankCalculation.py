#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Created on Thu May 25 09:16:28 2017

@author: julianissen

Finalized version of chem blank calculator

"""

import sys
import Tkinter as tk
import tkFileDialog as filedialog
import tkMessageBox as messagebox
import numpy as np
import pandas as pd
import openpyxl
import csv
import os
from itertools import islice
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

class Application(tk.Frame):
    """
    GUI for working with ChemBlankCalculation
    """
    
    def __init__(self, master):
        """
        Initiates Tkinter window for importing reagent blank information
        """
        tk.Frame.__init__(self,master)
        self.dialog_frame_top = tk.Frame(self)
        self.dialog_frame_top.pack()
        tk.Label(self.dialog_frame_top, text = "Welcome to the Chemblank Calculation Program!", font = ('TkDefaultFont', 10)  ).grid(row = 0, column = 0, sticky = 'e')
        self.master.title("ChemBlank Calculator")
        
        #assigns preset values for spike 233 and spike 229 concentrations. If these are unchanged, values from spike dictionary will be used
        global preset_values
        preset_values = [0.0, 0.0]
        
        self.create_widgets()
        self.pack()
        
    def create_widgets(self):
        """
        Prompts whether to change spike concentration values (233 pmol/g and 229 pmol/g)
        """
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = "Would you like to change spike conc values?", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
    
        self.CheckVar_preset_yes = tk.IntVar()
        self.CheckVar_preset_yes.set(0)
        self.option_preset_yes = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_preset_yes, command = self.preset_change).grid(row = 0, column = 1, sticky = 'w')
        
        self.CheckVar_preset_no = tk.IntVar()
        self.CheckVar_preset_no.set(0)
        self.option_preset_no = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_preset_no, command = self.parameter_input).grid(row = 0, column = 2, sticky = 'w')
    
    def parameter_input(self):
        """
        creates manual entry windows for blank name, spike info, spike weight, U weight, Th weight, uptake rate, ionization efficiency
        and chemblank export file name
        """
        
        self.dialog_frame = tk.Frame(self)
        self.dialog_frame.pack()
        
        #chemblank name 
        tk.Label(self.dialog_frame, text = "Enter blank name:  ", font = ('TkDefaultFont', 10) ).grid(row = 0, column = 0, sticky = 'w')
        self.blankname = tk.Entry(self.dialog_frame, background = 'white', width = 12)
        self.blankname.grid(row = 0, column = 1, sticky = 'w')
        self.blankname.focus_set()
        
        #spike used 
        tk.Label(self.dialog_frame, text = "Enter spike information (choose from: DIII-B, DIII-A, 1I, 1H):  ", font = ('TkDefaultFont', 10) ).grid(row = 1, column = 0, sticky = 'w')
        self.spikeinput = tk.Entry(self.dialog_frame, background = 'white', width = 12)
        self.spikeinput.grid(row = 1, column = 1, sticky = 'w')
        self.spikeinput.focus_set()
        
        #spike weight 
        tk.Label(self.dialog_frame, text = "Enter spike weight (g):  ", font = ('TkDefaultFont', 10) ).grid(row = 2, column = 0, sticky = 'w')
        self.spikewt = tk.Entry(self.dialog_frame, background = 'white', width = 12)
        self.spikewt.grid(row = 2, column = 1, sticky = 'w')
        self.spikewt.focus_set()
        
        #U weight
        tk.Label(self.dialog_frame, text = "Enter U weight (g):  ",font = ('TkDefaultFont', 10) ).grid(row = 3, column = 0, sticky = 'w')
        self.U_wt = tk.Entry(self.dialog_frame, background = 'white', width = 12)
        self.U_wt.grid(row = 3, column = 1, sticky = 'w')
        self.U_wt.focus_set()
        
        #Th weight
        tk.Label(self.dialog_frame, text = "Enter Th weight (g):  ", font = ('TkDefaultFont', 10) ).grid(row = 4, column = 0, sticky = 'w')
        self.Th_wt = tk.Entry(self.dialog_frame, background = 'white', width = 12)
        self.Th_wt.grid(row = 4, column = 1, sticky = 'w')
        self.Th_wt.focus_set()
        
        #uptake rate
        tk.Label(self.dialog_frame, text = "Enter uptake rate:  ", font = ('TkDefaultFont', 10) ).grid(row = 5, column = 0, sticky = 'w')
        self.uptake_rate = tk.Entry(self.dialog_frame, background = 'white', width = 12)
        self.uptake_rate.grid(row = 5, column = 1, sticky = 'w')
        self.uptake_rate.focus_set()
        
        #Iionization efficiency
        tk.Label(self.dialog_frame, text = "Enter ionization efficiency:  ", font = ('TkDefaultFont', 10) ).grid(row = 6, column = 0, sticky = 'w')
        self.IE = tk.Entry(self.dialog_frame, background = 'white', width = 12)
        self.IE.grid(row = 6, column = 1, sticky = 'w')
        self.IE.focus_set()
        
        #chemblank filename
        tk.Label(self.dialog_frame, text = "Enter chemblank export file name (include .xlsx):  ", font = ('TkDefaultFont', 10) ).grid(row = 7, column = 0, sticky = 'w')
        self.filename = tk.Entry(self.dialog_frame, background = 'white', width = 12)
        self.filename.grid(row = 7, column = 1, sticky = 'w')
        self.filename.focus_set()
        
        #option of altering Th method
        tk.Label(self.dialog_frame, text = 'Would you like to alter your Th method file before running?: ', font = ('TkDefaultFont', 10) ).grid(row = 8, column = 0, sticky = 'w')
        
        self.CheckVar_th_yes = tk.IntVar()
        self.CheckVar_th_yes.set(0)
        self.th_yes = tk.Checkbutton(self.dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10) , variable = self.CheckVar_th_yes, command = self.th_yes).grid(row = 8, column = 1, sticky = 'w')
        
        self.CheckVar_th_no = tk.IntVar()
        self.CheckVar_th_no.set(0)
        self.th_no = tk.Checkbutton(self.dialog_frame, text = 'No', font = ('TkDefaultFont', 10) , variable = self.CheckVar_th_no, command = self.th_no).grid(row = 8, column = 1, sticky = 'e')
        
    def th_yes(self):
        """
        Changing Th file by specifying which cycle to end on, uploading altered Th files
        """
        
        checkbutton_frame = tk.Frame(self)
        checkbutton_frame.pack()
        
        #altering Th file
        tk.Label(checkbutton_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10) ).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput_th = tk.Entry(checkbutton_frame, background = 'white', width = 12)
        self.rowinput_th.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_th.focus_set()
      
        #uploading Th file        
        self.th_chemblank_upload = tk.Button(checkbutton_frame, text = 'Upload Th chem blank file', font = ('TkDefaultFont', 10) , command = self.file_upload_th_chemblank_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_th_chemblank = tk.IntVar()
        self.CheckVar_th_chemblank.set(0)
        self.th_chemblank_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_th_chemblank).grid(row = 1, column = 1, sticky = 'w')
        
        #uploading Th wash file
        self.th_chemblankwash_upload = tk.Button(checkbutton_frame, text = 'Upload Th chem blank wash file', font = ('TkDefaultFont', 10) , command = self.file_upload_th_chemblankwash).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_th_chemblankwash = tk.IntVar()
        self.CheckVar_th_chemblankwash.set(0)
        self.th_chemblankwash_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_th_chemblankwash).grid(row = 2, column = 1, sticky = 'w')
        
        #option of altering U file
        tk.Label(checkbutton_frame, text = 'Would you like to alter your U method file before running?: ', font = ('TkDefaultFont', 10) ).grid(row = 3, column = 0, sticky = 'w')
        
        self.CheckVar_u_yes = tk.IntVar()
        self.CheckVar_u_yes.set(0)
        self.u_yes = tk.Checkbutton(checkbutton_frame, text = 'Yes', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_yes, command = self.u_yes).grid(row = 3, column = 1, sticky = 'w')
        
        self.CheckVar_u_no = tk.IntVar()
        self.CheckVar_u_no.set(0)
        self.u_no = tk.Checkbutton(checkbutton_frame, text = 'No', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_no, command = self.u_no).grid(row = 3, column = 1, sticky = 'e')
    
    def th_no(self):
        """
        Uploading unaltered Th files
        """
        
        checkbutton_frame = tk.Frame(self)
        checkbutton_frame.pack()
        
        #uploading Th file
        self.th_chemblank_upload = tk.Button(checkbutton_frame, text = 'Upload Th chem blank file', font = ('TkDefaultFont', 10) , command = self.file_upload_th_chemblank).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_th_chemblank = tk.IntVar()
        self.CheckVar_th_chemblank.set(0)
        self.th_chemblank_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_th_chemblank).grid(row = 0, column = 1, sticky = 'w')
        
        #uploading Th wash file
        self.th_chemblankwash_upload = tk.Button(checkbutton_frame, text = 'Upload Th chem blank wash file', font = ('TkDefaultFont', 10) , command = self.file_upload_th_chemblankwash).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_th_chemblankwash = tk.IntVar()
        self.CheckVar_th_chemblankwash.set(0)
        self.th_chemblankwash_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_th_chemblankwash).grid(row = 1, column = 1, sticky = 'w')
        
        #option of altering U file
        tk.Label(checkbutton_frame, text = 'Would you like to alter your U method file before running?: ', font = ('TkDefaultFont', 10) ).grid(row = 2, column = 0, sticky = 'w')
        
        self.CheckVar_u_yes = tk.IntVar()
        self.CheckVar_u_yes.set(0)
        self.u_yes = tk.Checkbutton(checkbutton_frame, text = 'Yes', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_yes, command = self.u_yes).grid(row = 2, column = 1, sticky = 'w')
        
        self.CheckVar_u_no = tk.IntVar()
        self.CheckVar_u_no.set(0)
        self.u_no = tk.Checkbutton(checkbutton_frame, text = 'No', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_no, command = self.u_no).grid(row = 2, column = 2, sticky = 'w')       
    
    def u_yes(self):
        """
        Changing U file by specifying which cycle to end on, uploading altered U files
        """
        
        checkbutton_frame = tk.Frame(self)
        checkbutton_frame.pack()
        
        #altering U file
        tk.Label(checkbutton_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10) ).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput_u = tk.Entry(checkbutton_frame, background = 'white', width = 12)
        self.rowinput_u.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_u.focus_set()
        
        #uploading U file
        self.u_chemblank_upload = tk.Button(checkbutton_frame, text = 'Upload U chem blank file', font = ('TkDefaultFont', 10) , command = self.file_upload_u_chemblank_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_u_chemblank = tk.IntVar()
        self.CheckVar_u_chemblank.set(0)
        self.u_chemblank_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_chemblank).grid(row = 1, column = 1, sticky = 'w')
        
        #uploading U wash file
        self.u_chemblankwash_upload = tk.Button(checkbutton_frame, text = 'Upload U chem blank wash file', font = ('TkDefaultFont', 10) , command = self.file_upload_u_chemblankwash).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_u_chemblankwash = tk.IntVar()
        self.CheckVar_u_chemblankwash.set(0)
        self.u_chemblankwash_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_chemblankwash).grid(row = 2, column = 1, sticky = 'w')
        
        #run chemblank calculation         
        self.chemblank = tk.Button(checkbutton_frame, text = 'Calculate chemblank and export data', font = ('TkDefaultFont', 10) , command = self.blank_calculate, default = 'active').grid(row = 3, column = 1, sticky = 'w')
        
        #quit
        self.quit = tk.Button(checkbutton_frame, text="QUIT", font = ('TkDefaultFont', 10) , command= self.quit_program).grid(row = 3, column = 2, sticky = 'w')
    
    def u_no(self):
        """
        Uploading unaltered U files
        """
        
        checkbutton_frame = tk.Frame(self)
        checkbutton_frame.pack()     
        
        #uploading U file
        self.u_chemblank_upload = tk.Button(checkbutton_frame, text = 'Upload U chem blank file', font = ('TkDefaultFont', 10) , command = self.file_upload_u_chemblank).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_u_chemblank = tk.IntVar()
        self.CheckVar_u_chemblank.set(0)
        self.u_chemblank_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_chemblank).grid(row = 0, column = 1, sticky = 'w')
        
        #uploading U wash file
        self.u_chemblankwash_upload = tk.Button(checkbutton_frame, text = 'Upload U chem blank wash file', font = ('TkDefaultFont', 10) , command = self.file_upload_u_chemblankwash).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_u_chemblankwash = tk.IntVar()
        self.CheckVar_u_chemblankwash.set(0)
        self.u_chemblankwash_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_chemblankwash).grid(row = 1, column = 1, sticky = 'w')
        
        #run chemblank calculation        
        self.chemblank = tk.Button(checkbutton_frame, text = 'Calculate chemblank and export data', font = ('TkDefaultFont', 10) , command = self.blank_calculate, default = 'active').grid(row = 2, column = 1, sticky = 'w')
        
        #quit
        self.quit = tk.Button(checkbutton_frame, text="QUIT", font = ('TkDefaultFont', 10) , command= self.quit_program).grid(row = 2, column = 2, sticky = 'w')
    
    def preset_change(self):
        """
        Runs function to change preset values for 233 and 229 spike concentration
        """
        
        """
        preset values refer to the following:
            [0]: 233 spike concentration (0.0 will be replaced by spike specific value if specified)
            [1]: 229 spike concentration (0.0 will be replaced by 1.0 if 233 changed)
        """
        self.master.withdraw()
        
        Application_preset(self)
        
    def show(self):
        """
        Returns to main window after changing spike preset values and continues to parameter_input()
        """
        
        self.master.update()
        self.master.deiconify()
        
        self.parameter_input()
        
    def quit_program(self):
        """
        Window destroy
        """
        
        self.master.destroy()
        root.quit()
    
    def file_upload_th_chemblank(self):
        """
        Uploads Th chemblank file and once loaded marks checkbox
        """
    
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_th_chemblank = openpyxl.Workbook()
            ws = filename_th_chemblank.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_th_chemblank.save("chemblankth.xlsx")
            self.filename_th_chemblank = "chemblankth.xlsx"
            self.CheckVar_th_chemblank.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_th_chemblank = filename_raw
            self.CheckVar_th_chemblank.set(1)
            
    def file_upload_th_chemblank_option(self):
        """
        Uploads altered Th chemblank file and once loaded marks checkbox
        """
    
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_th_chemblank = openpyxl.Workbook()
            ws = filename_th_chemblank.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in islice(reader, 0, int(self.rowinput_th.get()) + 9):
                    ws.append(row)
            filename_th_chemblank.save("chemblankth.xlsx")
            self.filename_th_chemblank = "chemblankth.xlsx"
            self.CheckVar_th_chemblank.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:])) 
    
    def file_upload_th_chemblankwash(self):
        """
        Uploads Th chemblank wash file and once loaded marks checkbox
        """
    
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_th_chemblankwash = openpyxl.Workbook()
            ws = filename_th_chemblankwash.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_th_chemblankwash.save("chemblankth_wash.xlsx")
            self.filename_th_chemblankwash = "chemblankth_wash.xlsx"
            self.CheckVar_th_chemblankwash.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_th_chemblankwash = filename_raw
            self.CheckVar_th_chemblankwash.set(1)
    
    def file_upload_u_chemblank(self):
        """
        Uploads U chemblank file and once loaded marks checkbox
        """
    
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_u_chemblank = openpyxl.Workbook()
            ws = filename_u_chemblank.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_u_chemblank.save("chemblanku.xlsx")
            self.filename_u_chemblank = "chemblanku.xlsx"
            self.CheckVar_u_chemblank.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_u_chemblank = filename_raw
            self.CheckVar_u_chemblank.set(1)

    def file_upload_u_chemblank_option(self):
        """
        Uploads altered U chemblank file and once loaded marks checkbox
        """
    
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_u_chemblank = openpyxl.Workbook()
            ws = filename_u_chemblank.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in islice(reader, 0, int(self.rowinput_u.get()) + 9):
                    ws.append(row)
            filename_u_chemblank.save("chemblanku.xlsx")
            self.filename_u_chemblank = "chemblanku.xlsx"
            self.CheckVar_u_chemblank.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:])) 
    
    
    def file_upload_u_chemblankwash(self):
        """
        Uploads U chemblank wash file and once loaded marks checkbox
        """
    
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_u_chemblankwash = openpyxl.Workbook()
            ws = filename_u_chemblankwash.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_u_chemblankwash.save("chemblanku_wash.xlsx")
            self.filename_u_chemblankwash = "chemblanku_wash.xlsx"
            self.CheckVar_u_chemblankwash.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_u_chemblankwash = filename_raw
            self.CheckVar_u_chemblankwash.set(1)
    
    
    def blank_calculate(self):
        """
        Calculates wash and chem blank values for all isotopes. Exports an excel file with isotope data.
        """
        self.spike_input = self.spikeinput.get()
        spike = self.spike_input
    
        #derives spike value based off dictionary entries
        spike_six_three_dictionary = {"DIII-B":1.008398,"DIII-A": 1.008398,"1I":1.010128,"1H":1.010128}
        spike_six_three_err_dictionary = {"DIII-B": 0.00015, "DIII-A": 0.00015, "1I": 0.00015, "1H": 0.00015}
        spike_three_dictionary = {"DIII-B": 0.78938, "DIII-A": 0.78933, "1I": 0.61351, "1H": 0.78997}
        spike_three_err_dictionary = {"DIII-B": 0.00002, "DIII-A": 0.00002, "1I": 0.00002, "1H": 0.00002}
        spike_nine_dictionary = {"DIII-B": 0.21734, "DIII-A": 0.21705, "1I": 0.177187, "1H": 0.22815}
        spike_nine_err_dictionary = {"DIII-B": 0.00001, "DIII-A": 0.00002, "1I": 0.00001, "1H": 0.00001}
        spike_zero_nine_dictionary = {"DIII-B": 0.0000625, "DIII-A": 0.0000625, "1I": 0.0000402, "1H": 0.0000402}
        spike_zero_nine_err_dictionary = {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.0000011, "1H": 0.0000011}
        spike_nine_two_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.00, "1H": 0.00}
        spike_nine_two_err_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.00, "1H": 0.00}
        spike_four_three_dictionary = {"DIII-B": 0.003195, "DIII-A": 0.003195, "1I":0.003180, "1H": 0.003180}
        spike_four_three_err_dictionary= {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.000003, "1H": 0.000003}
        spike_five_three_dictionary = {"DIII-B": 0.10532, "DIII-A": 0.10532, "1I": 0.10521, "1H":0.10521}
        spike_five_three_err_dictionary = {"DIII-B": 0.00003, "DIII-A": 0.00003, "1I": 0.00003, "1H": 0.00003}
        spike_eight_three_dictionary = {"DIII-B": 0.01680, "DIII-A": 0.01680, "1I": 0.01700, "1H":0.01700 }
        spike_eight_three_err_dictionary = {"DIII-B": 0.00001, "DIII-A": 0.00001,"1I": 0.00001, "1H": 0.00001}
        spike_three_nine_dictionary = {"DIII-B": 0.27533001, "DIII-A": 0.27498005, "1I": 0.28880866, "1H": 0.28880844}

        if spike in spike_six_three_dictionary:
            self.spike_six_three = float(spike_six_three_dictionary[spike]) #spike ratio
        else: 
            messagebox.showwarning("Error!", "No valid spike info entered! ")
        
        if spike in spike_six_three_err_dictionary: 
            self.spike_six_three_err = float(spike_six_three_err_dictionary[spike]) #error of spike ratio
        
        #if spike 233 concentration has been changed, uses input value. Otherwise uses spike dictionary value
        if preset_values[0] == 0.0:
            if spike in spike_three_dictionary:
                self.spike_three = float(spike_three_dictionary[spike]) #in pmol/g
            else:pass
        else: self.spike_three = float(preset_values[0])
    
        if spike in spike_three_err_dictionary:
            self.spike_three_err = float(spike_three_err_dictionary[spike]) #in pmol/g
        else:pass
    
        #if spike 233 concentration has been changed, multiplies input value by spike 233/229 ratio. Otherwises uses spike dictionary value
        if preset_values[1] == 1.0:
            if spike in spike_three_nine_dictionary:
                self.spike_nine = float(spike_three_nine_dictionary[spike]) * self.spike_three
            else: pass
        else: 
            if spike in spike_nine_dictionary:
                self.spike_nine = float(spike_nine_dictionary[spike]) #in pmol/g
            else: pass
    
        if spike in spike_nine_err_dictionary: 
            self.spike_nine_err = float(spike_nine_err_dictionary[spike]) #in pmol/g
        else: pass
    
        if spike in spike_zero_nine_dictionary:
            self.spike_zero_nine = float(spike_zero_nine_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_zero_nine_err_dictionary:
            self.spike_zero_nine_err = float(spike_zero_nine_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_nine_two_dictionary: 
            self.spike_nine_two = float(spike_nine_two_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_nine_two_err_dictionary:
            self.spike_nine_two_err = float(spike_nine_two_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_four_three_dictionary:
            self.spike_four_three = float(spike_four_three_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_four_three_err_dictionary:
            self.spike_four_three_err = float(spike_four_three_err_dictionary[spike]) #error of spike ratio
        else: pass
            
        if spike in spike_five_three_dictionary:
            self.spike_five_three = float(spike_five_three_dictionary[spike]) #spike ratio
        else: pass
        
        if spike in spike_five_three_err_dictionary:
            self.spike_five_three_err = float(spike_five_three_err_dictionary[spike]) #error of spike ratio
        else: pass
        
        if spike in spike_eight_three_dictionary:
            self.spike_eight_three = float(spike_eight_three_dictionary[spike]) #spike ratio
        else: pass
        
        if spike in spike_eight_three_err_dictionary:
            self.spike_eight_three_err = float(spike_eight_three_err_dictionary[spike]) #error of spike ratio
        else: pass
        
        #sample information
        self.blank_name = self.blankname.get()
        self.spike_wt = float(self.spikewt.get())
        self.U_wt = float(self.U_wt.get())
        self.Th_wt = float(self.Th_wt.get())
        self.uptake_rate = float(self.uptake_rate.get())
        self.IE = (float(self.IE.get())/100) 
        self.filename = self.filename.get()
        
        #spike in chem blank (pmol)
        self.spike_three_used = self.spike_three * self.spike_wt
        self.spike_nine_used = self.spike_nine * self.spike_wt
        
        #constants to be used throughout the class
        self.wt_229 = 229.031756
        self.wt_230 = 230.033128
        self.wt_232 = 232.038051
        self.wt_233 = 233.039629
        self.wt_234 = 234.040947
        self.wt_235 = 235.043924
        self.wt_236 = 236.045563
        self.wt_238 = 238.050785
        
        """
        Th wash and chem blank values
        """
        #wash 229 Th
        working_a = chem_blank(self.filename_th_chemblankwash, "C", "229")
        nine_wash = working_a.calc()
        
        
        #chem blank 229 Th
        working_b = chem_blank(self.filename_th_chemblank, "C", "229")
        nine = working_b.calc()
        
        
        #wash 230 Th
        working_c = chem_blank(self.filename_th_chemblankwash, "D", "230")
        zero_wash = working_c.calc()
        
        
        #chem blank 230 Th
        working_d = chem_blank(self.filename_th_chemblank, "D", "230")
        zero = working_d.calc()
        
        
        #wash 232 Th
        working_e = chem_blank(self.filename_th_chemblankwash, "E", "232")
        two_wash = working_e.calc()
        
        
        #chem blank 232Th
        working_f = chem_blank(self.filename_th_chemblank, "E", "232")
        two = working_f.calc()
        
        #230 unfiltered beam for plot
        self.zero_beam_array = working_d.array()
            
        #230 index array
        working_z = chem_blank(self.filename_th_chemblank, "A", "other")
        self.index_array_Th = working_z.array()
        
        
        """
        U wash and chem blank values
        """
        #two options for U wash and U chem blank values in case 232Th was not included in method file
        
        wb_Uwash = openpyxl.load_workbook(self.filename_u_chemblankwash)
        ws_Uwash = wb_Uwash.worksheets[0]
        
        if ws_Uwash['C1'].value == '1:232Th':
        
            #wash 233U
            working_g = chem_blank(self.filename_u_chemblankwash, "D", "233")
            three_wash = working_g.calc()
            
            #wash 234U
            working_h = chem_blank(self.filename_u_chemblankwash, "E", "234")
            four_wash = working_h.calc()
            
            #wash 235U
            working_i = chem_blank(self.filename_u_chemblankwash, "F", "235")
            five_wash = working_i.calc()
            
            #wash 236U
            working_j = chem_blank(self.filename_u_chemblankwash, "G", "236")
            six_wash = working_j.calc()
            
            #wash 238U
            working_k = chem_blank(self.filename_u_chemblankwash, "H", "238")
            eight_wash = working_k.calc()
            
        elif ws_Uwash['C1'].value == '1:233U':
            
            #wash 233U
            working_g = chem_blank(self.filename_u_chemblankwash, "C", "233")
            three_wash = working_g.calc()
            
            #wash 234U
            working_h = chem_blank(self.filename_u_chemblankwash, "D", "234")
            four_wash = working_h.calc()
            
            #wash 235U
            working_i = chem_blank(self.filename_u_chemblankwash, "E", "235")
            five_wash = working_i.calc()
            
            #wash 236U
            working_j = chem_blank(self.filename_u_chemblankwash, "F", "236")
            six_wash = working_j.calc()
            
            #wash 238U
            working_k = chem_blank(self.filename_u_chemblankwash, "G", "238")
            eight_wash = working_k.calc()
        
        wb_U = openpyxl.load_workbook(self.filename_u_chemblank)
        ws_U = wb_U.worksheets[0] 
        
        if ws_U['C1'].value == '1:232Th': 
            
            #chem blank 233U
            working_l = chem_blank(self.filename_u_chemblank, "D", "233")
            three = working_l.calc()
            
            #chem blank 234U
            working_m = chem_blank(self.filename_u_chemblank, "E", "234")
            four = working_m.calc()
            
            #chem blank 235U
            working_n = chem_blank(self.filename_u_chemblank, "F", "235")
            five = working_n.calc()
            
            #chem blank 236U
            working_o = chem_blank(self.filename_u_chemblank, "G", "236")
            six = working_o.calc()
            
            #chem blank 238U
            working_p = chem_blank(self.filename_u_chemblank, "H", "238")
            eight = working_p.calc()
            
            #234 unfiltered beam for plot
            self.four_beam_array = working_m.array()
            
            #234 index array
            working_q = chem_blank(self.filename_u_chemblank, "A", "other")
            self.index_array_U = working_q.array()

        elif ws_U['C1'].value == '1:233U':
            
            #chem blank 233U
            working_l = chem_blank(self.filename_u_chemblank, "C", "233")
            three = working_l.calc()
            
            #chem blank 234U
            working_m = chem_blank(self.filename_u_chemblank, "D", "234")
            four = working_m.calc()
            
            #chem blank 235U
            working_n = chem_blank(self.filename_u_chemblank, "E", "235")
            five = working_n.calc()
            
            #chem blank 236U
            working_o = chem_blank(self.filename_u_chemblank, "F", "236")
            six = working_o.calc()
            
            #chem blank 238U
            working_p = chem_blank(self.filename_u_chemblank, "G", "238")
            eight = working_p.calc()
            
            #234 index array
            working_q = chem_blank(self.filename_u_chemblank, "A", "other")
            self.index_array_U = working_q.array()
            
        
        #deleting excel files
        try:
            os.remove("chemblankth_wash.xlsx")
            os.remove("chemblankth.xlsx")
            os.remove("chemblanku_wash.xlsx")
            os.remove("chemblanku.xlsx")
        except: pass
        
        """
        Calculates signal isotopic ratio and 2s error
        
        Note: [0]: mean, [1]: counts, [2] = 2s rel error
        
        """
        #230/229
        zero_nine = (zero[0] - zero_wash[0]) / (nine[0] - nine_wash[0])
        zero_nine_err = np.sqrt( ((zero[0]*zero[2])**2/(nine[0] - nine_wash[0])**2) + 
                                 ((zero_wash[0]*zero_wash[2])**2/(nine_wash[0]-nine[0])**2) + 
                                 ((nine[0]*nine[2])**2 * ((zero_wash[0]-zero[0])/((nine[0]-nine_wash[0])**2))**2) + 
                                 ((nine_wash[0]*nine_wash[2])**2 * ((zero[0]-zero_wash[0])/((nine[0]-nine_wash[0])**2))**2)
                                 )/zero_nine
        
        #229/232 
        nine_two = (nine[0] - nine_wash[0]) / (two[0] - two_wash[0])
        nine_two_err = np.sqrt( ((nine[0]*nine[2])**2/(two[0] - two_wash[0])**2) + 
                                 ((nine_wash[0]*nine_wash[2])**2/(two_wash[0]-two[0])**2) + 
                                 ((two[0]*two[2])**2 * ((nine_wash[0]-nine[0])/((two[0]-two_wash[0])**2))**2) + 
                                 ((two_wash[0]*two_wash[2])**2 * ((nine[0]-nine_wash[0])/((two[0]-two_wash[0])**2))**2)
                                 )/nine_two
       
        #234/233
        four_three = (four[0] - four_wash[0])/(three[0] - three_wash[0])
        four_three_err = np.sqrt( ((four[0]*four[2])**2/(three[0] - three_wash[0])**2) + 
                                 ((four_wash[0]*four_wash[2])**2/(three_wash[0]-three[0])**2) + 
                                 ((three[0]*three[2])**2 * ((four_wash[0]-four[0])/((three[0]-three_wash[0])**2))**2) + 
                                 ((three_wash[0]*three_wash[2])**2 * ((four[0]-four_wash[0])/((three[0]-three_wash[0])**2))**2)
                                 )/four_three
        
        #235/233
        five_three = (five[0] - five_wash[0])/(three[0] - three_wash[0])
        five_three_err = np.sqrt( ((five[0]*five[2])**2/(three[0] - three_wash[0])**2) + 
                                 ((five_wash[0]*five_wash[2])**2/(three_wash[0]-three[0])**2) + 
                                 ((three[0]*three[2])**2 * ((five_wash[0]-five[0])/((three[0]-three_wash[0])**2))**2) + 
                                 ((three_wash[0]*three_wash[2])**2 * ((five[0]-five_wash[0])/((three[0]-three_wash[0])**2))**2)
                                 )/five_three
        
        #236/233 
        six_three = (six[0] - six_wash[0])/(three[0] - three_wash[0])
        
        #238/233
        eight_three = (eight[0] - eight_wash[0])/(three[0] - three_wash[0])
        eight_three_err = np.sqrt( ((eight[0]*eight[2])**2/(three[0] - three_wash[0])**2) + 
                                 ((eight_wash[0]*eight_wash[2])**2/(three_wash[0]-three[0])**2) + 
                                 ((three[0]*three[2])**2 * ((eight_wash[0]-eight[0])/((three[0]-three_wash[0])**2))**2) + 
                                 ((three_wash[0]*three_wash[2])**2 * ((eight[0]-eight_wash[0])/((three[0]-three_wash[0])**2))**2)
                                 )/eight_three
        
        """
        Corrects signal isotopic ratios for fractionation
    
        """
        #230/229 fract. corrected
        zero_nine_corr = zero_nine * (self.spike_six_three/six_three)**(np.log(self.wt_230/self.wt_229)/np.log(self.wt_236/self.wt_233))
        
        #229/232 fract. corrected
        nine_two_corr = nine_two * (self.spike_six_three/six_three)**(np.log(self.wt_229/self.wt_232)/np.log(self.wt_236/self.wt_233))
        
        #234/233 fract. corrected
        
        four_three_corr = four_three * (self.spike_six_three/six_three)**(np.log(self.wt_234/self.wt_233)/np.log(self.wt_236/self.wt_233))
        
        #235/233 fract. corrected
        
        five_three_corr = five_three * (self.spike_six_three/six_three)**(np.log(self.wt_235/self.wt_233)/np.log(self.wt_236/self.wt_233))
        
        #238/233 fract. corrected
        
        eight_three_corr = eight_three * (self.spike_six_three/six_three)**(np.log(self.wt_238/self.wt_233)/np.log(self.wt_236/self.wt_233))

        
        """
        2s relative spike errors
        """
        
        zero_nine_spike_err = self.spike_zero_nine_err/self.spike_zero_nine
        
        nine_two_spike_err = 0 #may need to change for different spikes
        
        four_three_spike_err = self.spike_four_three_err/self.spike_four_three
        
        five_three_spike_err = self.spike_five_three_err/self.spike_five_three
        
        eight_three_spike_err = self.spike_eight_three_err/self.spike_eight_three
        
        """
        Corrects for fractionation
        """
        
        #230/229 spike corrected
        
        zero_nine_corr_spike = zero_nine_corr - self.spike_zero_nine
        
        zero_nine_corr_spike_err = np.sqrt((zero_nine * zero_nine_err)**2 + (self.spike_zero_nine * zero_nine_spike_err)**2)/ abs(zero_nine_corr_spike)
        
        #229/232 spike corrected
        
        nine_two_corr_spike = nine_two_corr - self.spike_nine_two
        
        nine_two_corr_spike_err = np.sqrt((nine_two * nine_two_err)**2 + (self.spike_nine_two * nine_two_spike_err)**2)/ abs(nine_two_corr_spike)
        
        #234/233 spike corrected
        
        four_three_corr_spike = four_three_corr - self.spike_four_three
        
        four_three_corr_spike_err = np.sqrt((four_three * four_three_err)**2 + (self.spike_four_three * four_three_spike_err)**2)/ abs(four_three_corr_spike)
        
        #235/233 spike corrected
        
        five_three_corr_spike = five_three_corr - self.spike_five_three
        
        five_three_corr_spike_err = np.sqrt((five_three * five_three_err)**2 + (self.spike_five_three * five_three_spike_err)**2)/ abs(five_three_corr_spike)
        
        #238/233 spike corrected 
        
        eight_three_corr_spike = eight_three_corr - self.spike_eight_three
        
        eight_three_corr_spike_err = np.sqrt((eight_three * eight_three_err)**2 + (self.spike_eight_three * eight_three_spike_err)**2)/ abs(eight_three_corr_spike)
        
        """
        Chemistry yields in %
        """
        
        th_yield = ((nine[0] - nine_wash[0])/( (self.spike_nine_used/(10**12)) * (6.022E23) * self.IE * (self.uptake_rate/1000) * (1/self.Th_wt))) * 100 
        u_yield = ((six[0] - six_wash[0])/( (self.spike_three_used * self.spike_six_three/(10**12)) * (6.022E23) * self.IE * (self.uptake_rate/1000) * (1/self.U_wt))) * 100 
        
        
        """
        Chemistry blank values in grams
        """
        
        #230 chemblank ag
        zero_chemblank = ((self.spike_nine_used * zero_nine_corr_spike)/(10**12))* self.wt_230 * (10**18) 
        zero_chemblank_err = abs(zero_chemblank * zero_nine_corr_spike_err)
        
        #232 chemblank fg
        two_chemblank = ((self.spike_nine_used / nine_two_corr_spike)/(10**12))* self.wt_232 * (10**15)
        two_chemblank_err = abs(two_chemblank * nine_two_corr_spike_err)
        
        #234 chemblank ag
        four_chemblank = ((self.spike_three_used * four_three_corr_spike)/(10**12))* self.wt_234 * (10**18) 
        four_chemblank_err = abs(four_chemblank * four_three_corr_spike_err)
        
        #235 chemblank fg
        five_chemblank = ((self.spike_three_used * five_three_corr_spike)/(10**12))* self.wt_235 * (10**15) 
        five_chemblank_err = abs(five_chemblank * five_three_corr_spike_err)
        
        #238 chemblank fg
        eight_chemblank = ((self.spike_three_used * eight_three_corr_spike)/(10**12))* self.wt_238 * (10**15) 
        eight_chemblank_err = abs(eight_chemblank * eight_three_corr_spike_err)
        
        data = {'1_Chemblank': pd.Series([self.blank_name, "Run info"],index = ['1_fileinfo', '6_param']),
                '229Th': pd.Series([th_yield, "%", 'Spike used', self.spike_input], index = ['2_yields', '5_units', '7_param', '8_param']),
                '230Th': pd.Series([zero_chemblank, zero_chemblank_err, 'ag', 'Spike wt', self.spike_wt, 'g'], index = ['3_chemblank', '4_2s err', '5_units', '7_param', '8_param', '9_param']),
                '232Th': pd.Series([two_chemblank, two_chemblank_err, 'fg', 'Th wt', self.Th_wt, 'g'], index = ['3_chemblank', '4_2s err', '5_units', '7_param', '8_param', '9_param']),
                '234U': pd.Series([four_chemblank, four_chemblank_err, 'ag', 'U wt', self.U_wt, 'g'], index = ['3_chemblank', '4_2s err', '5_units', '7_param', '8_param', '9_param']),
                '233U': pd.Series([u_yield, "%", 'U.R.', self.uptake_rate, 'mg/sec'], index = ['2_yields', '5_units', '7_param', '8_param', '9_param']),
                '235U': pd.Series([five_chemblank, five_chemblank_err, 'fg', 'I.E.', (self.IE * 100), '%'], index = ['3_chemblank', '4_2s err', '5_units', '7_param', '8_param', '9_param']),
                '238U': pd.Series([eight_chemblank, eight_chemblank_err, 'fg'], index = ['3_chemblank', '4_2s err', '5_units'])}
                
        df = pd.DataFrame(data)
        
        writer = pd.ExcelWriter(self.filename, engine = 'openpyxl')
        
        df.to_excel(writer)
        
        writer.save()
        
        messagebox.showinfo("Chemblank data file saved ! ", "Chemblank data file name: "+ str(self.filename))
        
        #plots 234 and 230 beam to check for beam stability
        wb = plot_figure(self.four_beam_array, self.index_array_U, self.zero_beam_array, self.index_array_Th)
        
        wb.plot_fig()
        
class chem_blank():
        """
        Class for analyzing ICP-MS files
        """
        def __init__(self,filename, columnletter, int_time): 
            """
            Uploads column from specified file 
            """
            self.column = str(columnletter)+'{}:'+str(columnletter)+'{}'
            self.filename = str(filename)
            self.workbook = openpyxl.load_workbook(self.filename)
            self.ws = self.workbook.active
            
            int_time = str(int_time)
            
            int_dictionary = {"229":0.131, "230":1.049, "232":0.262, "233":0.131, "234":1.049,
                              "235": 0.262, "236":0.131, "238": 0.262, "other": 0.0}
            
            if int_time in int_dictionary:
                self.inttime = int_dictionary[int_time]
            else: print "Int_time not available"
                      
        def calc(self):
            """
            Code calculates the mean, total cycles, and 2s counting statistics error of Excel column
            """
            outlist = []
            outcounts = 0
            for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
                for cell in row:
                    if cell.value: 
                        try:
                            value = float(cell.value)
                            outlist.append(value)
                            outcounts += 1
                        except:
                            outlist.append(np.nan)
                    elif cell.value == 0:
                        value= 0.00
                        outlist.append(value)
                        outcounts += 1
                    else: outlist.append(np.nan)
            outarray = np.array(outlist, dtype = np.float)
            self.mean = np.nanmean(a = outarray)
            standdev = np.nanstd(a = outarray, ddof = 1)
            self.counts = outcounts
            err_abs =  2 * standdev/((self.counts)**0.5)
            err_rel_option1 = err_abs/self.mean
            err_rel_option2 = 2/((self.mean * self.counts*self.inttime)**0.5)
            self.err_rel = max(err_rel_option1, err_rel_option2)
            
            lst_Chem = [self.mean, self.counts, self.err_rel]
        
            return lst_Chem
        
        def array(self):
            """
            Code provides output array for Excel row. Includes NaN for non-values
            """
            outlist = []
            for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
                for cell in row:
                    if cell.value: 
                        try:
                            value = float(cell.value)
                            outlist.append(value)
                        except:
                            outlist.append(np.nan)
                    elif cell.value == 0:
                        value= 0.00
                        outlist.append(value)
                    else: outlist.append(np.nan)
            outarray = np.array(outlist, dtype = np.float)
            return outarray
        
        
class plot_figure(tk.Tk):
    """
    Provides plot of 234U and 230Th beam stability
    """
    
    def __init__(self, U, Uindex, Th, Thindex):
        """
        Init def
        """
        self.U = U
        self.Uindex = Uindex
        self.Th = Th
        self.Thindex = Thindex
      
    def plot_fig(self): 
        """
        plot of 234U and 230Th beam
        """
        toplevel = tk.Toplevel()
        toplevel.title("Beam intensity")
        fig = Figure(figsize = (8, 6))
        
        ax1 = fig.add_subplot(2,1,1)
        x1 = self.Uindex
        y1 = self.U
        y1mean = np.nanmean(y1)
        y1standdev = np.nanstd(a = y1, ddof = 1)
        ax1.scatter(x1, y1, color = 'b', marker = 'o')
        ax1.set_xlabel('Cycles', fontsize = 7, labelpad = 0.5)
        ax1.set_ylabel('234U', fontsize = 7, labelpad = 0.5)
        ax1.set_ylim([y1mean - 10*y1standdev, y1mean + 10*y1standdev])
        ax1.tick_params(labelsize = 5)
        ax1.set_title('Beam Intensity 234U' , fontsize = 8)
        
        ax2 = fig.add_subplot(2,1,2)
        x2 = self.Thindex
        y2 = self.Th
        y2mean = np.nanmean(y2)
        y2standdev = np.nanstd(a = y2, ddof = 1)
        ax2.scatter(x2, y2, color = 'c', marker = 'o')
        ax2.set_xlabel('Cycles', fontsize = 7, labelpad = 0.5)
        ax2.set_ylabel('230Th', fontsize = 7, labelpad = 0.5)
        ax2.set_ylim([y2mean - 10*y2standdev, y2mean + 10*y2standdev])
        ax2.tick_params(labelsize = 5)
        ax2.set_title('Beam Intensity 230Th', fontsize = 8)
        
        fig.set_tight_layout(True)
        
        canvas = FigureCanvasTkAgg(fig, master = toplevel)
        canvas.show()
        canvas.get_tk_widget().pack()
        toplevel.mainloop()

class Application_preset(tk.Toplevel):
    """
    Tkinter class for changing preset values
    """
    def __init__(self, original):
        """
        Opens new Tkinter window if you have chosen to change your spike concentration values
        """
        self.original_frame = original
        self.otherframe = tk.Toplevel()
        self.otherframe.protocol("WM_DELETE_WINDOW", on_closing)
        self.otherframe.title("Preset values for ChemBlank Calculation")
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        tk.Label(dialog_frame, text = "Options for changing preset values", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        
        
        self.spike_conc_option()
        
    def spike_conc_option(self):
        """
        Option for changing 233U concentration of spike
        """
        
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        
        #option for changing 233U concentration
        tk.Label(dialog_frame, text = "Would you like to change the 233U concentration of your spike?", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        
        self.CheckVar_spike_yes = tk.IntVar()
        self.CheckVar_spike_yes.set(0)
        self.CheckVar_spike_yes = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_spike_yes, command = self.spike_yes).grid(row = 0, column = 1, sticky = 'w')
        
        self.CheckVar_spike_no = tk.IntVar()
        self.CheckVar_spike_no.set(0)
        self.CheckVar_spike_no = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_spike_no, command = self.spike_no).grid(row = 0, column = 2, sticky = 'e')
        
    def spike_yes(self):
        """
        Prompt for changing 233U concentration and submit 
        """
        
        self.spike_yes = 1
        
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        
        #spike 233U concentration
        tk.Label(dialog_frame, text = "Enter 233U concentration in pmol/g:", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        self.spike_conc_three = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spike_conc_three.grid(row = 0, column = 1, sticky = 'w')
        self.spike_conc_three.focus_set()
        
        #submit value 
        self.submit_button = tk.Button(dialog_frame, text = "Submit", font = ('TkDefaultFont', 10), default = "active", command = self.click_submit).grid(row = 1, column = 0)
    
    def spike_no(self):
        """
        Submit
        """
        
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        
        self.submit_button = tk.Button(dialog_frame, text = "Submit", font = ('TkDefaultFont', 10), default = "active", command = self.click_submit).grid(row = 0, column = 0)
        
    def click_submit(self):
        """
        If spike concentration has been changed, updates 233U concentration in global preset values with input value. Updates 229 concentration 
        with 1.0. If spike concentration has not been changed, global preset values remain the same. Returns to oroginal window
        """
        
        #changing spike preset values
        if self.spike_yes == 1:
            spike_conc_three = self.spike_conc_three.get()
            spike_conc_nine = 1.0
        
            preset_values[0] = spike_conc_three
            preset_values[1] = spike_conc_nine
        
        #return to original window
        self.otherframe.destroy()
        self.original_frame.show()
    
    def on_closing(self):
        """
        If secondary window is X'ed out of, prompts you to quit 
        """
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            self.otherframe.destroy()       
            
root = tk.Tk()
    
app = Application(master=root)

def on_closing():
    """
    If primary window is X'ed out of, prompts you to quit 
    """
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        root.destroy()
        root.quit()

root.protocol("WM_DELETE_WINDOW", on_closing)

root.mainloop()    
        