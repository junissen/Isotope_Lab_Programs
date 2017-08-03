#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Tue Jul 18 11:23:56 2017

@author: julianissen

Finalized version of Age Calculator for MC-ICP-MS U/Th runs
"""

import sys
import Tkinter as tk
import tkFileDialog as filedialog
import tkMessageBox as messagebox
import openpyxl
import csv
import numpy as np
import os
from scipy.optimize import curve_fit
from scipy.optimize import fsolve
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from itertools import islice
import datetime

class Application(tk.Frame):
    """
    Root GUI for working with AgeCalculation. Gives option of either running SEM or Cups for U/Th.
    """
    def __init__(self, master):
        
        tk.Frame.__init__(self, master)
        
        self.master.title("Age Calculator")
        self.dialog_frame_top = tk.Frame(self)
        self.dialog_frame_top.pack()
        tk.Label(self.dialog_frame_top, text = "Welcome to the Age Calculation Program!", font = ('TkDefaultFont', 10) ).grid(row = 0, column = 0, sticky = 'e')
        
        global preset_values
        preset_values = [0.0, 0.0, 0.0001, 0.0003, 4.4E-6, 2.2E-6]
        
        self.pack()
        self.create_widgets()
    
    def create_widgets(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = "Would you like to change preset values?", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        
        self.CheckVar_preset_yes = tk.IntVar()
        self.CheckVar_preset_yes.set(0)
        self.option_preset_yes = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_preset_yes, command = self.preset_change).grid(row = 0, column = 1, sticky = 'w')
        
        self.CheckVar_preset_no = tk.IntVar()
        self.CheckVar_preset_no.set(0)
        self.option_preset_no = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_preset_no, command = self.method_used).grid(row = 0, column = 2, sticky = 'w')

    def method_used(self):    
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = "Check which method used for U run: ", font = ('TkDefaultFont', 10) ).grid(row = 0, column = 0, sticky = 'w')
        self.CheckVar_sem_U = tk.IntVar()
        self.CheckVar_sem_U.set(0)
        self.checkbutton_sem_U = tk.Checkbutton(dialog_frame, text = 'SEM', font = ('TkDefaultFont', 10), variable = self.CheckVar_sem_U, command = self.upload_sem).grid(row = 0, column = 1, sticky = 'w')
        
        self.CheckVar_cups_U = tk.IntVar()
        self.CheckVar_cups_U.set(0)
        self.checkbutton_cups_U = tk.Checkbutton(dialog_frame, text = 'CUPS', font = ('TkDefaultFont', 10), variable = self.CheckVar_cups_U, command = self.cups_command_U).grid(row = 0, column = 2, sticky = 'w')
        
    """
    SEM uploads
    """
        
    def upload_sem(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        self.parameter_input()
        
        
    def parameter_input(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = "Enter spike information (choose from: DIII-B, DIII-A, 1I, 1H):  ", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        self.spikeinput = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spikeinput.grid(row = 0, column = 1, sticky = 'w')
        self.spikeinput.focus_set()
        
        tk.Label(dialog_frame, text = "Enter abundant sensitivity for 237U-238U for U run: ", font = ('TkDefaultFont', 10)).grid(row = 1, column = 0, sticky = 'w')
        self.AS_U_input = tk.Entry(dialog_frame, background = 'white', width= 12)
        self.AS_U_input.grid(row = 1, column = 1, sticky = 'w')
        self.AS_U_input.focus_set()
        
        tk.Label(dialog_frame, text = "Is the AS the same for Th run? ", font = ('TkDefaultFont', 10) ).grid(row = 2, column = 0, sticky = 'w')
        
        self.CheckVar_AS_Th_yes = tk.IntVar()
        self.CheckVar_AS_Th_yes.set(0)
        self.checkbutton_AS_Th_yes = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_AS_Th_yes, command = self.AS_Th_yes).grid(row = 2, column = 1, sticky = 'w')
        
        self.CheckVar_AS_Th_no = tk.IntVar()
        self.CheckVar_AS_Th_no.set(0)
        self.checkbutton_AS_Th_no = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_AS_Th_no, command = self.AS_Th_no).grid(row = 2, column = 1, sticky = 'e')
        
    def AS_Th_yes(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = "Enter sample weight (g): ", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        self.sample_wt = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.sample_wt.grid(row = 0, column = 1, sticky = 'w')
        self.sample_wt.focus_set()
        
        tk.Label(dialog_frame, text = "Enter spike weight (g): ", font = ('TkDefaultFont', 10)).grid(row = 1, column = 0, sticky = 'w')
        self.spike_wt = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spike_wt.grid(row = 1, column = 1, sticky = 'w')
        self.spike_wt.focus_set()
        
        tk.Label(dialog_frame, text = "Enter sample ID: ", font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'w')
        self.sample_ID = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.sample_ID.grid(row = 2, column = 1, sticky = 'w')
        self.sample_ID.focus_set()
        
        tk.Label(dialog_frame, text = "Enter row # for age spreadsheet (starting at 6): ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'w')
        self.row_age = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.row_age.grid(row = 3, column = 1, sticky = 'w')
        self.row_age.focus_set()
        
        tk.Label(dialog_frame, text = "Would you like to alter your Th file? ", font = ('TkDefaultFont', 10)).grid(row = 4, column = 0, sticky = 'e')
        
        self.CheckVar_Th_yes_sem = tk.IntVar()
        self.CheckVar_Th_yes_sem.set(0)
        self.option_th_yes_sem = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_yes_sem, command = self.Th_yes_sem).grid(row = 4, column = 1, sticky = 'w')
        
        self.CheckVar_Th_no_sem = tk.IntVar()
        self.CheckVar_Th_no_sem.set(0)
        self.option_th_no_sem = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_no_sem, command = self.Th_no_sem).grid(row = 4, column = 1, sticky = 'e')
    
    def AS_Th_no(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        self.AS_Th_no = 1
        
        tk.Label(dialog_frame, text = "Enter abundant sensitivity for 237U-238U for Th run: ",  font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        self.AS_Th_input = tk.Entry(dialog_frame, background = 'white', width= 12)
        self.AS_Th_input.grid(row = 0, column = 1, sticky = 'w')
        self.AS_Th_input.focus_set()
        
        tk.Label(dialog_frame, text = "Enter sample weight (g): ", font = ('TkDefaultFont', 10)).grid(row = 1, column = 0, sticky = 'w')
        self.sample_wt = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.sample_wt.grid(row = 1, column = 1, sticky = 'w')
        self.sample_wt.focus_set()
        
        tk.Label(dialog_frame, text = "Enter spike weight (g): ", font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'w')
        self.spike_wt = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spike_wt.grid(row = 2, column = 1, sticky = 'w')
        self.spike_wt.focus_set()
        
        tk.Label(dialog_frame, text = "Enter sample ID: ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'w')
        self.sample_ID = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.sample_ID.grid(row = 3, column = 1, sticky = 'w')
        self.sample_ID.focus_set()
        
        tk.Label(dialog_frame, text = "Enter row # for age spreadsheet (starting at 6): ", font = ('TkDefaultFont', 10)).grid(row = 4, column = 0, sticky = 'w')
        self.row_age = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.row_age.grid(row = 4, column = 1, sticky = 'w')
        self.row_age.focus_set()
        
        tk.Label(dialog_frame, text = "Would you like to alter your Th file? ", font = ('TkDefaultFont', 10)).grid(row = 5, column = 0, sticky = 'e')
        
        self.CheckVar_Th_yes_sem = tk.IntVar()
        self.CheckVar_Th_yes_sem.set(0)
        self.option_th_yes_sem = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_yes_sem, command = self.Th_yes_sem).grid(row = 5, column = 1, sticky = 'w')
        
        self.CheckVar_Th_no_sem = tk.IntVar()
        self.CheckVar_Th_no_sem.set(0)
        self.option_th_no_sem = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_no_sem, command = self.Th_no_sem).grid(row = 5, column = 1, sticky = 'e')
    
    def Th_yes_sem(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10) ).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput_Th = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.rowinput_Th.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_Th.focus_set()
        
        self.Th_checkbutton = tk.Button(dialog_frame, text = "Upload Th file", font = ('TkDefaultFont', 10), command = self.file_Th_upload_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_Th_upload = tk.IntVar()
        self.CheckVar_Th_upload.set(0)
        self.Th_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_upload).grid(row = 1, column = 1, sticky = 'w')
        
        self.Th_wash_checkbutton = tk.Button(dialog_frame, text = "Upload Th wash file", font = ('TkDefaultFont', 10), command = self.file_Thwash_upload).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_Thwash_upload = tk.IntVar()
        self.CheckVar_Thwash_upload.set(0)
        self.Thwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_Thwash_upload).grid(row = 2, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "Would you like to alter your U file? ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'w')
        
        self.CheckVar_U_yes_sem = tk.IntVar()
        self.CheckVar_U_yes_sem.set(0)
        self.option_u_yes_sem = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_yes_sem, command = self.U_yes_sem).grid(row = 3, column = 1, sticky = 'w')
        
        self.CheckVar_U_no_sem = tk.IntVar()
        self.CheckVar_U_no_sem.set(0)
        self.option_u_no_sem = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_no_sem, command = self.U_no_sem).grid(row = 3, column = 1, sticky = 'e')
        
    def Th_no_sem(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        self.Th_checkbutton = tk.Button(dialog_frame, text = "Upload Th file", font = ('TkDefaultFont', 10), command = self.file_Th_upload).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_Th_upload = tk.IntVar()
        self.CheckVar_Th_upload.set(0)
        self.Th_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_upload).grid(row = 0, column = 1, sticky = 'w')
        
        self.Th_wash_checkbutton = tk.Button(dialog_frame, text = "Upload Th wash file", font = ('TkDefaultFont', 10), command = self.file_Thwash_upload).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_Thwash_upload = tk.IntVar()
        self.CheckVar_Thwash_upload.set(0)
        self.Thwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_Thwash_upload).grid(row = 1, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "Would you like to alter your U file? ", font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'w')
        
        self.CheckVar_U_yes_sem = tk.IntVar()
        self.CheckVar_U_yes_sem.set(0)
        self.option_u_yes_sem = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_yes_sem, command = self.U_yes_sem).grid(row = 2, column = 1, sticky = 'w')
        
        self.CheckVar_U_no_sem = tk.IntVar()
        self.CheckVar_U_no_sem.set(0)
        self.option_u_no_sem = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_no_sem, command = self.U_no_sem).grid(row = 2, column = 2, sticky = 'w')
       
    def U_yes_sem(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = 'Enter last cycle # to be analyzed: ' , font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput_U = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.rowinput_U.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_U.focus_set()
        
        self.U_checkbutton = tk.Button(dialog_frame, text = "Upload U file", font = ('TkDefaultFont', 10), command = self.file_U_upload_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_U_upload = tk.IntVar()
        self.CheckVar_U_upload.set(0)
        self.U_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_upload).grid(row = 1, column = 1, sticky = 'w')
        
        self.U_wash_checkbutton = tk.Button(dialog_frame, text = "Upload U wash file", font = ('TkDefaultFont', 10), command = self.file_Uwash_upload).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_Uwash_upload = tk.IntVar()
        self.CheckVar_Uwash_upload.set(0)
        self.Uwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_upload).grid(row = 2, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = 'Additional uploads:', font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'e')
        
        self.chemblank_checkbutton = tk.Button(dialog_frame, text = "Upload chemblank excel file", font = ('TkDefaultFont', 10), command = self.file_chemblank_upload).grid(row = 4, column = 0, sticky = 'e')
        
        self.CheckVar_chemblank_upload = tk.IntVar()
        self.CheckVar_chemblank_upload.set(0)
        self.chemblank_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_chemblank_upload).grid(row = 4, column = 1, sticky = 'w')
        
        self.file_checkbutton = tk.Button(dialog_frame, text = "Upload age excel file", font = ('TkDefaultFont', 10), command = self.file_upload_export).grid(row = 5, column = 0, sticky = 'e')
        
        self.CheckVar_file_export = tk.IntVar()
        self.CheckVar_file_export.set(0)
        self.file_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_file_export).grid(row = 5, column = 1, sticky = 'w')
        
        self.age = tk.Button(dialog_frame, text = 'Calculate Age', font = ('TkDefaultFont', 10), default = 'active', command = self.sem_command).grid(row = 6, column = 0)
        
        self.quit = tk.Button(dialog_frame, text="QUIT", command= root.destroy).grid(row = 6, column = 1)
       
    def U_no_sem(self):     
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        self.U_checkbutton = tk.Button(dialog_frame, text = "Upload U file", font = ('TkDefaultFont', 10), command = self.file_U_upload).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_U_upload = tk.IntVar()
        self.CheckVar_U_upload.set(0)
        self.U_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_upload).grid(row = 0, column = 1, sticky = 'w')
        
        self.U_wash_checkbutton = tk.Button(dialog_frame, text = "Upload U wash file", font = ('TkDefaultFont', 10), command = self.file_Uwash_upload).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_Uwash_upload = tk.IntVar()
        self.CheckVar_Uwash_upload.set(0)
        self.Uwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_upload).grid(row = 1, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = 'Additional uploads:', font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'e')
        
        self.chemblank_checkbutton = tk.Button(dialog_frame, text = "Upload chemblank excel file", font = ('TkDefaultFont', 10), command = self.file_chemblank_upload).grid(row = 3, column = 0, sticky = 'e')
        
        self.CheckVar_chemblank_upload = tk.IntVar()
        self.CheckVar_chemblank_upload.set(0)
        self.chemblank_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_chemblank_upload).grid(row = 3, column = 1, sticky = 'w')
        
        self.file_checkbutton = tk.Button(dialog_frame, text = "Upload age excel file", font = ('TkDefaultFont', 10), command = self.file_upload_export).grid(row = 4, column = 0, sticky = 'e')
        
        self.CheckVar_file_export = tk.IntVar()
        self.CheckVar_file_export.set(0)
        self.file_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_file_export).grid(row = 4, column = 1, sticky = 'w')
        
        self.age = tk.Button(dialog_frame, text = 'Calculate Age', default = 'active', font = ('TkDefaultFont', 10), command = self.sem_command).grid(row = 5, column = 0)
        
        self.quit = tk.Button(dialog_frame, text="QUIT", font = ('TkDefaultFont', 10), command= self.quit_program).grid(row = 5, column = 1)  
    
    
    def sem_command(self):
        
        spike_conc_three = preset_values[0]
        spike_conc_nine = preset_values[1]
        sample_wt_err = preset_values[2]
        spike_wt_err = preset_values[3]
        zerotwo = preset_values[4]
        zerotwo_err = preset_values[5]
        
        AS_U = self.AS_U_input.get()
        
        if self.AS_Th_no == 1:
            AS_Th = self.AS_Th_input.get()
        else: 
            AS_Th = AS_U
        
        sem_calc = Application_sem(self.spikeinput.get(), AS_U, AS_Th, self.sample_wt.get(), self.spike_wt.get(), self.sample_ID.get(), self.row_age.get(), 
                        self.filename_Th, self.filename_Thwash, self.filename_U, self.filename_Uwash, 
                        self.chem_lst, self.file_export, 
                        sample_wt_err, spike_wt_err, spike_conc_three, spike_conc_nine, zerotwo, zerotwo_err)
        sem_calc.age_calculate_sem()
        
    """
    CUPS
    """
        
    def cups_command_U(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = "Check which method used for Th run: ",  font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_sem_Th = tk.IntVar()
        self.CheckVar_sem_Th.set(0)
        self.option_sem_Th = tk.Checkbutton(dialog_frame, text = 'SEM',  font = ('TkDefaultFont', 10), variable = self.CheckVar_sem_Th, command = self.sem_command_Th).grid(row = 0, column = 1, sticky = 'w')
        
        self.CheckVar_cups_Th = tk.IntVar()
        self.CheckVar_cups_Th.set(0)
        self.option_cups_Th = tk.Checkbutton(dialog_frame, text = 'CUPS',  font = ('TkDefaultFont', 10), variable = self.CheckVar_cups_Th, command = self.cups_command_Th).grid(row = 0, column = 2, sticky = 'w')
    
    """
    Th on SEM
    """
    def sem_command_Th(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = "Enter spike information (choose from: DIII-B, DIII-A, 1I, 1H):  ", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        self.spikeinput = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spikeinput.grid(row = 0, column = 1, sticky = 'w')
        self.spikeinput.focus_set()
        
        tk.Label(dialog_frame, text = "Enter abundant sensitivity for 237U-238U for Th run: ", font = ('TkDefaultFont', 10)).grid(row = 1, column = 0, sticky = 'w')
        self.AS_input_Th = tk.Entry(dialog_frame, background = 'white', width= 12)
        self.AS_input_Th.grid(row = 1, column = 1, sticky = 'w')
        self.AS_input_Th.focus_set()
        
        tk.Label(dialog_frame, text = "Enter sample weight (g): ", font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'w')
        self.sample_wt = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.sample_wt.grid(row = 2, column = 1, sticky = 'w')
        self.sample_wt.focus_set()
        
        tk.Label(dialog_frame, text = "Enter spike weight (g): ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'w')
        self.spike_wt = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spike_wt.grid(row = 3, column = 1, sticky = 'w')
        self.spike_wt.focus_set()
        
        tk.Label(dialog_frame, text = "Enter sample ID: ", font = ('TkDefaultFont', 10) ).grid(row = 4, column = 0, sticky = 'w')
        self.sample_ID = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.sample_ID.grid(row = 4, column = 1, sticky = 'w')
        self.sample_ID.focus_set()
        
        tk.Label(dialog_frame, text = "Enter row # for age spreadsheet (starting at 6): ", font = ('TkDefaultFont', 10)).grid(row = 5, column = 0, sticky = 'w')
        self.row_age = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.row_age.grid(row = 5, column = 1, sticky = 'w')
        self.row_age.focus_set()
        
        tk.Label(dialog_frame, text = "Would you like to alter your unspiked standard file? ", font = ('TkDefaultFont', 10)).grid(row = 6, column = 0, sticky = 'e')
        
        self.CheckVar_unspiked_yes_semcups = tk.IntVar()
        self.CheckVar_unspiked_yes_semcups.set(0)
        self.option_unspiked_yes_semcups = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_unspiked_yes_semcups, command = self.unspiked_yes_semcups).grid(row = 6, column = 1, sticky = 'w')
        
        self.CheckVar_unspiked_no_semcups = tk.IntVar()
        self.CheckVar_unspiked_no_semcups.set(0)
        self.option_unspiked_no_semcups = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_unspiked_no_semcups, command = self.unspiked_no_semcups).grid(row = 6, column = 1, sticky = 'e')
   
    def unspiked_yes_semcups(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput_unspiked = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.rowinput_unspiked.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_unspiked.focus_set()
        
        self.unspiked_checkbutton = tk.Button(dialog_frame, text = "Upload unspiked standard file", font = ('TkDefaultFont', 10), command = self.file_unspiked_upload_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_unspiked_upload = tk.IntVar()
        self.CheckVar_unspiked_upload.set(0)
        self.unspiked_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_unspiked_upload).grid(row = 1, column = 1, sticky = 'w')
        
        self.unspikedwash_checkbutton = tk.Button(dialog_frame, text = "Upload unspiked standard wash file", font = ('TkDefaultFont', 10), command = self.file_unspikedwash_upload).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_unspikedwash_upload = tk.IntVar()
        self.CheckVar_unspikedwash_upload.set(0)
        self.unspikedwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_unspikedwash_upload).grid(row = 2, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "Enter your spiked standard 234/238 value in ppm: ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'e')
        self.spiked = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spiked.grid(row =3, column = 1, sticky = 'w')
        self.spiked.focus_set()
        
        tk.Label(dialog_frame, text = "Enter your spiked standard 234/238 error value in ppm: ", font = ('TkDefaultFont', 10)).grid(row = 4, column = 0, sticky = 'e')
        self.spiked_err = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spiked_err.grid(row = 4, column = 1, sticky = 'w')
        self.spiked_err.focus_set()
        
        tk.Label(dialog_frame, text = "Would you like to alter your Th file? ", font = ('TkDefaultFont', 10)).grid(row = 5, column = 0, sticky = 'e')
        
        self.CheckVar_Th_yes_semcups = tk.IntVar()
        self.CheckVar_Th_yes_semcups.set(0)
        self.option_th_yes_semcups = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_yes_semcups, command = self.Th_yes_semcups).grid(row = 5, column = 1, sticky = 'w')
        
        self.CheckVar_Th_no_semcups = tk.IntVar()
        self.CheckVar_Th_no_semcups.set(0)
        self.option_Th_no_semcups = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_no_semcups, command = self.Th_no_semcups).grid(row = 5, column = 1, sticky = 'e')
   
    def unspiked_no_semcups(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        self.unspiked_checkbutton = tk.Button(dialog_frame, text = "Upload unspiked standard file", font = ('TkDefaultFont', 10), command = self.file_unspiked_upload).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_unspiked_upload = tk.IntVar()
        self.CheckVar_unspiked_upload.set(0)
        self.unspiked_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10),  variable = self.CheckVar_unspiked_upload).grid(row = 0, column = 1, sticky = 'w')
        
        self.unspikedwash_checkbutton = tk.Button(dialog_frame, text = "Upload unspiked standard wash file", font = ('TkDefaultFont', 10), command = self.file_unspikedwash_upload).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_unspikedwash_upload = tk.IntVar()
        self.CheckVar_unspikedwash_upload.set(0)
        self.unspikedwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_unspikedwash_upload).grid(row = 1, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "Enter your spiked standard 234/238 value in ppm: ", font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'e')
        self.spiked = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spiked.grid(row = 2, column = 1, sticky = 'w')
        self.spiked.focus_set()
        
        tk.Label(dialog_frame, text = "Enter your spiked standard 234/238 error value in ppm: ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'e')
        self.spiked_err = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spiked_err.grid(row = 3, column = 1, sticky = 'w')
        self.spiked_err.focus_set()
        
        tk.Label(dialog_frame, text = "Would you like to alter your Th file? ", font = ('TkDefaultFont', 10)).grid(row = 4, column = 0, sticky = 'e')
        
        self.CheckVar_Th_yes_semcups = tk.IntVar()
        self.CheckVar_Th_yes_semcups.set(0)
        self.option_th_yes_semcups = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_yes_semcups, command = self.Th_yes_semcups).grid(row = 4, column = 1, sticky = 'w')
        
        self.CheckVar_Th_no_semcups = tk.IntVar()
        self.CheckVar_Th_no_semcups.set(0)
        self.option_Th_no_semcups = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_no_semcups, command = self.Th_no_semcups).grid(row = 4, column = 1, sticky = 'e')
   
    def Th_yes_semcups(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput_Th = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.rowinput_Th.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_Th.focus_set()
        
        self.Th_checkbutton = tk.Button(dialog_frame, text = "Upload Th file", font = ('TkDefaultFont', 10), command = self.file_Th_upload_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_Th_upload = tk.IntVar()
        self.CheckVar_Th_upload.set(0)
        self.Th_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_upload).grid(row = 1, column = 1, sticky = 'w')
        
        self.Thwash_checkbutton = tk.Button(dialog_frame, text = "Upload Th wash file", font = ('TkDefaultFont', 10), command = self.file_Thwash_upload).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_Thwash_upload = tk.IntVar()
        self.CheckVar_Thwash_upload.set(0)
        self.Thwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_Thwash_upload).grid(row = 2, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "Would you like to alter your U file? ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'w')
        
        self.CheckVar_U_yes_semcups = tk.IntVar()
        self.CheckVar_U_yes_semcups.set(0)
        self.option_U_yes_semcups = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_yes_semcups, command = self.U_yes_semcups).grid(row = 3, column = 1, sticky = 'w')
        
        self.CheckVar_U_no_semcups = tk.IntVar()
        self.CheckVar_U_no_semcups.set(0)
        self.option_U_no_semcups = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_no_semcups, command = self.U_no_semcups).grid(row = 3, column = 1, sticky = 'e')
    
    def Th_no_semcups(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        self.Th_checkbutton = tk.Button(dialog_frame, text = "Upload Th file", font = ('TkDefaultFont', 10), command = self.file_Th_upload).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_Th_upload = tk.IntVar()
        self.CheckVar_Th_upload.set(0)
        self.Th_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_upload).grid(row = 0, column = 1, sticky = 'w')
        
        self.Thwash_checkbutton = tk.Button(dialog_frame, text = "Upload Th wash file", font = ('TkDefaultFont', 10), command = self.file_Thwash_upload).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_Thwash_upload = tk.IntVar()
        self.CheckVar_Thwash_upload.set(0)
        self.Thwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_Thwash_upload).grid(row = 1, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "Would you like to alter your U file? ", font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'w')
        
        self.CheckVar_U_yes_semcups = tk.IntVar()
        self.CheckVar_U_yes_semcups.set(0)
        self.option_U_yes_semcups = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_yes_semcups, command = self.U_yes_semcups).grid(row = 2, column = 1, sticky = 'w')
        
        self.CheckVar_U_no_semcups = tk.IntVar()
        self.CheckVar_U_no_semcups.set(0)
        self.option_U_no_semcups = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_no_semcups, command = self.U_no_semcups).grid(row = 2, column = 2, sticky = 'w')
        
    def U_yes_semcups(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        
        tk.Label(dialog_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput_U = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.rowinput_U.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_U.focus_set()
        
        self.U_checkbutton = tk.Button(dialog_frame, text = "Upload U file", font = ('TkDefaultFont', 10), command = self.file_U_upload_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_U_upload = tk.IntVar()
        self.CheckVar_U_upload.set(0)
        self.U_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_U_upload).grid(row = 1, column = 1, sticky = 'w')
        
        self.Uwash_checkbutton = tk.Button(dialog_frame, text = "Upload U wash file", font = ('TkDefaultFont', 10), command = self.file_Uwash_upload).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_Uwash_upload = tk.IntVar()
        self.CheckVar_Uwash_upload.set(0)
        self.Uwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_upload).grid(row = 2, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "U wash run on: ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'w')
        
        self.CheckVar_Uwash_sem = tk.IntVar()
        self.CheckVar_Uwash_sem.set(0)
        self.Uwash_sem_checkbutton= tk.Checkbutton(dialog_frame, text = "SEM", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_sem, command = self.Uwash_sem).grid(row = 3, column = 1, sticky = 'w')
        
        self.CheckVar_Uwash_cups = tk.IntVar()
        self.CheckVar_Uwash_cups.set(0)
        self.Uwash_cups_checkbutton= tk.Checkbutton(dialog_frame, text = "CUPS", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_cups, command = self.Uwash_cups).grid(row = 3, column = 2, sticky = 'w')
        
        
        tk.Label(dialog_frame, text = 'Additional uploads:', font = ('TkDefaultFont', 10)).grid(row = 4, column = 0, sticky = 'e')
        
        self.chemblank_checkbutton = tk.Button(dialog_frame, text = "Upload chemblank excel file", font = ('TkDefaultFont', 10), command = self.file_chemblank_upload).grid(row = 5, column = 0, sticky = 'e')
        
        self.CheckVar_chemblank_upload = tk.IntVar()
        self.CheckVar_chemblank_upload.set(0)
        self.chemblank_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_chemblank_upload).grid(row = 5, column = 1, sticky = 'w')
        
        self.file_checkbutton = tk.Button(dialog_frame, text = "Upload age excel file", font = ('TkDefaultFont', 10), command = self.file_upload_export).grid(row = 6, column = 0, sticky = 'e')
        
        self.CheckVar_file_export = tk.IntVar()
        self.CheckVar_file_export.set(0)
        self.file_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_file_export).grid(row = 6, column = 1, sticky = 'w')
        
        self.age = tk.Button(dialog_frame, text = 'Calculate Age', font = ('TkDefaultFont', 10), default = 'active', command = self.semcups_command).grid(row = 7, column = 0)
        
        self.quit = tk.Button(dialog_frame, text="QUIT", font = ('TkDefaultFont', 10), command= self.quit_program).grid(row = 7, column = 1)  
        
    def U_no_semcups(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()

        self.U_checkbutton = tk.Button(dialog_frame, text = "Upload U file", font = ('TkDefaultFont', 10), command = self.file_U_upload).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_U_upload = tk.IntVar()
        self.CheckVar_U_upload.set(0)
        self.U_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_U_upload).grid(row = 0, column = 1, sticky = 'w')
        
        self.Uwash_checkbutton = tk.Button(dialog_frame, text = "Upload U wash file", font = ('TkDefaultFont', 10), command = self.file_Uwash_upload).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_Uwash_upload = tk.IntVar()
        self.CheckVar_Uwash_upload.set(0)
        self.Uwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_upload).grid(row = 1, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "U wash run on: ", font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'w')
        
        self.CheckVar_Uwash_sem = tk.IntVar()
        self.CheckVar_Uwash_sem.set(0)
        self.Uwash_sem_checkbutton= tk.Checkbutton(dialog_frame, text = "SEM", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_sem, command = self.Uwash_sem).grid(row = 2, column = 1, sticky = 'w')
        
        self.CheckVar_Uwash_cups = tk.IntVar()
        self.CheckVar_Uwash_cups.set(0)
        self.Uwash_cups_checkbutton= tk.Checkbutton(dialog_frame, text = "CUPS", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_cups, command = self.Uwash_cups).grid(row = 2, column = 2, sticky = 'w')
        
        tk.Label(dialog_frame, text = 'Additional uploads:', font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'e')
        
        self.chemblank_checkbutton = tk.Button(dialog_frame, text = "Upload chemblank excel file", font = ('TkDefaultFont', 10), command = self.file_chemblank_upload).grid(row = 4, column = 0, sticky = 'e')
        
        self.CheckVar_chemblank_upload = tk.IntVar()
        self.CheckVar_chemblank_upload.set(0)
        self.chemblank_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_chemblank_upload).grid(row = 4, column = 1, sticky = 'w')
        
        self.file_checkbutton = tk.Button(dialog_frame, text = "Upload age excel file", font = ('TkDefaultFont', 10), command = self.file_upload_export).grid(row = 5, column = 0, sticky = 'e')
        
        self.CheckVar_file_export = tk.IntVar()
        self.CheckVar_file_export.set(0)
        self.file_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_file_export).grid(row = 5, column = 1, sticky = 'w')
        
        self.age = tk.Button(dialog_frame, text = 'Calculate Age', font = ('TkDefaultFont', 10), default = 'active', command = self.semcups_command).grid(row = 6, column = 0)
        
        self.quit = tk.Button(dialog_frame, text="QUIT", font = ('TkDefaultFont', 10), command= self.quit_program).grid(row = 6, column = 2)

    def semcups_command(self):
        
        spike_conc_three = preset_values[0]
        spike_conc_nine = preset_values[1]
        sample_wt_err = preset_values[2]
        spike_wt_err = preset_values[3]
        zerotwo = preset_values[4]
        zerotwo_err = preset_values[5]
        
        semcups_calc = Application_semcups(self.spikeinput.get(), self.AS_input_Th.get(), self.sample_wt.get(), self.spike_wt.get(), self.sample_ID.get(), self.row_age.get(),
                                            self.spiked.get(), self.spiked_err.get(), self.filename_unspiked, self.filename_unspikedwash, 
                                            self.filename_Th, self.filename_Thwash, self.filename_U, self.filename_Uwash, 
                                            self.chem_lst, self.file_export,
                                            sample_wt_err, spike_wt_err, spike_conc_three, spike_conc_nine, zerotwo, zerotwo_err,
                                            self.Uwash)
        
        semcups_calc.age_calculate_semcups()
        
    """
    Th on Cups
    """    
    def cups_command_Th(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = "Enter spike information (choose from: DIII-B, DIII-A, 1I, 1H):  ", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        self.spikeinput = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spikeinput.grid(row = 0, column = 1, sticky = 'w')
        self.spikeinput.focus_set()
        
        tk.Label(dialog_frame, text = "Enter sample weight (g): ", font = ('TkDefaultFont', 10)).grid(row = 1, column = 0, sticky = 'w')
        self.sample_wt = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.sample_wt.grid(row = 1, column = 1, sticky = 'w')
        self.sample_wt.focus_set()
        
        tk.Label(dialog_frame, text = "Enter spike weight (g): ", font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'w')
        self.spike_wt = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spike_wt.grid(row = 2, column = 1, sticky = 'w')
        self.spike_wt.focus_set()
        
        tk.Label(dialog_frame, text = "Enter sample ID: ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'w')
        self.sample_ID = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.sample_ID.grid(row = 3, column = 1, sticky = 'w')
        self.sample_ID.focus_set()
        
        tk.Label(dialog_frame, text = "Enter row # for age spreadsheet (starting at 6): ", font = ('TkDefaultFont', 10)).grid(row = 4, column = 0, sticky = 'w')
        self.row_age = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.row_age.grid(row = 4, column = 1, sticky = 'w')
        self.row_age.focus_set()
        
        tk.Label(dialog_frame, text = "Would you like to alter your unspiked standard file? ", font = ('TkDefaultFont', 10)).grid(row = 5, column = 0, sticky = 'e')
        
        self.CheckVar_unspiked_yes_cups = tk.IntVar()
        self.CheckVar_unspiked_yes_cups.set(0)
        self.option_unspiked_yes_cups = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_unspiked_yes_cups, command = self.unspiked_yes_cups).grid(row = 5, column = 1, sticky = 'w')
        
        self.CheckVar_unspiked_no_cups = tk.IntVar()
        self.CheckVar_unspiked_no_cups.set(0)
        self.option_unspiked_no_cups = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_unspiked_no_cups, command = self.unspiked_no_cups).grid(row = 5, column = 1, sticky = 'e')
    

    def unspiked_yes_cups(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput_unspiked = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.rowinput_unspiked.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_unspiked.focus_set()
        
        self.unspiked_checkbutton = tk.Button(dialog_frame, text = "Upload unspiked standard file", font = ('TkDefaultFont', 10), command = self.file_unspiked_upload_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_unspiked_upload = tk.IntVar()
        self.CheckVar_unspiked_upload.set(0)
        self.unspiked_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_unspiked_upload).grid(row = 1, column = 1, sticky = 'w')
        
        self.unspikedwash_checkbutton = tk.Button(dialog_frame, text = "Upload unspiked standard wash file", font = ('TkDefaultFont', 10), command = self.file_unspikedwash_upload).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_unspikedwash_upload = tk.IntVar()
        self.CheckVar_unspikedwash_upload.set(0)
        self.unspikedwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_unspikedwash_upload).grid(row = 2, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "Enter your spiked standard 234/238 value in ppm: ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'e')
        self.spiked = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spiked.grid(row = 3, column = 1, sticky = 'w')
        self.spiked.focus_set()
        
        tk.Label(dialog_frame, text = "Enter your spiked standard 234/238 error value in ppm: ", font = ('TkDefaultFont', 10)).grid(row = 4, column = 0, sticky = 'e')
        self.spiked_err = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spiked_err.grid(row = 4, column = 1, sticky = 'w')
        self.spiked_err.focus_set()
              
        tk.Label(dialog_frame, text = "Would you like to alter your Th file? ", font = ('TkDefaultFont', 10)).grid(row = 5, column = 0, sticky = 'e')
        
        self.CheckVar_Th_yes_cups = tk.IntVar()
        self.CheckVar_Th_yes_cups.set(0)
        self.option_th_yes_cups = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_yes_cups, command = self.Th_yes_cups).grid(row = 5, column = 1, sticky = 'w')
        
        self.CheckVar_Th_no_cups = tk.IntVar()
        self.CheckVar_Th_no_cups.set(0)
        self.option_Th_no_cups = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_no_cups, command = self.Th_no_cups).grid(row = 5, column = 1, sticky = 'e')
   
    def unspiked_no_cups(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        self.unspiked_checkbutton = tk.Button(dialog_frame, text = "Upload unspiked standard file", font = ('TkDefaultFont', 10), command = self.file_unspiked_upload).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_unspiked_upload = tk.IntVar()
        self.CheckVar_unspiked_upload.set(0)
        self.unspiked_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_unspiked_upload).grid(row = 0, column = 1, sticky = 'w')
        
        self.unspikedwash_checkbutton = tk.Button(dialog_frame, text = "Upload unspiked standard wash file", font = ('TkDefaultFont', 10), command = self.file_unspikedwash_upload).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_unspikedwash_upload = tk.IntVar()
        self.CheckVar_unspikedwash_upload.set(0)
        self.unspikedwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_unspikedwash_upload).grid(row = 1, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "Enter your spiked standard 234/238 value in ppm: ", font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'e')
        self.spiked = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spiked.grid(row = 2, column = 1, sticky = 'w')
        self.spiked.focus_set()
        
        tk.Label(dialog_frame, text = "Enter your spiked standard 234/238 error value in ppm: ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'e')
        self.spiked_err = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spiked_err.grid(row = 3, column = 1, sticky = 'w')
        self.spiked_err.focus_set()
        
        tk.Label(dialog_frame, text = "Would you like to alter your Th file? ", font = ('TkDefaultFont', 10)).grid(row = 4, column = 0, sticky = 'e')
        
        self.CheckVar_Th_yes_cups = tk.IntVar()
        self.CheckVar_Th_yes_cups.set(0)
        self.option_th_yes_cups = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_yes_cups, command = self.Th_yes_cups).grid(row = 4, column = 1, sticky = 'w')
        
        self.CheckVar_Th_no_cups = tk.IntVar()
        self.CheckVar_Th_no_cups.set(0)
        self.option_Th_no_cups = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_no_cups, command = self.Th_no_cups).grid(row = 4, column = 1, sticky = 'e')
   
    def Th_yes_cups(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        self.rowinput_Th = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.rowinput_Th.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_Th.focus_set()
        
        self.Th_checkbutton = tk.Button(dialog_frame, text = "Upload Th file", font = ('TkDefaultFont', 10), command = self.file_Th_upload_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_Th_upload = tk.IntVar()
        self.CheckVar_Th_upload.set(0)
        self.Th_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_upload).grid(row = 1, column = 1, sticky = 'w')
        
        self.Thwash_checkbutton = tk.Button(dialog_frame, text = "Upload Th wash file", font = ('TkDefaultFont', 10), command = self.file_Thwash_upload).grid(row = 2, column = 0, sticky = 'e')
        self.CheckVar_Thwash_upload = tk.IntVar()
        self.CheckVar_Thwash_upload.set(0)
        self.Thwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_Thwash_upload).grid(row = 2, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "Th wash run on: ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'e')
        
        self.CheckVar_Thwash_sem = tk.IntVar()
        self.CheckVar_Thwash_sem.set(0)
        self.Thwash_sem_checkbutton= tk.Checkbutton(dialog_frame, text = "SEM", font = ('TkDefaultFont', 10), variable = self.CheckVar_Thwash_sem, command = self.Thwash_sem).grid(row = 3, column = 1, sticky = 'w')
        
        self.CheckVar_Thwash_cups = tk.IntVar()
        self.CheckVar_Thwash_cups.set(0)
        self.Thwash_cups_checkbutton= tk.Checkbutton(dialog_frame, text = "CUPS", font = ('TkDefaultFont', 10), variable = self.CheckVar_Thwash_cups, command = self.Thwash_cups).grid(row = 3, column = 2, sticky = 'w')
        
        tk.Label(dialog_frame, text = "Would you like to alter your U file? ", font = ('TkDefaultFont', 10)).grid(row = 4, column = 0, sticky = 'w')
        
        self.CheckVar_U_yes_cups = tk.IntVar()
        self.CheckVar_U_yes_cups.set(0)
        self.option_U_yes_cups = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_yes_cups, command = self.U_yes_cups).grid(row = 4, column = 1, sticky = 'w')
        
        self.CheckVar_U_no_cups = tk.IntVar()
        self.CheckVar_U_no_cups.set(0)
        self.option_U_no_cups = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_no_cups, command = self.U_no_cups).grid(row = 4, column = 1, sticky = 'e')
        
    def Th_no_cups(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()

        self.Th_checkbutton = tk.Button(dialog_frame, text = "Upload Th file", font = ('TkDefaultFont', 10), command = self.file_Th_upload).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_Th_upload = tk.IntVar()
        self.CheckVar_Th_upload.set(0)
        self.Th_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_Th_upload).grid(row = 0, column = 1, sticky = 'w')
        
        self.Thwash_checkbutton = tk.Button(dialog_frame, text = "Upload Th wash file", font = ('TkDefaultFont', 10), command = self.file_Thwash_upload).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_Thwash_upload = tk.IntVar()
        self.CheckVar_Thwash_upload.set(0)
        self.Thwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_Thwash_upload).grid(row = 1, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "Th wash run on: ", font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_Thwash_sem = tk.IntVar()
        self.CheckVar_Thwash_sem.set(0)
        self.Thwash_sem_checkbutton= tk.Checkbutton(dialog_frame, text = "SEM", font = ('TkDefaultFont', 10), variable = self.CheckVar_Thwash_sem, command = self.Thwash_sem).grid(row = 2, column = 1, sticky = 'w')
        
        self.CheckVar_Thwash_cups = tk.IntVar()
        self.CheckVar_Thwash_cups.set(0)
        self.Thwash_cups_checkbutton= tk.Checkbutton(dialog_frame, text = "CUPS", font = ('TkDefaultFont', 10), variable = self.CheckVar_Thwash_cups, command = self.Thwash_cups).grid(row = 2, column = 2, sticky = 'w')
        
        tk.Label(dialog_frame, text = "Would you like to alter your U file? ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'w')
        
        self.CheckVar_U_yes_cups = tk.IntVar()
        self.CheckVar_U_yes_cups.set(0)
        self.option_U_yes_cups = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_yes_cups, command = self.U_yes_cups).grid(row = 3, column = 1, sticky = 'w')
        
        self.CheckVar_U_no_cups = tk.IntVar()
        self.CheckVar_U_no_cups.set(0)
        self.option_U_no_cups = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_U_no_cups, command = self.U_no_cups).grid(row = 3, column = 2, sticky = 'w')
        
    def U_yes_cups(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput_U = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.rowinput_U.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_U.focus_set()
        
        self.U_checkbutton = tk.Button(dialog_frame, text = "Upload U file", font = ('TkDefaultFont', 10), command = self.file_U_upload_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_U_upload = tk.IntVar()
        self.CheckVar_U_upload.set(0)
        self.U_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_U_upload).grid(row = 1, column = 1, sticky = 'w')
        
        self.Uwash_checkbutton = tk.Button(dialog_frame, text = "Upload U wash file", font = ('TkDefaultFont', 10), command = self.file_Uwash_upload).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_Uwash_upload = tk.IntVar()
        self.CheckVar_Uwash_upload.set(0)
        self.Uwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_upload).grid(row = 2, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "U wash run on: ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'e')
        
        self.CheckVar_Uwash_sem = tk.IntVar()
        self.CheckVar_Uwash_sem.set(0)
        self.Uwash_sem_checkbutton= tk.Checkbutton(dialog_frame, text = "SEM", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_sem, command = self.Uwash_sem).grid(row = 3, column = 1, sticky = 'w')
        
        self.CheckVar_Uwash_cups = tk.IntVar()
        self.CheckVar_Uwash_cups.set(0)
        self.Uwash_cups_checkbutton= tk.Checkbutton(dialog_frame, text = "CUPS", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_cups, command = self.Uwash_cups).grid(row = 3, column = 2, sticky = 'w')
        
        tk.Label(dialog_frame, text = 'Additional uploads:', font = ('TkDefaultFont', 10)).grid(row = 4, column = 0, sticky = 'w')
        
        self.chemblank_checkbutton = tk.Button(dialog_frame, text = "Upload chemblank excel file", font = ('TkDefaultFont', 10), command = self.file_chemblank_upload).grid(row = 5, column = 0, sticky = 'e')
        
        self.CheckVar_chemblank_upload = tk.IntVar()
        self.CheckVar_chemblank_upload.set(0)
        self.chemblank_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_chemblank_upload).grid(row = 5, column = 1, sticky = 'w')
        
        self.file_checkbutton = tk.Button(dialog_frame, text = "Upload age excel file", font = ('TkDefaultFont', 10), command = self.file_upload_export).grid(row = 6, column = 0, sticky = 'e')
        
        self.CheckVar_file_export = tk.IntVar()
        self.CheckVar_file_export.set(0)
        self.file_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_file_export).grid(row = 6, column = 1, sticky = 'w')
        
        self.age = tk.Button(dialog_frame, text = 'Calculate Age', default = 'active', font = ('TkDefaultFont', 10), command = self.cups_command).grid(row = 7, column = 0)
        
        self.quit = tk.Button(dialog_frame, text="QUIT", font = ('TkDefaultFont', 10), command= self.quit_program).grid(row = 7, column = 1)  
           
    def U_no_cups(self):
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()

        self.U_checkbutton = tk.Button(dialog_frame, text = "Upload U file", font = ('TkDefaultFont', 10), command = self.file_U_upload).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_U_upload = tk.IntVar()
        self.CheckVar_U_upload.set(0)
        self.U_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_U_upload).grid(row = 0, column = 1, sticky = 'w')
        
        self.Uwash_checkbutton = tk.Button(dialog_frame, text = "Upload U wash file", font = ('TkDefaultFont', 10), command = self.file_Uwash_upload).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_Uwash_upload = tk.IntVar()
        self.CheckVar_Uwash_upload.set(0)
        self.Uwash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = "Uploaded", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_upload).grid(row = 1, column = 1, sticky = 'w')
        
        tk.Label(dialog_frame, text = "U wash run on: ", font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_Uwash_sem = tk.IntVar()
        self.CheckVar_Uwash_sem.set(0)
        self.Uwash_sem_checkbutton= tk.Checkbutton(dialog_frame, text = "SEM", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_sem, command = self.Uwash_sem).grid(row = 2, column = 1, sticky = 'w')
        
        self.CheckVar_Uwash_cups = tk.IntVar()
        self.CheckVar_Uwash_cups.set(0)
        self.Uwash_cups_checkbutton= tk.Checkbutton(dialog_frame, text = "CUPS", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_cups, command = self.Uwash_cups).grid(row = 2, column = 2, sticky = 'w')
        
        tk.Label(dialog_frame, text = 'Additional uploads:', font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'e')
        
        self.chemblank_checkbutton = tk.Button(dialog_frame, text = "Upload chemblank excel file", font = ('TkDefaultFont', 10), command = self.file_chemblank_upload).grid(row = 4, column = 0, sticky = 'e')
        
        self.CheckVar_chemblank_upload = tk.IntVar()
        self.CheckVar_chemblank_upload.set(0)
        self.chemblank_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_chemblank_upload).grid(row = 4, column = 1, sticky = 'w')
        
        self.file_checkbutton = tk.Button(dialog_frame, text = "Upload age excel file", font = ('TkDefaultFont', 10), command = self.file_upload_export).grid(row = 5, column = 0, sticky = 'e')
        
        self.CheckVar_file_export = tk.IntVar()
        self.CheckVar_file_export.set(0)
        self.file_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_file_export).grid(row = 5, column = 1, sticky = 'w')
        
        self.age = tk.Button(dialog_frame, text = 'Calculate Age', font = ('TkDefaultFont', 10), default = 'active', command = self.cups_command).grid(row = 6, column = 0)
        
        self.quit = tk.Button(dialog_frame, text="QUIT", font = ('TkDefaultFont', 10), command= self.quit_program).grid(row = 6, column = 1)

    def cups_command(self):
        
        spike_conc_three = preset_values[0]
        spike_conc_nine = preset_values[1]
        sample_wt_err = preset_values[2]
        spike_wt_err = preset_values[3]
        zerotwo = preset_values[4]
        zerotwo_err = preset_values[5]
        
        cups_calc = Application_cups(self.spikeinput.get(), self.sample_wt.get(), self.spike_wt.get(), self.sample_ID.get(), self.row_age.get(),
                                            self.spiked.get(), self.spiked_err.get(), self.filename_unspiked, self.filename_unspikedwash, 
                                            self.filename_Th, self.filename_Thwash, self.filename_U, self.filename_Uwash, 
                                            self.chem_lst, self.file_export,
                                            sample_wt_err, spike_wt_err, spike_conc_three, spike_conc_nine, zerotwo, zerotwo_err,
                                            self.Uwash, self.Thwash)
        cups_calc.age_calculate_cups()
    
    """
    Wash options for cups measurements
    """
    
    def Uwash_cups(self):
        
        self.Uwash = "cups"
    
    def Uwash_sem(self):
        
        self.Uwash = "sem"
        
    def Thwash_cups(self):
        
        self.Thwash = "cups"
        
    def Thwash_sem(self):
        
        self.Thwash = "sem"
        
    """
    Changing preset options
    """
    
    def preset_change(self):
        
        """
        preset values refer to the following:
            [0]: 233 spike concentration (0.0 will be replaced by spike specific value if not specified)
            [1]: 229 spike concentration (0.0 will be replaced by spike specific value if not specified)
            [2]: sample weight error
            [3]: spike weight error
            [4]: 230/232 initial value
            [5]: 230/232 initial error value
        """
        self.master.withdraw()
        
        Application_preset(self)
        
    def show(self):
        
        self.master.update()
        self.master.deiconify()
        
        self.method_used()
        
    """
    Program quit function
    """
        
    def quit_program(self):
        
        self.master.destroy()
        root.quit()
        
    """
    Upload functions
    """
    
    def file_unspiked_upload_option(self):
        filename_raw = filedialog.askopenfilename(parent = self)
        try: 
            filename_unspiked = openpyxl.Workbook()
            ws = filename_unspiked.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in islice(reader, 0, int(self.rowinput_unspiked.get()) + 9):
                    ws.append(row)
            filename_unspiked.save("unspiked.xlsx")
            self.filename_unspiked = "unspiked.xlsx"
            self.CheckVar_unspiked_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error: ", str(sys.exc_info()[:]))
    
    def file_unspiked_upload(self):
        filename_raw = filedialog.askopenfilename(parent = self)
        try: 
            filename_unspiked = openpyxl.Workbook()
            ws = filename_unspiked.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_unspiked.save("unspiked.xlsx")
            self.filename_unspiked = "unspiked.xlsx"
            self.CheckVar_unspiked_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_unspiked = filename_raw
            self.CheckVar_unspiked_upload.set(1)
    
    def file_unspikedwash_upload(self):
        filename_raw = filedialog.askopenfilename(parent = self)
        try: 
            filename_unspikedwash = openpyxl.Workbook()
            ws = filename_unspikedwash.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_unspikedwash.save("unspiked_wash.xlsx")
            self.filename_unspikedwash = "unspiked_wash.xlsx"
            self.CheckVar_unspikedwash_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_unspikedwash = filename_raw
            self.CheckVar_unspikedwash_upload.set(1)
    
    def file_Th_upload_option(self):
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_Th = openpyxl.Workbook()
            ws = filename_Th.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in islice(reader, 0, int(self.rowinput_Th.get())+ 9):
                    ws.append(row)
            filename_Th.save("Th.xlsx")
            self.filename_Th = "Th.xlsx"
            self.CheckVar_Th_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))  
    
    def file_Th_upload(self):
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_Th = openpyxl.Workbook()
            ws = filename_Th.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_Th.save("Th.xlsx")
            self.filename_Th = "Th.xlsx"
            self.CheckVar_Th_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_Th = filename_raw
            self.CheckVar_Th_upload.set(1)
        #else:
        #    messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
            
    def file_Thwash_upload(self):
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_Thwash = openpyxl.Workbook()
            ws = filename_Thwash.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_Thwash.save("Thwash.xlsx")
            self.filename_Thwash = "Thwash.xlsx"
            self.CheckVar_Thwash_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_Thwash = filename_raw
            self.CheckVar_Thwash_upload.set(1)
        #else:
        #    messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
    
    def file_U_upload_option(self):
        
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_U = openpyxl.Workbook()
            ws = filename_U.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in islice(reader, 0, int(self.rowinput_U.get())+ 9):
                    ws.append(row)
            filename_U.save("U.xlsx")
            self.filename_U = "U.xlsx"
            self.CheckVar_U_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
    
    def file_U_upload(self):
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_U = openpyxl.Workbook()
            ws = filename_U.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_U.save("U.xlsx")
            self.filename_U = "U.xlsx"
            self.CheckVar_U_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_U = filename_raw
            self.CheckVar_U_upload.set(1)
        #else:
        #    messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
    
    def file_Uwash_upload(self):
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_Uwash = openpyxl.Workbook()
            ws = filename_Uwash.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_Uwash.save("Uwash.xlsx")
            self.filename_Uwash = "Uwash.xlsx"
            self.CheckVar_Uwash_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_Uwash = filename_raw
            self.CheckVar_Uwash_upload.set(1)
        #else:
        #    messagebox.showerror("Unexpected error:", str(sys.exc_info()[:])) 
     
    def file_chemblank_upload(self):
        filename_raw = filedialog.askopenfilename(parent = self)
        try:
            filename_chemblank = openpyxl.load_workbook(filename_raw)
            ws = filename_chemblank.worksheets[0]
            zero = ws['D4'].value
            zero_err = ws['D5'].value
            two = ws['E4'].value
            two_err = ws['E5'].value
            four = ws['G4'].value
            four_err = ws['G5'].value
            five = ws['H4'].value
            five_err = ws['H5'].value
            eight = ws['I4'].value
            eight_err = ws['I5'].value
            self.chem_lst = [zero, zero_err, two, two_err, four, four_err, five, five_err, eight, eight_err]
            self.CheckVar_chemblank_upload.set(1)
            return self.chem_lst
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:])) 
    
    def file_upload_export(self):
        
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            self.file_export =  filename_raw
            self.CheckVar_file_export.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
        
        
class Application_sem():
    
    def __init__(self, spike, AS_U, AS_Th, sample_wt, spike_wt, sample_ID, row_age, Th, Thwash, U, Uwash, chem_lst, export_age, sample_wt_err, spike_wt_err, spike_conc_three, spike_conc_nine, zerotwo, zerotwo_err):
        
        spike = str(spike)

        #Spike values
        spike_six_three_dictionary = {"DIII-B":1.008398,"DIII-A": 1.008398,"1I":1.010128,"1H":1.010128}
        spike_six_three_err_dictionary = {"DIII-B": 0.00015, "DIII-A": 0.00015, "1I": 0.00015, "1H": 0.00015}
        spike_three_dictionary = {"DIII-B": 0.78938, "DIII-A": 0.78933, "1I": 0.61351, "1H": 0.78997}
        spike_three_err_dictionary = {"DIII-B": 0.00002, "DIII-A": 0.00002, "1I": 0.00002, "1H": 0.00002}
        spike_nine_dictionary = {"DIII-B": 0.21734, "DIII-A": 0.21705, "1I": 0.177187, "1H": 0.22815}
        spike_nine_err_dictionary = {"DIII-B": 0.00001, "DIII-A": 0.00001, "1I": 0.00001, "1H": 0.00001}
        spike_zero_nine_dictionary = {"DIII-B": 0.0000625, "DIII-A": 0.0000625, "1I": 0.0000402, "1H": 0.0000402}
        spike_zero_nine_err_dictionary = {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.0000011, "1H": 0.0000011}
        spike_nine_two_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.00, "1H": 0.00}
        spike_nine_two_err_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.00, "1H": 0.00}
        spike_four_three_dictionary = {"DIII-B": 0.003195, "DIII-A": 0.003195, "1I":0.003180, "1H": 0.003180}
        spike_four_three_err_dictionary= {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.000003, "1H": 0.000003}
        spike_five_three_dictionary = {"DIII-B": 0.105321, "DIII-A": 0.10532, "1I": 0.10521, "1H":0.10521}
        spike_five_three_err_dictionary = {"DIII-B": 0.00003, "DIII-A": 0.00003, "1I": 0.00003, "1H": 0.00003}
        spike_eight_three_dictionary = {"DIII-B": 0.016802, "DIII-A": 0.01680, "1I": 0.01700, "1H":0.01700 }
        spike_eight_three_err_dictionary = {"DIII-B": 0.00001, "DIII-A": 0.00001,"1I": 0.00001, "1H": 0.00001}
        spike_three_nine_dictionary = {"DIII-B": 0.27533001, "DIII-A": 0.27498005, "1I": 0.28880866, "1H": 0.28880844}

        if spike in spike_six_three_dictionary:
            self.spike = float(spike_six_three_dictionary[spike]) #spike ratio
        else: 
            messagebox.showwarning("Error!", "No valid spike info entered! ")
        
        if spike in spike_six_three_err_dictionary: 
            self.spike_six_three_err = float(spike_six_three_err_dictionary[spike]) #error of spike ratio
        
        if spike_conc_three == 0.0:
        
            if spike in spike_three_dictionary:
                self.spike_three = float(spike_three_dictionary[spike]) #in pmol/g
            else:pass
        
        else: 
            self.spike_three = float(spike_conc_three)
        
        if spike_conc_nine == 1.0:
            
            if spike in spike_three_nine_dictionary:
                self.spike_nine = float(spike_three_nine_dictionary[spike]) * self.spike_three
            else: pass
        
        else:
            if spike in spike_nine_dictionary: 
                self.spike_nine = float(spike_nine_dictionary[spike])
            else: pass
    
        if spike in spike_three_err_dictionary:
            self.spike_three_err = float(spike_three_err_dictionary[spike]) #in pmol/g
        else:pass
    
        if spike in spike_nine_err_dictionary: 
            self.spike_nine_err = float(spike_nine_err_dictionary[spike]) #in pmol/g
        else: pass
    
        if spike in spike_zero_nine_dictionary:
            self.spike_zero_nine = float(spike_zero_nine_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_zero_nine_err_dictionary:
            self.spike_zero_nine_err = float(spike_zero_nine_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_nine_two_dictionary: 
            self.spike_nine_two = float(spike_nine_two_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_nine_two_err_dictionary:
            self.spike_nine_two_err = float(spike_nine_two_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_four_three_dictionary:
            self.spike_four_three = float(spike_four_three_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_four_three_err_dictionary:
            self.spike_four_three_err = float(spike_four_three_err_dictionary[spike]) #error of spike ratio
        else: pass
            
        if spike in spike_five_three_dictionary:
            self.spike_five_three = float(spike_five_three_dictionary[spike]) #spike ratio
        else: pass
        
        if spike in spike_five_three_err_dictionary:
            self.spike_five_three_err = float(spike_five_three_err_dictionary[spike]) #error of spike ratio
        else: pass
        
        if spike in spike_eight_three_dictionary:
            self.spike_eight_three = float(spike_eight_three_dictionary[spike]) #spike ratio
        else: pass
        
        if spike in spike_eight_three_err_dictionary:
            self.spike_eight_three_err = float(spike_eight_three_err_dictionary[spike]) #error of spike ratio
        else: pass
        
        #Other input parameters
        self.AS_U = float(AS_U)
        self.AS_Th = float(AS_Th)
        self.sample_wt = float(sample_wt)
        self.spike_wt = float(spike_wt)
        self.sample_name = sample_ID
        self.row = row_age
        self.file_Th = Th
        self.file_Thwash = Thwash
        self.file_U = U
        self.file_Uwash = Uwash
        self.filename_export = export_age
        
        #constants needed in age calculation
        self.wt_229 = 229.031756
        self.wt_230 = 230.033128
        self.wt_232 = 232.038051
        self.wt_233 = 233.039629
        self.wt_234 = 234.040947
        self.wt_235 = 235.043924
        self.wt_236 = 236.045563
        self.wt_238 = 238.050785
        self.five_counttime = 0.131
        self.four_counttime = 1.049
        self.three_counttime = 0.393
        self.two_nine_counttime = 1.049
        self.eight_five_rat = 137.82 #why not 137.83?  
        self.eight_filament_blank = 0.0001
        self.eight_filament_blank_err = 0.1
        self.sample_wt_err = float(sample_wt_err)
        self.spike_wt_err = float(spike_wt_err)
        self.two_nine_spike = 0.00065
        self.two_nine_spike_err = 0.00005
        self.AS_1amu = 1.00E-10
        self.AS_1amu_err = 0.25 * self.AS_1amu
        self.AS_2amu = self.AS_1amu/2.5
        self.AS_2amu_err = 0.25 * self.AS_2amu
        self.lambda_238 = 0.000000000155125
        self.lambda_234 = 0.0000028263*0.9985
        self.lambda_230 = 0.0000091577*1.0014
        self.threefive_four = 1E-11
        self.fourfour_four = 1E-11
        
        self.zerotwo_initial = float(zerotwo)
        self.zerotwo_initial_err = float(zerotwo_err)
        
        #chemblank values
        
        self.chemblank_eight = float(chem_lst[8] * (1/10.**15) * (1/self.wt_238) * (10.**12)) #in pmol
        self.chemblank_eight_err = float(chem_lst[9] * (1/10.**15) * (1/self.wt_238) * (10.**12)) #in pmol
        self.chemblank_two = float(chem_lst[2] * (1/10.**15) * (1/self.wt_232) * (10.**12)) #in pmol 
        self.chemblank_two_err = float(chem_lst[3] * (1/10.**15) * (1/self.wt_232) * (10.**12)) #in pmol 
        self.chemblank_zero = float(chem_lst[0] * (1/10.**18) * (1/self.wt_230) * (10.**15))#in fmol
        self.chemblank_zero_err = float(chem_lst[1] * (1/10.**18) * (1/self.wt_230) * (10.**15)) #in fmol


    def age_calculate_sem(self):
       
        """
        Import functions for U, Th and wash values
        """
        #234 unfiltered beam for plot
        working_e = isofilter(self.file_U, "D")
        self.four_beam_array = working_e.array()
        
        #index array for plot
        working_f = isofilter(self.file_U, "A")
        self.index_array_U = working_f.array()
        
        wb_U = Ucalculation(self.spike, self.AS_U, self.file_U)
        
        lstU_Th = wb_U.U_normalization_forTh() #provides a list for use in Th normalization
        """
            lst U_Th output is a list of the following values: 
                [0]: 236/233 measured ratio
                [1]: 236/233 measured ratio error
                [2]: 235/233 normalized ratio
                [3]: 235/233 normalized ratio error
                [4]: 236/233 corrected ratio
                [5]: 236/233 corrected ratio error
        """
        self.lstU_Age = wb_U.U_normalization_forAge() #provides a list for use in Age Calculation
        """
            lstU_Age output is a list of the following values: 
                [0]: 235/233 normalized ratio
                [1]: 235/233 normalized ratio error
                [2]: 235/234 normalized and corrected ratio
                [3]: 235/234 normalized and corrected ratio error
                [4]: Unfiltered 233 counts
                [5]: Filtered 234/235 counts
                [6]: Unfiltered 233 mean
        """
        
        #230 unfiltered array for plot
        working_e = isofilter(self.file_Th, "D")
        self.zero_beam_array = working_e.array()
        
        #index array for plot
        working_f = isofilter(self.file_Th, "A")
        self.index_array_Th = working_f.array()
        
        wb_Th = Thcalculation(self.spike, self.AS_Th, self.file_Th, lstU_Th)
        
        self.lstTh_Age = wb_Th.Th_normalization_forAge() #provides a list for use in Age Calculation
  
        """
            Th_normalization_forAge output is a list of the following values: 
                [0]: 230/229 corrected and normalized ratio
                [1]: 230/229 corrected and normalized ratio error
                [2]: 232/229 corrected and normalized ratio
                [3]: 232/229 corrected and normalized ratio error
                [4]: Unfiltered 229 mean
                [5]: Unfiltered 229 counts
        """
        
        background = background_values(self.file_Uwash, self.file_Thwash)
        
        self.lstU_wash = background.U_wash()
        """
            U_wash provides a list the following outputs for the Age Calculation: 
                [0]: 233 unfiltered wash in cps
                [1]: 234 unfiltered wash in cps
                [2]: 235 unfiltered wash in cps
            
        """
        
        self.Th_wash = background.Th_wash()
        """
            Th_wash provides the 230 unfiltered wash in cpm
        """
    
        """
        Age Calculation equations
        """
        
        #238 ppb
        
        five_three_max_err = ( (self.lstU_Age[6] * self.lstU_Age[0]) - self.lstU_wash[2] ) / (self.lstU_Age[6] - self.lstU_wash[0])
        
        eight_nmol = (((five_three_max_err -  self.spike_five_three) * self.spike_wt * self.spike_three * self.eight_five_rat)/1000) /self.sample_wt  
        
        chemblank_corr_238 = ((eight_nmol * self.sample_wt) - (self.chemblank_eight/1000)) / self.sample_wt
        
        filament_blank_corr_238 = chemblank_corr_238 * (1 - (self.eight_filament_blank/ (self.lstU_Age[6] * five_three_max_err
                                                        * self.eight_five_rat)))
        
        eight_ppb = filament_blank_corr_238 * self.wt_238
        
        
        #238 ppb error
        
        rel_err_1 = (self.lstU_Age[1]/self.lstU_Age[0]) 
        
        three_counting_err = 2 / np.sqrt(self.lstU_Age[6] * self.lstU_Age[4] * self.three_counttime)
    
        five_counting_err = 2 / np.sqrt(self.lstU_Age[6] * self.lstU_Age[0] * self.five_counttime * self.lstU_Age[4])
        
        rel_err_2 = np.sqrt( (five_counting_err**2) + (three_counting_err**2) + (three_counting_err**2)*(8.0/9.0) )
        
        rel_err_five_three = max(rel_err_1, rel_err_2)
        
        
        abs_err_five_three = rel_err_five_three * five_three_max_err
        
        eight_nmol_err = eight_nmol * np.sqrt( (np.sqrt((abs_err_five_three**2) + (0.0000527**2))/(five_three_max_err - self.spike_five_three))**2 +
                                              (self.spike_three_err/self.spike_three)**2)
        eight_nmol_err_rel = eight_nmol_err/eight_nmol
        
        chemblank_corr_238_err = np.sqrt( (eight_nmol_err**2) + ((self.chemblank_eight_err/1000)**2) )
        
        chemblank_corr_238_err_rel = chemblank_corr_238_err / chemblank_corr_238
        
        filament_blank_corr_238_err_rel = np.sqrt( (chemblank_corr_238_err_rel**2) + 
                                                   ( ((self.eight_filament_blank/(self.lstU_Age[6]*self.lstU_Age[0]*self.eight_five_rat)) *
                                                     np.sqrt((self.eight_filament_blank_err/self.eight_filament_blank)**2 + 
                                                              ((self.lstU_Age[6]*0.05)/self.lstU_Age[6])**2 + 
                                                              ((self.lstU_Age[1]/self.lstU_Age[0])**2)))
                                                    / (1 - (self.eight_filament_blank/(self.lstU_Age[6]*self.lstU_Age[0]*self.eight_five_rat))))**2)
        
        filament_blank_corr_238_err = filament_blank_corr_238_err_rel * filament_blank_corr_238
        
        eight_ppb_err = filament_blank_corr_238_err * self.wt_238
        
        eight_ppb_wt_err = eight_ppb * np.sqrt((eight_ppb_err/eight_ppb)**2 + (self.sample_wt_err/self.sample_wt)**2 + (self.spike_wt_err/self.spike_wt)**2 )
        
        
        #232 ppt
        
        two_nine_max_err = self.lstTh_Age[2]
        
        two_nine_spike_corr = two_nine_max_err - self.two_nine_spike
        
        two_nine_chemblank_corr = two_nine_spike_corr - ( self.chemblank_two/(self.spike_wt * self.spike_nine)  )
        
        two_pmol = two_nine_chemblank_corr * self.spike_wt * self.spike_nine/self.sample_wt
                 
        two_ppt = two_pmol * self.wt_232

        #232 ppt error
        
        abs_err_two_nine = self.lstTh_Age[3]
        
        two_nine_spike_corr_err = np.sqrt( (abs_err_two_nine**2) + (self.two_nine_spike_err **2) )
        
        two_nine_chemblank_corr_err = np.sqrt( (self.chemblank_two/(self.spike_wt*self.spike_nine) * 
                                                np.sqrt( (self.chemblank_two_err/self.chemblank_two)**2 + 
                                                        (self.spike_nine_err/self.spike_nine)**2))**2 +
                                                two_nine_spike_corr_err**2)
        
        two_pmol_err = two_pmol * np.sqrt( (two_nine_chemblank_corr_err/two_nine_chemblank_corr)**2 + 
                                           (self.spike_nine_err/self.spike_nine)**2)
        
        two_pmol_err_rel = two_pmol_err / two_pmol
        
        two_ppt_err = two_ppt * two_pmol_err_rel
        
        two_ppt_wt_err = two_ppt * np.sqrt( (two_ppt_err/two_ppt)**2 + (self.sample_wt_err/self.sample_wt)**2 + (self.spike_wt_err/self.spike_wt)**2)
        
        #230 pmol/g
        
        zero_nine_max_err = self.lstTh_Age[0]
        
        zero_nine_spike_corr = zero_nine_max_err - self.spike_zero_nine
        
        zero_nine_AS_corr = zero_nine_spike_corr - self.AS_1amu - (self.AS_2amu * self.lstTh_Age[2])
        
        zero_nine_darknoise_corr = zero_nine_AS_corr * (1 - ((self.Th_wash/60)/(self.lstTh_Age[4]*zero_nine_AS_corr)) )
        
        zero_nine_chemblank_corr = zero_nine_darknoise_corr - ( self.chemblank_zero/(self.spike_wt * self.spike_nine * 1000) )
        
        zero_pmol = (zero_nine_chemblank_corr * self.spike_wt * self.spike_nine) / self.sample_wt
        
        #230 pmol/g error
        
        zero_nine_counting_err = self.lstTh_Age[0] * 2 * np.sqrt( (1 / ((self.lstTh_Age[4]*self.lstTh_Age[0])*self.lstTh_Age[5]*self.two_nine_counttime)) + 
                                                             (1 / (self.lstTh_Age[4]*self.lstTh_Age[5]*self.two_nine_counttime)  ) )
        
        abs_err_zero_nine = max((zero_nine_max_err*0.00001), zero_nine_counting_err, self.lstTh_Age[1]  )
        
        zero_nine_spike_corr_err = np.sqrt( (abs_err_zero_nine**2) + (0.000003**2) )
        
        zero_nine_AS_corr_err = np.sqrt( (zero_nine_spike_corr_err**2) + (self.AS_1amu_err**2) + 
                                        ( self.AS_2amu * self.lstTh_Age[2] * np.sqrt( (self.AS_2amu_err/self.AS_2amu)**2 + 
                                         (self.lstTh_Age[3]/self.lstTh_Age[2])**2 ) )**2 )
        
        zero_nine_darknoise_corr_err = zero_nine_darknoise_corr * np.sqrt( (zero_nine_AS_corr_err/zero_nine_AS_corr)**2 + 
                                                                      (((self.Th_wash/60)/(self.lstTh_Age[4]*zero_nine_AS_corr)) * 
                                                                      np.sqrt((0.2**2) + (10/self.lstTh_Age[4])**2 + (zero_nine_AS_corr_err/zero_nine_AS_corr)**2 
                                                                              / (1 - ((self.Th_wash/60)/self.lstTh_Age[4]*zero_nine_AS_corr) ))
                                                                              )**2)
        
        zero_nine_chemblank_corr_err = np.sqrt( zero_nine_darknoise_corr_err**2 +
                                               ( (self.chemblank_zero/(self.spike_wt * self.spike_nine * 1000)) * 
                                                np.sqrt( (self.chemblank_zero_err/self.chemblank_zero)**2 + 
                                                         (self.spike_nine_err/self.spike_nine)**2 ))**2)
    
        zero_pmol_err = zero_pmol * np.sqrt((zero_nine_chemblank_corr_err/zero_nine_chemblank_corr)**2 + 
                                            (self.spike_nine_err/self.spike_nine)**2)
        
       
        zero_pmol_err_rel = zero_pmol_err / zero_pmol
        
        
        #230/232 atomic ratio
            
        zero_two_atomic = zero_pmol / two_pmol
        
        zero_two_atomic_final = zero_two_atomic * 10**6
        
        #230/232 atomic ratio error 
            
        zero_two_atomic_err_rel = np.sqrt( two_pmol_err_rel**2 + zero_pmol_err_rel**2 )
        
        zero_two_atomic_err = zero_two_atomic_err_rel * zero_two_atomic
        
        zero_two_atomic_err_final = zero_two_atomic_err * 10**6
        
        #d234U measured
            
        zero_nine_measuredU = self.lstU_Age[2] * (1 - self.lstU_wash[1]/(self.lstU_Age[6] * self.lstU_Age[2] * self.lstU_Age[0]))
        
        four_five_wt_avg = zero_nine_measuredU
        
        four_three_max_err = four_five_wt_avg * self.lstU_Age[0]
        
        four_five_tail_corr = four_five_wt_avg * (1 - ((4.0/9.0 * self.threefive_four) + (5.0/9.0 * self.fourfour_four)))
        
        four_five_spike_corr_234 = four_five_tail_corr * (1 - (self.spike_four_three/four_three_max_err))
        
        four_five_spike_corr_235 = four_five_spike_corr_234 * (1 / (1- (self.spike_five_three/five_three_max_err)))
        
        four_eight_ppm = (four_five_spike_corr_235 * 10**6) / self.eight_five_rat
        
        d234U_m = (( four_eight_ppm / ((self.lambda_238/self.lambda_234) * 10**6)) - 1) * 1000
        
        #d234U measured error
        
        zero_nine_measuredU_err_rel = self.lstU_Age[3] / zero_nine_measuredU
        
        rel_err_1 = np.sqrt(zero_nine_measuredU_err_rel**2 + (self.lstU_Age[1]/self.lstU_Age[0])**2)
        
        four_counting_err = 2 / (self.lstU_Age[6] * four_three_max_err * self.four_counttime * self.lstU_Age[4])**0.5
        
        rel_err_2 = np.sqrt(four_counting_err**2 + 2*three_counting_err**2 + (2.0/9.0)*three_counting_err**2)
        
        rel_err_four_three = max(rel_err_1, rel_err_2)
        
        four_five_wt_avg_err_rel = max(zero_nine_measuredU_err_rel**2, 
                                       np.sqrt(four_counting_err**2 + five_counting_err**2 + (2.0/9.0 * three_counting_err**2) ))
        
        four_five_tail_corr_err_rel = np.sqrt((four_five_wt_avg_err_rel**2) + 
                                              (np.sqrt((4.0/9.0 * self.threefive_four)**2 + (5.0/9.0 * self.fourfour_four)**2)/
                                               (1 - (4.0/9.0 * self.threefive_four + 5/9 * self.fourfour_four)) )**2)
        
        four_five_spike_corr_234_err_rel = np.sqrt((four_five_tail_corr_err_rel**2) + 
                                                   ((self.spike_four_three/four_three_max_err) * np.sqrt(0.002**2 + rel_err_four_three**2) /
                                                    (1 - self.spike_four_three/four_three_max_err))**2)
        
        four_five_spike_corr_235_err_rel = np.sqrt((four_five_spike_corr_234_err_rel**2) + 
                                                   ((self.spike_five_three/five_three_max_err) * np.sqrt(0.0005**2 + (rel_err_five_three/1000)**2) /
                                                    (1 - self.spike_five_three/five_three_max_err))**2)
        
        four_five_spike_corr_235_err = four_five_spike_corr_235_err_rel * four_five_spike_corr_235
        
        four_eight_ppm_err = (four_five_spike_corr_235_err * 10**6)/ self.eight_five_rat
        
        d234U_m_err = (four_eight_ppm_err / ((self.lambda_238/self.lambda_234) * 10**6)) * 1000
        
        #230Th/238U activity ratio
        
        zero_eight_atomic = (zero_pmol/(eight_ppb/self.wt_238))/1000
        
        zero_eight_activity = zero_eight_atomic * (self.lambda_230/self.lambda_238)
        
        #230Th/238U activity ratio error 
        
        zero_eight_atomic_err_rel = np.sqrt(zero_pmol_err_rel**2 + eight_nmol_err_rel **2 )
        
        zero_eight_activity_err = zero_eight_atomic_err_rel * zero_eight_activity
        
        #Uncorrected age calculation and error
        
        age_func = lambda t : zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        t_initial_guess = 0
        uncorrected_t = fsolve(age_func, t_initial_guess) #returns the value for t at which the solution is 0. This is true of all fsolve functions following this. 
        
        age_func_ThUmax = lambda t : (zero_eight_activity+zero_eight_activity_err) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        uncorrected_ThUmax = fsolve(age_func_ThUmax, t_initial_guess)
        
        age_func_ThUmin = lambda t : (zero_eight_activity-zero_eight_activity_err) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        uncorrected_ThUmin = fsolve(age_func_ThUmin, t_initial_guess)
        
        age_func_d234Umax = lambda t : zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + ((d234U_m + d234U_m_err)/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        uncorrected_d234Umax = fsolve(age_func_d234Umax, t_initial_guess)
        
        age_func_d234Umin = lambda t : zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + ((d234U_m - d234U_m_err)/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        uncorrected_d234Umin = fsolve(age_func_d234Umin, t_initial_guess)
        
        uncorrected_t_maxerr = np.sqrt((uncorrected_ThUmax - uncorrected_t)**2 + (uncorrected_d234Umin - uncorrected_t)**2)
        
        uncorrected_t_minerr = np.sqrt((uncorrected_ThUmin - uncorrected_t)**2 + (uncorrected_d234Umax - uncorrected_t)**2)
        
        uncorrected_t_err = (uncorrected_t_maxerr + uncorrected_t_minerr)/2
        
        #Corrected age calculation and error
        
        zero_two_initial = self.zerotwo_initial
        zero_two_initial_err = self.zerotwo_initial_err
        
        age_func_corrected_t = lambda t : (((zero_pmol - zero_two_initial*np.exp(-self.lambda_230*t)*two_pmol) * self.lambda_230/(filament_blank_corr_238 * 1000 * self.lambda_238)) - 
                                  (1 - np.exp(-self.lambda_230 * t) + (d234U_m/1000 * (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                  (1 - np.exp((self.lambda_234-self.lambda_230)*t)))))
        
        t_initial_guess = 0
        corrected_t = fsolve(age_func_corrected_t, t_initial_guess)
        
        zero_two_initial_now = zero_two_initial * np.exp(-self.lambda_230 * corrected_t)
        
        zero_two_initial_now_err = zero_two_initial_now * (zero_two_initial_err / zero_two_initial)
        
        corrected_zero_eight_activity = (zero_pmol - zero_two_initial_now*two_pmol) * self.lambda_230/(filament_blank_corr_238 * 1000 * self.lambda_238)
        
        corrected_zero_eight_activity_err = corrected_zero_eight_activity * np.sqrt( 
                                                                            (np.sqrt(((zero_two_initial_now * two_pmol) * np.sqrt((zero_two_initial_now_err/zero_two_initial_now)**2 
                                                                                    + (two_pmol_err/two_pmol)**2))**2 + zero_pmol_err**2) / 
                                                                                    (zero_pmol - zero_two_initial_now*two_pmol))**2 +
                                                                                    (filament_blank_corr_238_err/filament_blank_corr_238)**2)
        
        age_func_ThUmax = lambda t : (corrected_zero_eight_activity+corrected_zero_eight_activity_err) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_ThUmax = fsolve(age_func_ThUmax, t_initial_guess)
        
        age_func_ThUmin = lambda t : (corrected_zero_eight_activity-corrected_zero_eight_activity_err) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_ThUmin = fsolve(age_func_ThUmin, t_initial_guess)
    
        age_func_d234Umax = lambda t : corrected_zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + ((d234U_m + d234U_m_err)/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_d234Umax = fsolve(age_func_d234Umax, t_initial_guess)
        
        age_func_d234Umin = lambda t : corrected_zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + ((d234U_m - d234U_m_err)/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_d234Umin = fsolve(age_func_d234Umin, t_initial_guess)
        
        age_func_low = lambda t: ((zero_pmol - ((zero_two_initial + zero_two_initial_err) * np.exp(-self.lambda_230 * t)) *two_pmol) 
                                * self.lambda_230/(filament_blank_corr_238 * 1000. * self.lambda_238)) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000.) * 
                                             (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                             (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
    
        age_func_high = lambda t: ((zero_pmol - ((zero_two_initial - zero_two_initial_err) * np.exp(-self.lambda_230 * t)) *two_pmol) 
                                * self.lambda_230/(filament_blank_corr_238 * 1000. * self.lambda_238)) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000.) * 
                                             (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                             (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_age_low = fsolve(age_func_low, t_initial_guess)
        
        corrected_age_high = fsolve(age_func_high, t_initial_guess)
    
        corrected_t_maxerr = np.sqrt((corrected_ThUmax - corrected_t)**2 + (corrected_d234Umin - corrected_t)**2 + (corrected_age_high - corrected_t)**2 )
        
        corrected_t_minerr = np.sqrt((corrected_ThUmin - corrected_t)**2 + (corrected_d234Umax - corrected_t)**2 + (corrected_age_low - corrected_t)**2 )
        
        corrected_t_err = (corrected_t_maxerr + corrected_t_minerr)/2
    
        #Corrected initial d234U and error
        
        d234U_i = d234U_m * np.exp(self.lambda_234 * corrected_t)
        
        d234U_i_maxerr = np.sqrt( (d234U_m_err * np.exp(self.lambda_234 * corrected_t))**2 + 
                                 (d234U_m * np.exp(self.lambda_234 * (corrected_t + corrected_t_maxerr)) - d234U_i)**2)
        
        d234U_i_minerr = np.sqrt( (d234U_m_err * np.exp(self.lambda_234 * corrected_t))**2 + 
                                 (d234U_m * np.exp(self.lambda_234 * (corrected_t - corrected_t_minerr)) - d234U_i)**2)
        
        d234U_i_err = (d234U_i_maxerr + d234U_i_minerr)/2
        
        
        #Corrected age BP
        
        corrected_t_BP = corrected_t - (datetime.datetime.today().year - 1950.0)
        
        corrected_t_BP_err = corrected_t_err
        age_file = openpyxl.load_workbook(self.filename_export)
        sheet = age_file.worksheets[0]
        row = str(self.row)
        sheet['B' + row] = self.sample_name
        sheet['C' + row] = "{0:.1f}".format(eight_ppb)
        sheet['D' + row] = "± " + "{0:.1f}".format(eight_ppb_wt_err)
        sheet['E' + row] = "{0:.0f}".format(two_ppt)
        sheet['F' + row] = "± " + "{0:.0f}".format(two_ppt_wt_err)
        sheet['G' + row] = "{0:.1f}".format(zero_two_atomic_final)
        sheet['H' + row] = "± " + "{0:.1f}".format(zero_two_atomic_err_final)
        sheet['I' + row] = "{0:.1f}".format(d234U_m)
        sheet['J' + row] = "± " + "{0:.1f}".format(d234U_m_err)
        sheet['K' + row] = "{0:.5f}".format(zero_eight_activity)
        sheet['L' + row] = "± " + "{0:.5f}".format(zero_eight_activity_err)
        sheet['M' + row] = "%.0f" % uncorrected_t
        sheet['N' + row] = "± %.0f" % uncorrected_t_err
        sheet['O' + row] = "%.0f" % corrected_t
        sheet['P' + row] = "± %.0f" % corrected_t_err
        sheet['Q' + row] = "%.1f" % d234U_i
        sheet['R' + row] = "± %.1f" % d234U_i_err
        sheet['S' + row] = "%.0f" % corrected_t_BP
        sheet['T' + row] = "± %.0f" % corrected_t_BP_err
        
        age_file.save(self.filename_export)
        
        messagebox.showinfo("Success! ", "Age calculation finished! ")
        
        wb = plot_figure(self.four_beam_array, self.index_array_U, self.zero_beam_array, self.index_array_Th)
        
        wb.plot_fig()
        
        
class Application_semcups():
    
    def __init__(self, spike, AS_Th, sample_wt, spike_wt, sample_ID, row_age, spiked, spiked_err, unspiked, unspikedwash, Th, Thwash, U, Uwash, chem_lst, export_age, sample_wt_err, spike_wt_err, spike_conc_three, spike_conc_nine, zerotwo, zerotwo_err, Uwash_option):
     
        
        #spike input for Th calculation
        spike_six_three_dictionary = {"DIII-B":1.008398,"DIII-A": 1.008398,"1I":1.010128,"1H":1.010128}
        
        if spike in spike_six_three_dictionary:
            self.spike_six_three = float(spike_six_three_dictionary[spike]) #spike ratio
        else: 
            messagebox.showwarning("Error!", "No valid spike info entered! ")
            
        self.spike = str(spike)
        
        #Other input parameters
        self.AS_Th = float(AS_Th)
        self.sample_wt = float(sample_wt)
        self.spike_wt = float(spike_wt)
        self.sample_name = sample_ID
        self.row = row_age
        self.spiked_stand = float(spiked)
        self.spiked_stand_err = float(spiked_err)
        self.file_unspiked = unspiked
        self.file_unspikedwash = unspikedwash
        self.file_Th = Th
        self.file_Thwash = Thwash
        self.file_U = U
        self.file_Uwash = Uwash
        self.filename_export = export_age
        
        #constants needed in age calculation
        self.wt_229 = 229.031756
        self.wt_230 = 230.033128
        self.wt_232 = 232.038051
        self.wt_233 = 233.039629
        self.wt_234 = 234.040947
        self.wt_235 = 235.043924
        self.wt_236 = 236.045563
        self.wt_238 = 238.050785
        self.five_counttime = 0.131
        self.four_counttime = 1.049
        self.three_counttime = 0.393
        self.two_nine_counttime = 1.049
        self.eight_five_rat = 137.82 #why not 137.83?  
        self.eight_filament_blank = 0.0001
        self.eight_filament_blank_err = 0.1
        self.sample_wt_err = float(sample_wt_err)
        self.spike_wt_err = float(spike_wt_err)
        self.two_nine_spike = 0.00065
        self.two_nine_spike_err = 0.00005
        self.AS_1amu = 1.00E-10
        self.AS_1amu_err = 0.25 * self.AS_1amu
        self.AS_2amu = self.AS_1amu/2.5
        self.AS_2amu_err = 0.25 * self.AS_2amu
        self.lambda_238 = 0.000000000155125
        self.lambda_234 = 0.0000028263*0.9985
        self.lambda_232 = 0.000000000049475
        self.lambda_230 = 0.0000091577*1.0014
        self.threefive_four = 1E-11
        self.fourfour_four = 1E-11
        self.zerotwo_initial = float(zerotwo)
        self.zerotwo_initial_err = float(zerotwo_err)
        
        self.chem_lst = chem_lst
        
        self.spike_conc_three = spike_conc_three
        self.spike_conc_nine = spike_conc_nine
        
        self.Uwash_option = Uwash_option
        
        
    def age_calculate_semcups(self):
        
        wb_unspiked = unspiked_standard(self.file_unspiked, self.file_unspikedwash)
        
        lst_unspiked = wb_unspiked.unspiked_calc()
        
        """
            lst_unspiked is a list of the following values:
                [0]: 233/237 tail ratio mean
                [1]: 233/237 tail ratio mean error
                [2]: 234/237 tail ratio mean
                [3]: 234/237 tail ratio mean error
                [4]: 235/237 tail ratio mean
                [5]: 235/237 tail ratio mean error
                [6]: 236/237 tail ratio mean
                [7]: 236/237 tail ratio mean error
        """
        #234 unfiltered beam for plot
        working_e = isofilter(self.file_U, "D")
        self.four_beam_array = working_e.array()
        
        #index array for plot
        working_f = isofilter(self.file_U, "A")
        self.index_array_U = working_f.array()
        
        wb_cups = Calculation_forCups(self.spike, lst_unspiked, self.chem_lst, self.spike_wt, self.sample_wt, self.spike_conc_three, self.spike_conc_nine, self.sample_wt_err, self.spike_wt_err)
        
        lst_Ucups = wb_cups.U_calc(self.file_U, self.file_Uwash, self.spiked_stand, self.spiked_stand_err, self.Uwash_option)
        
        """
            lst_Ucups is a list of the following values: 
                [0]: 238 ppb
                [1]: 238 ppb err
                [2]: 238 ppb wt err 
                [3]: d234U
                [4]: d234U err
                [5]: 236/233 corrected ratio
                [6]: 236/233 corrected ratio err
                [7]: 237/238 blank corrected ratio
                [8]: 237/238 blank corrected ratio error
        """
        
        lst_UTh = [lst_Ucups[5], lst_Ucups[6]]
        
        #230 unfiltered array for plot
        working_e = isofilter(self.file_Th, "D")
        self.zero_beam_array = working_e.array()
        
        #index array for plot
        working_f = isofilter(self.file_Th, "A")
        self.index_array_Th = working_f.array()
        
        wb_Thsem = Thcalculation(self.spike_six_three, self.AS_Th, self.file_Th, lst_UTh)
        
        lst_ThAge = wb_Thsem.Th_normalization_forAge() 
        
        """
            lst_ThAge is a list of the following values: 
                [0]: 230/229 corrected and normalized ratio
                [1]: 230/229 corrected and normalized ratio error
                [2]: 232/229 corrected and normalized ratio
                [3]: 232/229 corrected and normalized ratio error
                [4]: unfiltered 229 mean (cps)
                [5]: unfiltered 229 counts
        """
        
        lst_Thsemcups = wb_cups.Thsem_calc(self.file_Th, self.file_Thwash, lst_ThAge)

        """
            lst_Thsemcups is a list of the following values:
                [0]: 230 fmol/g
                [1]: 230 ppt
                [2]: 230 ppt error
                [3]: 230 ppt wt err
                [4]: 232 pmol/g
                [5]: 232 pmol/g wt error
                [6]: 232 ppt
                [7]: 232 ppt error
                [8]: 232 ppt wt err
                [9]: 230/232 ratio
                [10]: 230/232 ratio error
        """
    
        #results from U cups and Th sem calculations
        eight_ppb = lst_Ucups[0]
        eight_ppb_err = lst_Ucups[1]
        eight_ppb_wt_err = lst_Ucups[2]
        d234U_m = lst_Ucups[3]
        d234U_m_err = lst_Ucups[4]
        zero_fmolg = lst_Thsemcups[0] * (10.**15)
        zero_pgg = lst_Thsemcups[1]
        zero_pgg_err = lst_Thsemcups[2]
        zero_pgg_wt_err = lst_Thsemcups[3]
        two_pmol = lst_Thsemcups[4] * (10.**12)
        two_pmol_err = lst_Thsemcups[5] * (10.**12)
        two_ppt = lst_Thsemcups[6]
        two_ppt_err = lst_Thsemcups[7]
        two_ppt_wt_err = lst_Thsemcups[8]
        zero_two_atomic_final = lst_Thsemcups[9] * (10.**6)
        zero_two_atomic_err_final = lst_Thsemcups[10] * (10.**6)
        
        eight_nmol = eight_ppb / self.wt_238
        eight_nmol_err = eight_nmol * (eight_ppb_err/eight_ppb)
        
        zero_pmol = zero_pgg / self.wt_230
        zero_pmol_err = zero_pmol * (zero_pgg_err/zero_pgg)
        
        """
        Calculated atomic ratios
        """
        
        #232/238 atomic ratio
        two_eight_atomic = (two_pmol/(10.**12))/((eight_ppb/(10.**9))/self.wt_238)
        two_eight_atomic_err = two_eight_atomic * np.sqrt((two_ppt_err/two_ppt)**2 + (eight_ppb_err/eight_ppb)**2)
        
        #230/238 atomic ratio
        zero_eight_atomic = (zero_fmolg/(10.**15))/((eight_ppb/(10.**9))/self.wt_238)
        zero_eight_atomic_err = abs(zero_eight_atomic * np.sqrt((zero_pgg_err/zero_pgg)**2 + (eight_ppb_err/eight_ppb)**2))
        
        
        #230/234 atomic ratio
        zero_four_atomic = (zero_fmolg/(10.**15))/((eight_ppb/(10.**9))/self.wt_238 * ((d234U_m/(10.**3)) + 1)  * (self.lambda_238/self.lambda_234) )
        zero_four_atomic_err = zero_four_atomic * np.sqrt((zero_pgg_err/zero_pgg)**2 + (eight_ppb_err/eight_ppb)**2 + 
                                                                      ((d234U_m_err/(10.**3))/((d234U_m/(10.**3))+1))**2 )
        
        
        """
        Calculated activity ratios
        """
        
        #232/238 activity ratio
        two_eight_activity= two_eight_atomic * (self.lambda_232/self.lambda_238)
        two_eight_activity_err = (two_eight_atomic_err/two_eight_atomic) * two_eight_activity
        
        #230/238 activity ratio
        zero_eight_activity = zero_eight_atomic * (self.lambda_230/self.lambda_238)
        zero_eight_activity_err = (zero_eight_atomic_err/zero_eight_atomic) * zero_eight_activity
        
        #230/234 activity ratio
        zero_four_activity = zero_four_atomic * (self.lambda_230/self.lambda_234)
        zero_four_activity_err = (zero_four_atomic_err/zero_four_atomic) * zero_four_activity
        
        """
        Age calculation
        """
        
        #Uncorrected age calculation and error
        
        age_func = lambda t : zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        t_initial_guess = 0
        uncorrected_t = fsolve(age_func, t_initial_guess) #returns the value for t at which the solution is 0. This is true of all fsolve functions following this. 
        
        age_func_ThUmax = lambda t : (zero_eight_activity+zero_eight_activity_err) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        uncorrected_ThUmax = fsolve(age_func_ThUmax, t_initial_guess)
        
        age_func_ThUmin = lambda t : (zero_eight_activity-zero_eight_activity_err) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        uncorrected_ThUmin = fsolve(age_func_ThUmin, t_initial_guess)
        
        age_func_d234Umax = lambda t : zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + ((d234U_m + d234U_m_err)/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        uncorrected_d234Umax = fsolve(age_func_d234Umax, t_initial_guess)
        
        age_func_d234Umin = lambda t : zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + ((d234U_m - d234U_m_err)/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        uncorrected_d234Umin = fsolve(age_func_d234Umin, t_initial_guess)
        
        uncorrected_t_maxerr = np.sqrt((uncorrected_ThUmax - uncorrected_t)**2 + (uncorrected_d234Umin - uncorrected_t)**2)
        
        uncorrected_t_minerr = np.sqrt((uncorrected_ThUmin - uncorrected_t)**2 + (uncorrected_d234Umax - uncorrected_t)**2)
        
        uncorrected_t_err = (uncorrected_t_maxerr + uncorrected_t_minerr)/2
        
        #Corrected age calculation and error
        
        zero_two_initial = self.zerotwo_initial
        zero_two_initial_err = self.zerotwo_initial_err
        
        age_func_corrected_t = lambda t : (((zero_pmol - zero_two_initial*np.exp(-self.lambda_230*t)*two_pmol) * self.lambda_230/(eight_nmol * 1000 * self.lambda_238)) - 
                                  (1 - np.exp(-self.lambda_230 * t) + (d234U_m/1000 * (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                  (1 - np.exp((self.lambda_234-self.lambda_230)*t)))))
        
        t_initial_guess = 0
        corrected_t = fsolve(age_func_corrected_t, t_initial_guess)
        
        zero_two_initial_now = zero_two_initial * np.exp(-self.lambda_230 * corrected_t)
        
        zero_two_initial_now_err = zero_two_initial_now * (zero_two_initial_err / zero_two_initial)
        
        corrected_zero_eight_activity = (zero_pmol - zero_two_initial_now*two_pmol) * self.lambda_230/(eight_nmol * 1000 * self.lambda_238)
        
        corrected_zero_eight_activity_err = corrected_zero_eight_activity * np.sqrt( 
                                                                            (np.sqrt(((zero_two_initial_now * two_pmol) * np.sqrt((zero_two_initial_now_err/zero_two_initial_now)**2 
                                                                                    + (two_pmol_err/two_pmol)**2))**2 + zero_pmol_err**2) / 
                                                                                    (zero_pmol - zero_two_initial_now*two_pmol))**2 +
                                                                                    (eight_nmol_err/eight_nmol)**2)
        
        age_func_ThUmax = lambda t : (corrected_zero_eight_activity+corrected_zero_eight_activity_err) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_ThUmax = fsolve(age_func_ThUmax, t_initial_guess)
        
        age_func_ThUmin = lambda t : (corrected_zero_eight_activity-corrected_zero_eight_activity_err) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_ThUmin = fsolve(age_func_ThUmin, t_initial_guess)
    
        age_func_d234Umax = lambda t : corrected_zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + ((d234U_m + d234U_m_err)/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_d234Umax = fsolve(age_func_d234Umax, t_initial_guess)
        
        age_func_d234Umin = lambda t : corrected_zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + ((d234U_m - d234U_m_err)/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_d234Umin = fsolve(age_func_d234Umin, t_initial_guess)
        
        age_func_low = lambda t: ((zero_pmol - ((zero_two_initial + zero_two_initial_err) * np.exp(-self.lambda_230 * t)) *two_pmol) 
                                * self.lambda_230/(eight_nmol * 1000. * self.lambda_238)) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000.) * 
                                             (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                             (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
    
        age_func_high = lambda t: ((zero_pmol - ((zero_two_initial - zero_two_initial_err) * np.exp(-self.lambda_230 * t)) *two_pmol) 
                                * self.lambda_230/(eight_nmol * 1000. * self.lambda_238)) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000.) * 
                                             (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                             (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_age_low = fsolve(age_func_low, t_initial_guess)
        
        corrected_age_high = fsolve(age_func_high, t_initial_guess)
    
        corrected_t_maxerr = np.sqrt((corrected_ThUmax - corrected_t)**2 + (corrected_d234Umin - corrected_t)**2 + (corrected_age_high - corrected_t)**2 )
        
        corrected_t_minerr = np.sqrt((corrected_ThUmin - corrected_t)**2 + (corrected_d234Umax - corrected_t)**2 + (corrected_age_low - corrected_t)**2 )
        
        corrected_t_err = (corrected_t_maxerr + corrected_t_minerr)/2
    
        #Corrected initial d234U and error
        
        d234U_i = d234U_m * np.exp(self.lambda_234 * corrected_t)
        
        d234U_i_maxerr = np.sqrt( (d234U_m_err * np.exp(self.lambda_234 * corrected_t))**2 + 
                                 (d234U_m * np.exp(self.lambda_234 * (corrected_t + corrected_t_maxerr)) - d234U_i)**2)
        
        d234U_i_minerr = np.sqrt( (d234U_m_err * np.exp(self.lambda_234 * corrected_t))**2 + 
                                 (d234U_m * np.exp(self.lambda_234 * (corrected_t - corrected_t_minerr)) - d234U_i)**2)
        
        d234U_i_err = (d234U_i_maxerr + d234U_i_minerr)/2
        
        
        #Corrected age BP
        
        corrected_t_BP = corrected_t - (datetime.datetime.today().year - 1950.0)
        
        corrected_t_BP_err = corrected_t_err
        age_file = openpyxl.load_workbook(self.filename_export)
        sheet = age_file.worksheets[0]
        row = str(self.row)
        sheet['B' + row] = self.sample_name
        sheet['C' + row] = "{0:.1f}".format(eight_ppb)
        sheet['D' + row] = "± " + "{0:.1f}".format(eight_ppb_wt_err)
        sheet['E' + row] = "{0:.0f}".format(two_ppt)
        sheet['F' + row] = "± " + "{0:.0f}".format(two_ppt_wt_err)
        sheet['G' + row] = "{0:.1f}".format(zero_two_atomic_final)
        sheet['H' + row] = "± " + "{0:.1f}".format(zero_two_atomic_err_final)
        sheet['I' + row] = "{0:.1f}".format(d234U_m)
        sheet['J' + row] = "± " + "{0:.1f}".format(d234U_m_err)
        sheet['K' + row] = "{0:.5f}".format(zero_eight_activity)
        sheet['L' + row] = "± " + "{0:.5f}".format(zero_eight_activity_err)
        sheet['M' + row] = "%.0f" % uncorrected_t
        sheet['N' + row] = "± %.0f" % uncorrected_t_err
        sheet['O' + row] = "%.0f" % corrected_t
        sheet['P' + row] = "± %.0f" % corrected_t_err
        sheet['Q' + row] = "%.1f" % d234U_i
        sheet['R' + row] = "± %.1f" % d234U_i_err
        sheet['S' + row] = "%.0f" % corrected_t_BP
        sheet['T' + row] = "± %.0f" % corrected_t_BP_err
        
        age_file.save(self.filename_export)
        
        messagebox.showinfo("Success! ", "Age calculation finished! ")
        
        wb = plot_figure(self.four_beam_array, self.index_array_U, self.zero_beam_array, self.index_array_Th)
        
        wb.plot_fig()
              
        
class Application_cups():
    
    def __init__(self, spike, sample_wt, spike_wt, sample_ID, row_age, spiked, spiked_err, unspiked, unspikedwash, Th, Thwash, U, Uwash, chem_lst, export_age, sample_wt_err, spike_wt_err, spike_conc_three, spike_conc_nine, zerotwo, zerotwo_err, Uwash_option, Thwash_option):
     
        self.spike = str(spike)
        
        #Other input parameters
        self.sample_wt = float(sample_wt)
        self.spike_wt = float(spike_wt)
        self.sample_name = sample_ID
        self.row = row_age
        self.spiked_stand = float(spiked)
        self.spiked_stand_err = float(spiked_err)
        self.file_unspiked = unspiked
        self.file_unspikedwash = unspikedwash
        self.file_Th = Th
        self.file_Thwash = Thwash
        self.file_U = U
        self.file_Uwash = Uwash
        self.filename_export = export_age
        
        #constants needed in age calculation
        self.wt_229 = 229.031756
        self.wt_230 = 230.033128
        self.wt_232 = 232.038051
        self.wt_233 = 233.039629
        self.wt_234 = 234.040947
        self.wt_235 = 235.043924
        self.wt_236 = 236.045563
        self.wt_238 = 238.050785
        self.five_counttime = 0.131
        self.four_counttime = 1.049
        self.three_counttime = 0.393
        self.two_nine_counttime = 1.049
        self.eight_five_rat = 137.82 #why not 137.83?  
        self.eight_filament_blank = 0.0001
        self.eight_filament_blank_err = 0.1
        self.sample_wt_err = float(sample_wt_err)
        self.spike_wt_err = float(spike_wt_err)

        self.two_nine_spike = 0.00065
        self.two_nine_spike_err = 0.00005
        self.AS_1amu = 1.00E-10
        self.AS_1amu_err = 0.25 * self.AS_1amu
        self.AS_2amu = self.AS_1amu/2.5
        self.AS_2amu_err = 0.25 * self.AS_2amu
        self.lambda_238 = 0.000000000155125
        self.lambda_234 = 0.0000028263*0.9985
        self.lambda_232 = 0.000000000049475
        self.lambda_230 = 0.0000091577*1.0014
        self.threefive_four = 1E-11
        self.fourfour_four = 1E-11
        
        self.zerotwo_initial = float(zerotwo)
        self.zerotwo_initial_err = float(zerotwo_err)
        
        self.chem_lst = chem_lst
        
        self.spike_conc_three = spike_conc_three
        self.spike_conc_nine = spike_conc_nine
        
        self.Uwash_option = Uwash_option
        self.Thwash_option = Thwash_option
   
    def age_calculate_cups(self):
        
        wb_unspiked = unspiked_standard(self.file_unspiked, self.file_unspikedwash)
        
        lst_unspiked = wb_unspiked.unspiked_calc()
        
        """
            lst_unspiked is a list of the following values:
                [0]: 233/237 tail ratio mean
                [1]: 233/237 tail ratio mean error
                [2]: 234/237 tail ratio mean
                [3]: 234/237 tail ratio mean error
                [4]: 235/237 tail ratio mean
                [5]: 235/237 tail ratio mean error
                [6]: 236/237 tail ratio mean
                [7]: 236/237 tail ratio mean error
        """
        #234 unfiltered beam for plot
        working_e = isofilter(self.file_U, "D")
        self.four_beam_array = working_e.array()
        
        #index array for plot
        working_f = isofilter(self.file_U, "A")
        self.index_array_U = working_f.array()
        
        wb_cups = Calculation_forCups(self.spike, lst_unspiked, self.chem_lst, self.spike_wt, self.sample_wt, self.spike_conc_three, self.spike_conc_nine, self.sample_wt_err, self.spike_wt_err)
        
        lst_Ucups = wb_cups.U_calc(self.file_U, self.file_Uwash, self.spiked_stand, self.spiked_stand_err, self.Uwash_option)
        
        """
            lst_Ucups is a list of the following values: 
                [0]: 238 ppb
                [1]: 238 ppb err
                [2]: 238 ppb wt err
                [3]: d234U 
                [4]: d234U err
                [5]: 236/233 corrected ratio
                [6]: 236/233 corrected ratio error
                [7]: 237/238 blank corrected ratio 
                [8]: 237/238 blank corrected ratio error
        """
        
        lst_UTh = [lst_Ucups[5], lst_Ucups[6], lst_Ucups[7], lst_Ucups[8]]
        
        
        #230 unfiltered array for plot
        working_e = isofilter(self.file_Th, "D")
        self.zero_beam_array = working_e.array()
        
        #index array for plot
        working_f = isofilter(self.file_Th, "A")
        self.index_array_Th = working_f.array()
        
        lst_Thcups = wb_cups.Thcups_calc(self.file_Th, self.file_Thwash, lst_UTh, self.Thwash_option)
        
        """
            lst_Thcups is a list of the following values:
                [0]: 230 fmol/g
                [1]: 230 ppt
                [2]: 230 ppt error
                [3]: 230 ppt wt err
                [4]: 232 pmol/g
                [5]: 232 pmol/g wt error
                [6]: 232 ppt
                [7]: 232 ppt error
                [8]: 232 ppt wt err
                [9]: 230/232 ratio
                [10]: 230/232 ratio error
        """
       
        #results from U cups and Th cups calculations
        eight_ppb = lst_Ucups[0]
        eight_ppb_err = lst_Ucups[1]
        eight_ppb_wt_err = lst_Ucups[2]
        d234U_m = lst_Ucups[3]
        d234U_m_err = lst_Ucups[4]
        zero_fmolg = lst_Thcups[0] * (10.**15)
        zero_pgg = lst_Thcups[1]
        zero_pgg_err = lst_Thcups[2]
        zero_pgg_wt_err = lst_Thcups[3]
        two_pmol = lst_Thcups[4] * (10.**12)
        two_pmol_err = lst_Thcups[5] * (10.**12)
        two_ppt = lst_Thcups[6]
        two_ppt_err = lst_Thcups[7]
        two_ppt_wt_err = lst_Thcups[8]
        zero_two_atomic_final = lst_Thcups[9] * (10.**6)
        zero_two_atomic_err_final = lst_Thcups[10] * (10.**6)
        
        eight_nmol = eight_ppb / self.wt_238
        eight_nmol_err = eight_nmol * (eight_ppb_err/eight_ppb)
        
        zero_pmol = zero_pgg / self.wt_230
        zero_pmol_err = zero_pmol * (zero_pgg_err/zero_pgg)
        
        
        """
        Calculated atomic ratios
        """
        
        #232/238 atomic ratio
        two_eight_atomic = (two_pmol/(10.**12))/((eight_ppb/(10.**9))/self.wt_238)
        two_eight_atomic_err = two_eight_atomic * np.sqrt((two_ppt_err/two_ppt)**2 + (eight_ppb_err/eight_ppb)**2)
        
        #230/238 atomic ratio
        zero_eight_atomic = (zero_fmolg/(10.**15))/((eight_ppb/(10.**9))/self.wt_238)
        zero_eight_atomic_err = abs(zero_eight_atomic * np.sqrt((zero_pgg_err/zero_pgg)**2 + (eight_ppb_err/eight_ppb)**2))
        
        
        #230/234 atomic ratio
        zero_four_atomic = (zero_fmolg/(10.**15))/((eight_ppb/(10.**9))/self.wt_238 * ((d234U_m/(10.**3)) + 1)  * (self.lambda_238/self.lambda_234) )
        zero_four_atomic_err = zero_four_atomic * np.sqrt((zero_pgg_err/zero_pgg)**2 + (eight_ppb_err/eight_ppb)**2 + 
                                                                      ((d234U_m_err/(10.**3))/((d234U_m/(10.**3))+1))**2 )
        
        
        
        """
        Calculated activity ratios
        """
        
        #232/238 activity ratio
        two_eight_activity= two_eight_atomic * (self.lambda_232/self.lambda_238)
        two_eight_activity_err = (two_eight_atomic_err/two_eight_atomic) * two_eight_activity
        
        #230/238 activity ratio
        zero_eight_activity = zero_eight_atomic * (self.lambda_230/self.lambda_238)
        zero_eight_activity_err = (zero_eight_atomic_err/zero_eight_atomic) * zero_eight_activity
        
        #230/234 activity ratio
        zero_four_activity = zero_four_atomic * (self.lambda_230/self.lambda_234)
        zero_four_activity_err = (zero_four_atomic_err/zero_four_atomic) * zero_four_activity
        
        """
        Age calculation
        """
        
        #Uncorrected age calculation and error
        
        age_func = lambda t : zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        t_initial_guess = 0
        uncorrected_t = fsolve(age_func, t_initial_guess) #returns the value for t at which the solution is 0. This is true of all fsolve functions following this. 
        
        age_func_ThUmax = lambda t : (zero_eight_activity+zero_eight_activity_err) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        uncorrected_ThUmax = fsolve(age_func_ThUmax, t_initial_guess)
        
        age_func_ThUmin = lambda t : (zero_eight_activity-zero_eight_activity_err) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        uncorrected_ThUmin = fsolve(age_func_ThUmin, t_initial_guess)
        
        age_func_d234Umax = lambda t : zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + ((d234U_m + d234U_m_err)/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        uncorrected_d234Umax = fsolve(age_func_d234Umax, t_initial_guess)

        age_func_d234Umin = lambda t : zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + ((d234U_m - d234U_m_err)/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        uncorrected_d234Umin = fsolve(age_func_d234Umin, t_initial_guess)
        
        uncorrected_t_maxerr = np.sqrt((uncorrected_ThUmax - uncorrected_t)**2 + (uncorrected_d234Umin - uncorrected_t)**2)
        
        uncorrected_t_minerr = np.sqrt((uncorrected_ThUmin - uncorrected_t)**2 + (uncorrected_d234Umax - uncorrected_t)**2)
        
        uncorrected_t_err = (uncorrected_t_maxerr + uncorrected_t_minerr)/2
        
        #Corrected age calculation and error
        
        zero_two_initial = self.zerotwo_initial
        zero_two_initial_err = self.zerotwo_initial_err
        
        age_func_corrected_t = lambda t : (((zero_pmol - zero_two_initial*np.exp(-self.lambda_230*t)*two_pmol) * self.lambda_230/(eight_nmol * 1000 * self.lambda_238)) - 
                                  (1 - np.exp(-self.lambda_230 * t) + ((d234U_m/1000) * (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                  (1 - np.exp((self.lambda_234-self.lambda_230)*t)))))
        
        t_initial_guess = 0
        corrected_t = fsolve(age_func_corrected_t, t_initial_guess)
        
        zero_two_initial_now = zero_two_initial * np.exp(-self.lambda_230 * corrected_t)
        
        zero_two_initial_now_err = zero_two_initial_now * (zero_two_initial_err / zero_two_initial)
        
        corrected_zero_eight_activity = (zero_pmol - zero_two_initial_now*two_pmol) * self.lambda_230/(eight_nmol * 1000 * self.lambda_238)
        
        corrected_zero_eight_activity_err = corrected_zero_eight_activity * np.sqrt( 
                                                                            (np.sqrt(((zero_two_initial_now * two_pmol) * np.sqrt((zero_two_initial_now_err/zero_two_initial_now)**2 
                                                                                    + (two_pmol_err/two_pmol)**2))**2 + zero_pmol_err**2) / 
                                                                                    (zero_pmol - zero_two_initial_now*two_pmol))**2 +
                                                                                    (eight_nmol_err/eight_nmol)**2)
        
        age_func_ThUmax = lambda t : (corrected_zero_eight_activity+corrected_zero_eight_activity_err) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_ThUmax = fsolve(age_func_ThUmax, t_initial_guess)
        
        age_func_ThUmin = lambda t : (corrected_zero_eight_activity-corrected_zero_eight_activity_err) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_ThUmin = fsolve(age_func_ThUmin, t_initial_guess)
        
        age_func_d234Umax = lambda t : corrected_zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + ((d234U_m + d234U_m_err)/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_d234Umax = fsolve(age_func_d234Umax, t_initial_guess)
        
        age_func_d234Umin = lambda t : corrected_zero_eight_activity - (1 - np.exp(-self.lambda_230*t) + ((d234U_m - d234U_m_err)/1000) * 
                                                 (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                                 (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_d234Umin = fsolve(age_func_d234Umin, t_initial_guess)
        
        age_func_low = lambda t: ((zero_pmol - ((zero_two_initial + zero_two_initial_err) * np.exp(-self.lambda_230 * t)) *two_pmol) 
                                * self.lambda_230/(eight_nmol * 1000. * self.lambda_238)) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000.) * 
                                             (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                             (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
    
        age_func_high = lambda t: ((zero_pmol - ((zero_two_initial - zero_two_initial_err) * np.exp(-self.lambda_230 * t)) *two_pmol) 
                                * self.lambda_230/(eight_nmol * 1000. * self.lambda_238)) - (1 - np.exp(-self.lambda_230*t) + (d234U_m/1000.) * 
                                             (self.lambda_230/(self.lambda_230-self.lambda_234)) * 
                                             (1 - np.exp((self.lambda_234 - self.lambda_230)*t)))
        
        corrected_age_low = fsolve(age_func_low, t_initial_guess)
        
        corrected_age_high = fsolve(age_func_high, t_initial_guess)
    
        corrected_t_maxerr = np.sqrt((corrected_ThUmax - corrected_t)**2 + (corrected_d234Umin - corrected_t)**2 + (corrected_age_high - corrected_t)**2 )
        
        corrected_t_minerr = np.sqrt((corrected_ThUmin - corrected_t)**2 + (corrected_d234Umax - corrected_t)**2 + (corrected_age_low - corrected_t)**2 )
        
        corrected_t_err = (corrected_t_maxerr + corrected_t_minerr)/2
    
        #Corrected initial d234U and error
        
        d234U_i = d234U_m * np.exp(self.lambda_234 * corrected_t)
        
        d234U_i_maxerr = np.sqrt( (d234U_m_err * np.exp(self.lambda_234 * corrected_t))**2 + 
                                 (d234U_m * np.exp(self.lambda_234 * (corrected_t + corrected_t_maxerr)) - d234U_i)**2)
        
        d234U_i_minerr = np.sqrt( (d234U_m_err * np.exp(self.lambda_234 * corrected_t))**2 + 
                                 (d234U_m * np.exp(self.lambda_234 * (corrected_t - corrected_t_minerr)) - d234U_i)**2)
        
        d234U_i_err = (d234U_i_maxerr + d234U_i_minerr)/2
        
        
        #Corrected age BP
        
        corrected_t_BP = corrected_t - (datetime.datetime.today().year - 1950.0)
        
        corrected_t_BP_err = corrected_t_err
        age_file = openpyxl.load_workbook(self.filename_export)
        sheet = age_file.worksheets[0]
        row = str(self.row)
        sheet['B' + row] = self.sample_name
        sheet['C' + row] = "{0:.1f}".format(eight_ppb)
        sheet['D' + row] = "± " + "{0:.1f}".format(eight_ppb_wt_err)
        sheet['E' + row] = "{0:.0f}".format(two_ppt)
        sheet['F' + row] = "± " + "{0:.0f}".format(two_ppt_wt_err)
        sheet['G' + row] = "{0:.1f}".format(zero_two_atomic_final)
        sheet['H' + row] = "± " + "{0:.1f}".format(zero_two_atomic_err_final)
        sheet['I' + row] = "{0:.1f}".format(d234U_m)
        sheet['J' + row] = "± " + "{0:.1f}".format(d234U_m_err)
        sheet['K' + row] = "{0:.5f}".format(zero_eight_activity)
        sheet['L' + row] = "± " + "{0:.5f}".format(zero_eight_activity_err)
        sheet['M' + row] = "%.0f" % uncorrected_t
        sheet['N' + row] = "± %.0f" % uncorrected_t_err
        sheet['O' + row] = "%.0f" % corrected_t
        sheet['P' + row] = "± %.0f" % corrected_t_err
        sheet['Q' + row] = "%.1f" % d234U_i
        sheet['R' + row] = "± %.1f" % d234U_i_err
        sheet['S' + row] = "%.0f" % corrected_t_BP
        sheet['T' + row] = "± %.0f" % corrected_t_BP_err
        
        age_file.save(self.filename_export)
        
        messagebox.showinfo("Success! ", "Age calculation finished! ")
        
        wb = plot_figure(self.four_beam_array, self.index_array_U, self.zero_beam_array, self.index_array_Th)
        
        wb.plot_fig()
        
        
"""
FUNCTIONS USED IN AGE CALCULATION
"""

class unspiked_standard():
    
    def __init__(self, unspiked_file, unspiked_wash_file):
        
        self.filename_unspiked = unspiked_file
        self.filename_unspiked_wash = unspiked_wash_file
        
    def unspiked_calc(self):
        
        """
        Function for calculating 237 tail values from unspiked standard run
        """

        #constants used
        three_wt = 233.039629
        three_five_wt = 233.5
        four_wt = 234.040947
        four_five_wt = 234.5
        five_wt = 235.043924
        six_wt = 236.045563
        
        """
        UNSPIKED STANDARD
        """
        
        """
        Unspiked standard wash values
        """
        
        #233U
        unspiked_three_wash_working = isofilter(self.filename_unspiked_wash,"C")
        unspiked_three_wash = unspiked_three_wash_working.getMean()
        if unspiked_three_wash < 0: unspiked_three_wash = 0.0
        unspiked_three_wash_err = abs( (2 * unspiked_three_wash_working.getStanddev()) / (unspiked_three_wash_working.getCounts())**0.5)
        
        #233.5U
        unspiked_three_five_wash_working = isofilter(self.filename_unspiked_wash, "D")
        unspiked_three_five_wash = unspiked_three_five_wash_working.getMean()
        if unspiked_three_five_wash < 0: unspiked_three_five_wash = 0.0
        unspiked_three_five_wash_err = abs( (2 * unspiked_three_five_wash_working.getStanddev()) / (unspiked_three_five_wash_working.getCounts())**0.5)
        
        #234.5U
        unspiked_four_five_wash_working = isofilter(self.filename_unspiked_wash, "E")
        unspiked_four_five_wash = unspiked_four_five_wash_working.getMean()
        if unspiked_four_five_wash < 0: unspiked_four_five_wash = 0.0
        unspiked_four_five_wash_err = abs( (2 * unspiked_four_five_wash_working.getStanddev()) / (unspiked_four_five_wash_working.getCounts())**0.5)
        
        #236U
        unspiked_six_wash_working = isofilter(self.filename_unspiked_wash, "F")
        unspiked_six_wash = unspiked_six_wash_working.getMean()
        if unspiked_six_wash < 0: unspiked_six_wash = 0.0
        unspiked_six_wash_err = abs( (2 * unspiked_six_wash_working.getStanddev()) / (unspiked_six_wash_working.getCounts())**0.5)
        
        #237U
        unspiked_seven_wash_working = isofilter(self.filename_unspiked_wash, "G")
        unspiked_seven_wash = unspiked_seven_wash_working.getMean()
        if unspiked_seven_wash < 0: unspiked_seven_wash = 0
        unspiked_seven_wash_err = abs( (2 * unspiked_six_wash_working.getStanddev()) / (unspiked_six_wash_working.getCounts())**0.5)
        
        """
        Unspiked standard values
        """
        
        #233U
        unspiked_three_working = isofilter(self.filename_unspiked, "C")
        unspiked_three = unspiked_three_working.getMean()
        unspiked_three_err = abs( (2 * unspiked_three_working.getStanddev()) / (unspiked_three_working.getCounts()) ** 0.5 )
        
        #233.5U
        unspiked_three_five_working = isofilter(self.filename_unspiked, "D")
        unspiked_three_five = unspiked_three_five_working.getMean()
        unspiked_three_five_err = abs( (2 * unspiked_three_five_working.getStanddev()) / (unspiked_three_five_working.getCounts()) ** 0.5 )
        
        #234.5U
        unspiked_four_five_working = isofilter(self.filename_unspiked, "E")
        unspiked_four_five = unspiked_four_five_working.getMean()
        unspiked_four_five_err = abs( (2 * unspiked_four_five_working.getStanddev()) / (unspiked_four_five_working.getCounts()) ** 0.5 )
        
        #236U
        unspiked_six_working = isofilter(self.filename_unspiked, "F")
        unspiked_six = unspiked_six_working.getMean()
        unspiked_six_err = abs( (2 * unspiked_six_working.getStanddev()) / (unspiked_six_working.getCounts()) ** 0.5 )
        
        #237U
        unspiked_seven_working = isofilter(self.filename_unspiked, "G")
        unspiked_seven = unspiked_seven_working.getMean()
        unspiked_seven_err = abs( (2 * unspiked_seven_working.getStanddev()) / (unspiked_seven_working.getCounts()) ** 0.5 )
        
        #233/237U
        unspiked_three_seven_working = isofilter(self.filename_unspiked, "H")
        unspiked_three_seven = unspiked_three_seven_working.getMean()
        unspiked_three_seven_err = abs( (2 * unspiked_three_seven_working.getStanddev()) / (unspiked_three_seven_working.getCounts()) ** 0.5 )
        
        #233.5/237U
        unspiked_three_five_seven_working = isofilter(self.filename_unspiked, "I")
        unspiked_three_five_seven = unspiked_three_five_seven_working.getMean()
        unspiked_three_five_seven_err = abs( (2 * unspiked_three_five_seven_working.getStanddev()) / (unspiked_three_five_seven_working.getCounts()) ** 0.5 )
        
        #234.5/237U
        unspiked_four_five_seven_working = isofilter(self.filename_unspiked, "J")
        unspiked_four_five_seven = unspiked_four_five_seven_working.getMean()
        unspiked_four_five_seven_err = abs( (2 * unspiked_four_five_seven_working.getStanddev()) / (unspiked_four_five_seven_working.getCounts()) ** 0.5 )
        
        #236/237U
        unspiked_six_seven_working = isofilter(self.filename_unspiked, "K")
        unspiked_six_seven = unspiked_six_seven_working.getMean()
        unspiked_six_seven_err = abs( (2 * unspiked_six_seven_working.getStanddev()) / (unspiked_six_seven_working.getCounts()) ** 0.5 )
        
        """
        Unspiked blank corrected tail values
        """
        unspiked_three_seven_blankcorr = unspiked_three_seven - ((unspiked_three_wash/unspiked_seven)/(1 - (unspiked_seven_wash/unspiked_seven) ))
        
        unspiked_three_five_seven_blankcorr = unspiked_three_five_seven - ((unspiked_three_five_wash/unspiked_seven)/(1 - (unspiked_seven_wash/unspiked_seven) ))
        
        unspiked_four_five_seven_blankcorr = unspiked_four_five_seven - ((unspiked_four_five_wash/unspiked_seven)/(1 - (unspiked_seven_wash/unspiked_seven) ))
        
        unspiked_six_seven_blankcorr = unspiked_six_seven - ((unspiked_six_wash/unspiked_seven)/(1 - (unspiked_seven_wash/unspiked_seven) ))
        
        """
        Measured tail error values
        """
        
        #2s relative errors (int)
            
        three_2s_rel_err = max((unspiked_three_seven_err/unspiked_three_seven), (unspiked_three_wash/unspiked_three))
        
        three_five_2s_rel_err = max((unspiked_three_five_seven_err/unspiked_three_five_seven), (unspiked_three_five_wash/unspiked_three_five))
        
        four_five_2s_rel_err = max((unspiked_four_five_seven_err/unspiked_four_five_seven), (unspiked_four_five_wash/unspiked_four_five))
        
        six_2s_rel_err = max((unspiked_six_seven_err/unspiked_six_seven), (unspiked_six_wash/unspiked_six))
        
        #Max tails
        
        three_max_tail = unspiked_three_seven_blankcorr * (1 + three_2s_rel_err)
        
        three_five_max_tail = unspiked_three_five_seven_blankcorr * (1 + three_five_2s_rel_err)
        
        four_five_max_tail = unspiked_four_five_seven_blankcorr * (1 + four_five_2s_rel_err)
        
        six_max_tail = unspiked_six_seven_blankcorr * (1 + six_2s_rel_err)
        
        #Min tails
        
        three_min_tail = unspiked_three_seven_blankcorr * (1 - three_2s_rel_err)
        
        three_five_min_tail = unspiked_three_five_seven_blankcorr * (1 - three_five_2s_rel_err)
        
        four_five_min_tail = unspiked_four_five_seven_blankcorr * (1 - four_five_2s_rel_err)
        
        six_min_tail = unspiked_six_seven_blankcorr * (1 - six_2s_rel_err)
        
        """
        Corrected tail error values
        """
        
        #calculate log line of best fit
        
        y_log_min = np.array([np.log10(three_min_tail), np.log10(three_five_min_tail), np.log10(four_five_min_tail), np.log10(six_min_tail)])
        
        y_log_max = np.array([np.log10(three_max_tail), np.log10(three_five_max_tail), np.log10(four_five_max_tail), np.log10(six_max_tail)])
        
        x = np.array([three_wt, three_five_wt, four_five_wt, six_wt])
        
        def line(x, a, b):
            return a + b * x
    
        popt_min, pcov_min = curve_fit(line, x, y_log_min)
        
        b_min = 10 ** popt_min[0]
        
        m_min = 10 ** popt_min[1]
    
        popt_max, pcov_max = curve_fit(line, x, y_log_max)
        
        b_max = 10 ** popt_max[0]
        
        m_max = 10 ** popt_max[1]
        
        #Max corrected tails
        
        three_max_corr_tail = b_max * (m_max ** three_wt)
        
        three_five_max_corr_tail = b_max * (m_max ** three_five_wt)
        
        four_max_corr_tail = b_max * (m_max ** four_wt)
        
        four_five_max_corr_tail = b_max * (m_max ** four_five_wt)
        
        five_max_corr_tail = b_max * (m_max ** five_wt)
        
        six_max_corr_tail = b_max * (m_max ** six_wt)
        
        #Min corrected tails
    
        three_min_corr_tail = b_min * (m_min ** three_wt)
        
        three_five_min_corr_tail = b_min * (m_min ** three_five_wt)
        
        four_min_corr_tail = b_min * (m_min ** four_wt)
        
        four_five_min_corr_tail = b_min * (m_min ** four_five_wt)
        
        five_min_corr_tail = b_min * (m_min ** five_wt)
        
        six_min_corr_tail = b_min * (m_min ** six_wt)
        
        #Offset from measurement
        
        three_tail_offset = (((three_max_corr_tail + three_min_corr_tail)/2) - unspiked_three_seven_blankcorr) / unspiked_three_seven_blankcorr
        
        three_five_tail_offset = (((three_five_max_corr_tail + three_five_min_corr_tail)/2) - unspiked_three_five_seven_blankcorr) / unspiked_three_five_seven_blankcorr
        
        four_five_tail_offset = (((four_five_max_corr_tail + four_five_min_corr_tail)/2) - unspiked_four_five_seven_blankcorr) / unspiked_four_five_seven_blankcorr
        
        six_tail_offset = (((six_max_corr_tail + six_min_corr_tail)/2) - unspiked_six_seven_blankcorr) / unspiked_six_seven_blankcorr
        
        four_tail_offset = ((four_five_tail_offset * (four_wt - three_five_wt)) + (three_five_tail_offset * (four_five_wt - four_wt))) / (four_five_wt - three_five_wt)
                            
        five_tail_offset = ((six_tail_offset * (five_wt - four_five_wt)) + (four_five_tail_offset * (six_wt - five_wt))) / (six_wt - four_five_wt)
    
        """
        Finalized tail values
        """
        
        #Tail/237
        
        self.three_seven_tail = unspiked_three_seven_blankcorr
        
        three_five_seven_tail = unspiked_three_five_seven_blankcorr
        
        self.four_seven_tail = np.average([four_max_corr_tail, four_min_corr_tail]) * ( 1 - four_tail_offset)
        
        four_five_seven_tail = unspiked_four_five_seven_blankcorr
        
        self.five_seven_tail = five_max_corr_tail #why not the same as 234/237 tail calc?
        
        self.six_seven_tail = unspiked_six_seven_blankcorr
        
        #2s relative error calculated
        
        three_2s_rel_err_corr = three_2s_rel_err
        
        three_five_2s_rel_err_corr = three_five_2s_rel_err
        
        four_2s_rel_err_corr = max( np.sqrt((three_five_2s_rel_err_corr**2) + (four_five_2s_rel_err**2)), (((four_max_corr_tail-four_min_corr_tail)/2)/self.four_seven_tail)  )
        
        four_five_2s_rel_err_corr = four_five_2s_rel_err
        
        five_2s_rel_err_corr = max( np.sqrt((four_five_2s_rel_err_corr**2) + (six_2s_rel_err**2)), (((five_max_corr_tail-five_min_corr_tail)/2)/self.five_seven_tail)  )
        
        six_2s_rel_err_corr = six_2s_rel_err
        
        #Final error
        
        self.three_seven_err = self.three_seven_tail * max(three_2s_rel_err_corr, 0.05)
        
        self.four_seven_err = self.four_seven_tail * max(four_2s_rel_err_corr, 0.05)
        
        self.five_seven_err = self.five_seven_tail * max(five_2s_rel_err_corr, 0.05)
        
        self.six_seven_err = self.six_seven_tail * max(six_2s_rel_err_corr, 0.05)
        
        """
        Optional message box for displaying unspiked tail values. Delete quotation marks to run.
        # message box 
        messagebox.showinfo( "UNSPIKED STANDARD TAIL VALUES: ",
        "\n233/237: " + str("{0:.4f}".format(self.three_seven_tail)) + " ± " + str("{0:.4f}".format(self.three_seven_err)) +\
        "\n234/237: " + str("{0:.4f}".format(self.four_seven_tail)) + " ± " + str("{0:.4f}".format(self.four_seven_err)) +\
        "\n235/237: " + str("{0:.4f}".format(self.five_seven_tail)) + " ± " + str("{0:.4f}".format(self.five_seven_err)) +\
        "\n236/237: " + str("{0:.4f}".format(self.six_seven_tail)) + " ± " + str("{0:.4f}".format(self.six_seven_err)))
        """
        # deleting excel files
        try:
            os.remove("unspiked.xlsx")
        except: pass
    
        try: 
            os.remove("unspiked_wash.xlsx")
        except: pass
    
        lst_unspiked = [self.three_seven_tail, self.three_seven_err, self.four_seven_tail, self.four_seven_err, self.five_seven_tail, 
                        self.five_seven_err, self.six_seven_tail, self.six_seven_err]
        
        return lst_unspiked

class Calculation_forCups():
    
    def __init__(self, spike_input, lst_unspiked, chem_lst, spike_wt, sample_wt, spike_conc_three, spike_conc_nine, sample_wt_err, spike_wt_err):
        
        spike = spike_input
        
        #derives spike value based off dictionary entries
        spike_six_three_dictionary = {"DIII-B":1.008398,"DIII-A": 1.008398,"1I":1.010128,"1H":1.010128}
        spike_six_three_err_dictionary = {"DIII-B": 0.00015, "DIII-A": 0.00015, "1I": 0.00015, "1H": 0.00015}
        spike_three_dictionary = {"DIII-B": 0.78938, "DIII-A": 0.78933, "1I": 0.61351, "1H": 0.78997}
        spike_three_err_dictionary = {"DIII-B": 0.00002, "DIII-A": 0.00002, "1I": 0.00002, "1H": 0.00002}
        spike_nine_dictionary = {"DIII-B": 0.21734, "DIII-A": 0.21705, "1I": 0.177187, "1H": 0.22815}
        spike_nine_err_dictionary = {"DIII-B": 0.00001, "DIII-A": 0.00002, "1I": 0.00001, "1H": 0.00001}
        spike_zero_nine_dictionary = {"DIII-B": 0.0000625, "DIII-A": 0.0000625, "1I": 0.0000402, "1H": 0.0000402}
        spike_zero_nine_err_dictionary = {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.0000011, "1H": 0.0000011}
        spike_nine_two_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.0018, "1H": 0.0018}
        spike_nine_two_err_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.00, "1H": 0.00}
        spike_four_three_dictionary = {"DIII-B": 0.003195, "DIII-A": 0.003195, "1I":0.003180, "1H": 0.003180}
        spike_four_three_err_dictionary= {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.000003, "1H": 0.000003}
        spike_five_three_dictionary = {"DIII-B": 0.10532, "DIII-A": 0.10532, "1I": 0.10521, "1H":0.10521}
        spike_five_three_err_dictionary = {"DIII-B": 0.00003, "DIII-A": 0.00003, "1I": 0.00003, "1H": 0.00003}
        spike_eight_three_dictionary = {"DIII-B": 0.01680, "DIII-A": 0.01680, "1I": 0.01700, "1H":0.01700 }
        spike_eight_three_err_dictionary = {"DIII-B": 0.00001, "DIII-A": 0.00001,"1I": 0.00001, "1H": 0.00001}
        spike_three_nine_dictionary = {"DIII-B": 0.27533001, "DIII-A": 0.27498005, "1I": 0.28880866, "1H": 0.28880844}

        if spike in spike_six_three_dictionary:
            self.spike_six_three = float(spike_six_three_dictionary[spike]) #spike ratio
        else: 
            messagebox.showwarning("Error!", "No valid spike info entered! ")
        
        if spike in spike_six_three_err_dictionary: 
            self.spike_six_three_err = float(spike_six_three_err_dictionary[spike]) #error of spike ratio
        
        if spike_conc_three == 0.0:
            if spike in spike_three_dictionary:
                self.spike_three = float(spike_three_dictionary[spike]) #in pmol/g
            else:pass
        else: self.spike_three = float(spike_conc_three)
    
        if spike in spike_three_err_dictionary:
            self.spike_three_err = float(spike_three_err_dictionary[spike]) #in pmol/g
        else:pass
    
        if spike_conc_nine == 1.0:
            if spike in spike_three_nine_dictionary:
                self.spike_nine = float(spike_three_nine_dictionary[spike]) * self.spike_three #in pmol/g
            else: pass
        else:
            if spike in spike_nine_dictionary: 
                self.spike_nine = float(spike_nine_dictionary[spike])
    
        if spike in spike_nine_err_dictionary: 
            self.spike_nine_err = float(spike_nine_err_dictionary[spike]) #in pmol/g
        else: pass
    
        if spike in spike_zero_nine_dictionary:
            self.spike_zero_nine = float(spike_zero_nine_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_zero_nine_err_dictionary:
            self.spike_zero_nine_err = float(spike_zero_nine_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_nine_two_dictionary: 
            self.spike_nine_two = float(spike_nine_two_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_nine_two_err_dictionary:
            self.spike_nine_two_err = float(spike_nine_two_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_four_three_dictionary:
            self.spike_four_three = float(spike_four_three_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_four_three_err_dictionary:
            self.spike_four_three_err = float(spike_four_three_err_dictionary[spike]) #error of spike ratio
        else: pass
            
        if spike in spike_five_three_dictionary:
            self.spike_five_three = float(spike_five_three_dictionary[spike]) #spike ratio
        else: pass
        
        if spike in spike_five_three_err_dictionary:
            self.spike_five_three_err = float(spike_five_three_err_dictionary[spike]) #error of spike ratio
        else: pass
        
        if spike in spike_eight_three_dictionary:
            self.spike_eight_three = float(spike_eight_three_dictionary[spike]) #spike ratio
        else: pass
        
        if spike in spike_eight_three_err_dictionary:
            self.spike_eight_three_err = float(spike_eight_three_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        #weight constants
        self.wt_229 = 229.031756
        self.wt_230 = 230.033128
        self.wt_232 = 232.038051
        self.wt_233 = 233.039629
        self.wt_234 = 234.040947
        self.wt_235 = 235.043924
        self.wt_236 = 236.045563
        self.wt_238 = 238.050785
        
        self.spike_wt = spike_wt
        self.sample_wt = sample_wt
        self.sample_wt_err = sample_wt_err
        self.spike_wt_err = spike_wt_err
        
    
        #tail ratios and errors
        self.three_seven_tail = lst_unspiked[0]
        self.three_seven_tail_err = lst_unspiked[1]
        self.four_seven_tail = lst_unspiked[2]
        self.four_seven_tail_err = lst_unspiked[3]
        self.five_seven_tail = lst_unspiked[4]
        self.five_seven_tail_err = lst_unspiked[5]
        self.six_seven_tail = lst_unspiked[6]
        self.six_seven_tail_err = lst_unspiked[7]
    
        #chemblank values
        self.chemblank_zero = float(chem_lst[0]) #in ag
        self.chemblank_zero_err = float(chem_lst[1]) #in ag
        self.chemblank_two = float(chem_lst[2]) #in fg
        self.chemblank_two_err = float(chem_lst[3]) #in fg
        self.chemblank_four = float(chem_lst[4]) #in ag
        self.chemblank_four_err = float(chem_lst[5]) #in ag
        self.chemblank_five = float(chem_lst[6]) #in fg
        self.chemblank_five_err = float(chem_lst[7]) #in fg
        self.chemblank_eight = float(chem_lst[8]) #in fg
        self.chemblank_eight_err = float(chem_lst[9]) #in fg
        
        #chemblank ratio values
        self.chemblank_zero_nine = ((self.chemblank_zero/(10.**18))/self.wt_234)/(self.spike_wt * self.spike_nine/(10.**12))
        self.chemblank_zero_nine_err = (self.chemblank_zero_err/self.chemblank_zero) * self.chemblank_zero_nine
        self.chemblank_two_nine = ((self.chemblank_two/(10.**15))/self.wt_235)/(self.spike_wt * self.spike_nine/(10.**12))
        self.chemblank_two_nine_err = (self.chemblank_two_err/self.chemblank_two) * self.chemblank_two_nine
        self.chemblank_four_three = ((self.chemblank_four/(10.**18))/self.wt_234)/(self.spike_wt * self.spike_three/(10.**12))
        self.chemblank_four_three_err = (self.chemblank_four_err/self.chemblank_four) * self.chemblank_four_three
        self.chemblank_five_three = ((self.chemblank_five/(10.**15))/self.wt_235)/(self.spike_wt * self.spike_three/(10.**12))
        self.chemblank_five_three_err = (self.chemblank_five_err/self.chemblank_five) * self.chemblank_five_three
        self.chemblank_eight_three = ((self.chemblank_eight/(10.**15))/self.wt_238)/(self.spike_wt * self.spike_three/(10.**12))
        self.chemblank_eight_three_err = (self.chemblank_eight_err/self.chemblank_eight) * self.chemblank_eight_three
        
        
    def U_calc(self, Ufile, Uwash, spiked_standard, spiked_standard_err, Uwash_option):
        
        """"
        U cups calculation
        """
        
        int_time = 1.049
        
        self.filename_U = Ufile
        self.filename_Uwash = Uwash
        
        self.spiked_standard = spiked_standard
        self.spiked_standard_err = spiked_standard_err
        
        """
        Machine blanks
        """
        int_time_234 = 1.049
        int_time_other = 0.262 #233, 235, 236, 237, 238
        
        if Uwash_option == "sem":
            
            #233 
            working_three_mb = isofilter(self.filename_Uwash, "D")
            three_mb = working_three_mb.getMean()
            three_mb_err = max( (2 * working_three_mb.getStanddev()/np.sqrt(working_three_mb.getCounts())), 
                               (three_mb * 2 / np.sqrt(working_three_mb.getCounts() * int_time_other * three_mb)) )
            
            #234 
            working_four_mb = isofilter(self.filename_Uwash, "E")
            four_mb = working_four_mb.getMean()
            four_mb_err = max( (2 * working_four_mb.getStanddev()/np.sqrt(working_four_mb.getCounts())), 
                               (four_mb * 2 / np.sqrt(working_four_mb.getCounts() * int_time_234 * four_mb)) )
            
            #235
            working_five_mb = isofilter(self.filename_Uwash, "F")
            five_mb = working_five_mb.getMean()
            five_mb_err = max( (2 * working_five_mb.getStanddev()/np.sqrt(working_five_mb.getCounts())), 
                               (five_mb * 2 / np.sqrt(working_five_mb.getCounts() * int_time_other * five_mb)) )
            
            #236
            working_six_mb = isofilter(self.filename_Uwash, "G")
            six_mb = working_six_mb.getMean()
            six_mb_err = max( (2 * working_six_mb.getStanddev()/np.sqrt(working_six_mb.getCounts())), 
                               (six_mb * 2 / np.sqrt(working_six_mb.getCounts() * int_time_other * six_mb)) )
            
            #237
            working_seven_mb = isofilter(self.filename_Uwash, "H")
            seven_mb = working_seven_mb.getMean()
            if seven_mb < 0: seven_mb = 0.0
            seven_mb_err = max( (2 * working_seven_mb.getStanddev()/np.sqrt(working_seven_mb.getCounts())), 
                               (seven_mb * 2 / np.sqrt(working_seven_mb.getCounts() * int_time_other * seven_mb)) )
            
            #234/238
            working_four_eight_mb = isofilter(self.filename_Uwash, "L")
            four_eight_mb = working_four_eight_mb.getMean()
            four_eight_mb_err = 2 * working_four_eight_mb.getStanddev()/np.sqrt(working_four_eight_mb.getCounts())
            
            #238
            eight_mb = four_mb / four_eight_mb
            eight_mb_err = eight_mb * (four_eight_mb_err/four_eight_mb)
            
        elif Uwash_option == "cups":
            
            #233 
            working_three_mb = isofilter(self.filename_Uwash, "D")
            three_mb = working_three_mb.getMean() * 62422000
            three_mb_err = (2 * working_three_mb.getStanddev()/np.sqrt(working_three_mb.getCounts())) * 6242000
            
            #234
            working_four_mb = isofilter(self.filename_Uwash, "E")
            four_mb = working_four_mb.getMean() * 62422000
            four_mb_err = (2 * working_four_mb.getStanddev()/np.sqrt(working_four_mb.getCounts())) * 6242000
            
            #235
            working_five_mb = isofilter(self.filename_Uwash, "F")
            five_mb = working_five_mb.getMean() * 62422000
            five_mb_err = (2 * working_five_mb.getStanddev()/np.sqrt(working_five_mb.getCounts())) * 6242000
            
            #236
            working_six_mb = isofilter(self.filename_Uwash, "G")
            six_mb = working_six_mb.getMean() * 62422000
            six_mb_err = (2 * working_six_mb.getStanddev()/np.sqrt(working_six_mb.getCounts())) * 6242000
            
            #237
            working_seven_mb = isofilter(self.filename_Uwash, "H")
            seven_mb = working_seven_mb.getMean()
            if seven_mb < 0: seven_mb = 0.0
            seven_mb_err = 2 * working_seven_mb.getStanddev()/np.sqrt(working_seven_mb.getCounts())
            
            #238
            working_eight_mb = isofilter(self.filename_Uwash, "I")
            eight_mb = working_eight_mb.getMean() * 62422000
            eight_mb_err = (2 * working_eight_mb.getStanddev()/np.sqrt(working_eight_mb.getCounts())) * 6242000
        
        """
        Measured beam intensities
        """
        #233
        working_three_beam = isofilter(self.filename_U,"C")
        three_beam_volts = working_three_beam.getMean()
        three_beam_cps = three_beam_volts/(10**11)/(1.602/(10**19))
        three_beam_abs_err = 2 * working_three_beam.getStanddev()/np.sqrt(working_three_beam.getCounts())
        three_beam_rel_err = 2 / np.sqrt(three_beam_cps * working_three_beam.getCounts() * int_time) * 1000
        
        #234
        working_four_beam = isofilter(self.filename_U,"D")
        four_beam_volts = working_four_beam.getMean()
        four_beam_cps = four_beam_volts/(10**11)/(1.602/(10**19))
        four_beam_abs_err = 2 * working_four_beam.getStanddev()/np.sqrt(working_four_beam.getCounts())
        four_beam_rel_err = 2 / np.sqrt(four_beam_cps * working_four_beam.getCounts() * int_time) * 1000
        
        #235
        working_five_beam = isofilter(self.filename_U,"E")
        five_beam_volts = working_five_beam.getMean()
        five_beam_cps = five_beam_volts/(10**11)/(1.602/(10**19))
        five_beam_abs_err = 2 * working_five_beam.getStanddev()/np.sqrt(working_five_beam.getCounts())
        five_beam_rel_err = 2 / np.sqrt(five_beam_cps * working_five_beam.getCounts() * int_time) * 1000
        
        #236
        working_six_beam = isofilter(self.filename_U,"F")
        six_beam_volts = working_six_beam.getMean()
        six_beam_cps = six_beam_volts/(10**11)/(1.602/(10**19))
        six_beam_abs_err = 2 * working_six_beam.getStanddev()/np.sqrt(working_six_beam.getCounts())
        six_beam_rel_err = 2 / np.sqrt(six_beam_cps * working_six_beam.getCounts() * int_time) * 1000
        
        #237
        working_seven_beam = isofilter(self.filename_U,"G")
        seven_beam_cps = working_seven_beam.getMean()
        seven_beam_abs_err = 2 * working_seven_beam.getStanddev()/np.sqrt(working_seven_beam.getCounts())
        seven_beam_rel_err = 2 / np.sqrt(seven_beam_cps * working_seven_beam.getCounts() * int_time) * 1000
        
        #238
        working_eight_beam = isofilter(self.filename_U,"H")
        eight_beam_volts = working_eight_beam.getMean()
        eight_beam_cps = eight_beam_volts/(10**11)/(1.602/(10**19))
        eight_beam_abs_err = 2 * working_eight_beam.getStanddev()/np.sqrt(working_eight_beam.getCounts())
        eight_beam_rel_err = 2 / np.sqrt(eight_beam_cps * working_eight_beam.getCounts() * int_time) * 1000
        
        """
        Measured ratios and corrected errors and arrays
        """
        
        #234/233 measured ratio
        four_three_working = isofilter(self.filename_U, "I")
        spiked_four_three_meas = four_three_working.getMean()
        spiked_four_three_meas_err = max((2 * four_three_working.getStanddev() / np.sqrt(four_three_working.getCounts())), 
                                          spiked_four_three_meas * np.sqrt(four_beam_rel_err**2)/(10**3))
        spiked_four_three_meas_array = isocorrection().array(self.filename_U, "I")
        
        
        #235/233 measured ratio
        five_three_working = isofilter(self.filename_U, "J")
        spiked_five_three_meas = five_three_working.getMean()
        spiked_five_three_meas_err = max((2 * five_three_working.getStanddev() / np.sqrt(five_three_working.getCounts())),
                                         spiked_five_three_meas * np.sqrt(five_beam_rel_err**2)/(10**3))
        spiked_five_three_meas_array = isocorrection().array(self.filename_U, "I")
        
        #236/233 measured ratio
        six_three_working = isofilter(self.filename_U, "K")
        spiked_six_three_meas = six_three_working.getMean()
        spiked_six_three_meas_err = max((2 * six_three_working.getStanddev() / np.sqrt(six_three_working.getCounts())),
                                         spiked_six_three_meas * np.sqrt(six_beam_rel_err**2 + three_beam_rel_err**2)/(10**3))
        spiked_six_three_meas_array = isocorrection().array(self.filename_U, "K")
        
        #238/233 measured ratio
        eight_three_working = isofilter(self.filename_U, "L")
        spiked_eight_three_meas = eight_three_working.getMean()
        spiked_eight_three_meas_err = max((2 * eight_three_working.getStanddev() / np.sqrt(eight_three_working.getCounts())),
                                         spiked_eight_three_meas * np.sqrt(eight_beam_rel_err**2)/(10**3))
        spiked_eight_three_meas_array = isocorrection().array(self.filename_U, "L")
        
        #237/238 measured ratio
        seven_eight_working = isofilter(self.filename_U, "M")
        spiked_seven_eight_meas = seven_eight_working.getMean()
        spiked_seven_eight_meas_err = max((2 * seven_eight_working.getStanddev() / np.sqrt(seven_eight_working.getCounts())), 
                                          spiked_seven_eight_meas * np.sqrt(eight_beam_rel_err**2 + seven_beam_rel_err**2)/(10**3))
        spiked_seven_eight_meas_array = isocorrection().array(self.filename_U, "M")
     
        
        """
        Drift correction for 234/233
        """
        
        #234/233 drift correction
    
        four_array = isocorrection().array(self.filename_U, "D")
        three_array = isocorrection().array(self.filename_U, "C")
        four_three_array = isocorrection().array(self.filename_U, "I")
        drift_array = isocorrection().drift_correction_offset(four_array, four_three_array)
        four_three_drift_corrected_array = isocorrection().drift_correction(drift_array, three_array)
        
        
        four_three_drift_corrected_err = max( (2 * np.nanstd(a = four_three_drift_corrected_array, ddof = 1)/ np.sqrt(len(four_three_drift_corrected_array[np.logical_not(np.isnan(four_three_drift_corrected_array))]))),
                                             (np.nanmean(four_three_drift_corrected_array) * np.sqrt(four_beam_rel_err**2)/(10**3)) )
        
        """
        Machine blank correction
        """
        
        #234/233 mb corrected
        four_three_mb_corrected_array = isocorrection().machine_blank_correction(four_three_drift_corrected_array, 
                                                     three_beam_volts, three_mb, four_mb)
        four_three_mb_corrected_err = np.sqrt(four_three_drift_corrected_err**2 + (four_mb_err/three_beam_cps)**2)
        
        #235/233 mb corrected
        five_three_mb_corrected_array = isocorrection().machine_blank_correction(isocorrection().array(self.filename_U, "J"), 
                                                      three_beam_volts, three_mb, five_mb)
        five_three_mb_corrected_err = np.sqrt(spiked_five_three_meas_err**2 + (five_mb_err/three_beam_cps)**2)
        
        #236/233 mb corrected
        six_three_mb_corrected_array = isocorrection().machine_blank_correction(isocorrection().array(self.filename_U, "K"), 
                                                     three_beam_volts, three_mb, six_mb)
        six_three_mb_corrected_err = np.sqrt(spiked_six_three_meas_err**2 + (six_mb_err/three_beam_cps)**2)
        
        #238/233 mb corrected
        eight_three_mb_corrected_array = isocorrection().machine_blank_correction(isocorrection().array(self.filename_U, "L"),
                                                       three_beam_volts, three_mb, eight_mb)
        eight_three_mb_corrected_err = np.sqrt(spiked_eight_three_meas_err**2 + (eight_mb_err/three_beam_cps)**2)
        
        #237/238 mb corrected
        seven_eight_mb_corrected_array = isocorrection().machine_blank_correction(isocorrection().array(self.filename_U, "M"), 
                                                       eight_beam_volts, eight_mb, seven_mb)
        
        AS_seven_eight = np.nanmean(seven_eight_mb_corrected_array)
        AS_seven_eight_err = np.sqrt(spiked_seven_eight_meas_err**2 + (seven_mb_err/three_beam_cps/(np.nanmean(eight_three_mb_corrected_array)))**2)

       
        """
        Tail correction
        """
        
        #233/238 tail ratios ppm
        three_eight_tail = self.three_seven_tail * AS_seven_eight * (10**6)
        three_eight_tail_err = three_eight_tail * np.sqrt( (self.three_seven_tail_err/self.three_seven_tail)**2 + (AS_seven_eight_err/AS_seven_eight)**2 ) 
        
        #234/238 tail ratios ppm
        four_eight_tail = self.four_seven_tail * AS_seven_eight * (10**6)
        four_eight_tail_err = four_eight_tail * np.sqrt( (self.four_seven_tail_err/self.four_seven_tail)**2 + (AS_seven_eight_err/AS_seven_eight)**2 )
        
        #235/238 tail ratios ppm
        five_eight_tail = self.five_seven_tail * AS_seven_eight * (10**6)
        five_eight_tail_err = five_eight_tail * np.sqrt( (self.five_seven_tail_err/self.five_seven_tail)**2 + (AS_seven_eight_err/AS_seven_eight)**2 )
        
        #236/238 tail ratios ppm
        six_eight_tail = self.six_seven_tail * AS_seven_eight * (10**6)
        six_eight_tail_err = six_eight_tail * np.sqrt( (self.six_seven_tail_err/self.six_seven_tail)**2 + (AS_seven_eight_err/AS_seven_eight)**2 )
        
        #238/233 machine blank corrected mean
        eight_three_mb_corr = np.nanmean(eight_three_mb_corrected_array)
        
        #234/233 tail corrected
        four_three_tail_corrected_array = isocorrection().tail_correction(four_three_mb_corrected_array, four_eight_tail, three_eight_tail, eight_three_mb_corr, option = 'norm')
        four_three_tail_corrected_err = np.sqrt(four_three_mb_corrected_err**2 + (four_eight_tail_err/(10**6)*(np.nanmean(eight_three_mb_corrected_array)))**2)
        
        #235/233 tail corrected
        five_three_tail_corrected_array = isocorrection().tail_correction(five_three_mb_corrected_array, five_eight_tail, three_eight_tail, eight_three_mb_corr, option = 'norm')
        five_three_tail_corrected_err = np.sqrt(five_three_mb_corrected_err**2 + (five_eight_tail_err/(10**6) * (np.nanmean(eight_three_mb_corrected_array)))**2)
        
        #236/233 tail corrected
        six_three_tail_corrected_array = isocorrection().tail_correction(six_three_mb_corrected_array, six_eight_tail, three_eight_tail, eight_three_mb_corr, option = 'norm')
        six_three_tail_corrected_mean = np.nanmean(six_three_tail_corrected_array)
        six_three_tail_corrected_err = np.sqrt(six_three_mb_corrected_err**2 + (six_eight_tail_err/(10**6) * (np.nanmean(eight_three_mb_corrected_array)))**2)
        
        #238/233 tail corrected
        eight_three_tail_corrected_array = isocorrection().tail_correction(eight_three_mb_corrected_array, three_eight_tail, three_eight_tail, eight_three_mb_corr, option = '238/233')
        eight_three_tail_corrected_err = eight_three_mb_corrected_err
        
        #238/235 tail corrected
        eight_five_tail_corrected_array = isocorrection().tail_correction_alt(eight_three_tail_corrected_array, five_three_tail_corrected_array)
        eight_five_tail_corrected_err = np.sqrt((2 * np.nanstd(a = eight_five_tail_corrected_array, ddof = 1)/np.sqrt( len(eight_five_tail_corrected_array[np.logical_not(np.isnan(eight_five_tail_corrected_array))])))**2 + 
                                                 ((five_eight_tail_err/(10**6) * (eight_three_mb_corr/(np.nanmean(five_three_mb_corrected_array)))) * np.nanmean(eight_five_tail_corrected_array))**2 )
        
        #234/238 tail corrected
        four_eight_tail_corrected_array = isocorrection().tail_correction_alt(four_three_tail_corrected_array, eight_three_tail_corrected_array)
        four_eight_tail_corrected_err = np.sqrt((2 * np.nanstd(a = four_eight_tail_corrected_array, ddof = 1)/np.sqrt( len(four_eight_tail_corrected_array[np.logical_not(np.isnan(four_eight_tail_corrected_array))])))**2 + 
                                                 (four_eight_tail_err/(10**6))**2)
        
        
        """
        Fractionation correction
        """
        
        #234/233 fractionation corrected
        four_three_fract_corrected_array = isocorrection().fractionation_correction(four_three_tail_corrected_array,    
                                                        six_three_tail_corrected_mean, '234', '233', self.spike_six_three)
        four_three_fract_corrected_mean = np.nanmean(four_three_fract_corrected_array)
        four_three_fract_corrected_err = np.sqrt((four_three_tail_corrected_err/np.nanmean(four_three_tail_corrected_array))**2 + 
                                                 (1 * (six_three_tail_corrected_err/six_three_tail_corrected_mean)/3)**2) * four_three_fract_corrected_mean
        
        #235/233 fractionation corrected                                         
        five_three_fract_corrected_array = isocorrection().fractionation_correction(five_three_tail_corrected_array, 
                                                        six_three_tail_corrected_mean, '235', '233', self.spike_six_three)
        five_three_fract_corrected_mean = np.nanmean(five_three_fract_corrected_array)
        five_three_fract_corrected_err = np.sqrt((five_three_tail_corrected_err/np.nanmean(five_three_tail_corrected_array))**2 + 
                                                 (2 * (six_three_tail_corrected_err/six_three_tail_corrected_mean)/3)**2) * five_three_fract_corrected_mean
        
        #238/233 fractionation corrected                                         
        eight_three_fract_corrected_array = isocorrection().fractionation_correction(eight_three_tail_corrected_array,  
                                                         six_three_tail_corrected_mean, '238', '233', self.spike_six_three)
        eight_three_fract_corrected_mean = np.nanmean(eight_three_fract_corrected_array)
        eight_three_fract_corrected_err = np.sqrt((eight_three_tail_corrected_err/np.nanmean(eight_three_tail_corrected_array))**2 + 
                                                  (5 * (six_three_tail_corrected_err/six_three_tail_corrected_mean)/3)**2) * eight_three_fract_corrected_mean
        
        #238/235 fractionation corrected                                          
        eight_five_fract_corrected_array = isocorrection().fractionation_correction(eight_five_tail_corrected_array,  
                                                        six_three_tail_corrected_mean, '238', '235', self.spike_six_three)
        eight_five_fract_corrected_mean = np.nanmean(eight_five_fract_corrected_array)
        eight_five_fract_corrected_err = np.sqrt((eight_five_tail_corrected_err/np.nanmean(eight_five_tail_corrected_array))**2 + 
                                                 (3 * (six_three_tail_corrected_err/six_three_tail_corrected_mean)/3)**2) * eight_five_fract_corrected_mean
        
        #238/234 fractionation corrected                                         
        four_eight_fract_corrected_array = isocorrection().fractionation_correction(four_eight_tail_corrected_array, 
                                                        six_three_tail_corrected_mean, '234', '238', self.spike_six_three)
        four_eight_fract_corrected_mean = np.nanmean(four_eight_fract_corrected_array)
        four_eight_fract_corrected_err = np.sqrt((four_eight_tail_corrected_err/np.nanmean(four_eight_tail_corrected_array))**2 + 
                                                 (4 * (six_three_tail_corrected_err/six_three_tail_corrected_mean)/3)**2) * four_eight_fract_corrected_mean
                      
                                                                                     
        """
        Spike correction
        """
        
        #234/233 spike corrected
        four_three_spike_corrected_mean = four_three_fract_corrected_mean  - self.spike_four_three
        four_three_spike_corrected_err = np.sqrt(four_three_fract_corrected_err**2 + self.spike_four_three_err**2)
        
        #235/233 spike corrected
        five_three_spike_corrected_mean = five_three_fract_corrected_mean - self.spike_five_three
        five_three_spike_corrected_err = np.sqrt(five_three_fract_corrected_err**2 + self.spike_five_three_err**2)
        
        #238/233 spike corrected
        eight_three_spike_corrected_mean = eight_three_fract_corrected_mean - self.spike_eight_three
        eight_three_spike_corrected_err = np.sqrt(eight_three_fract_corrected_err**2 + self.spike_eight_three_err**2)
        
        #238/235 spike corrected
        eight_five_spike_corrected_mean = (eight_five_fract_corrected_mean - self.spike_eight_three/five_three_fract_corrected_mean)/(1 - self.spike_five_three/five_three_fract_corrected_mean)
        eight_five_spike_corrected_err = np.sqrt(eight_five_fract_corrected_err**2 + (eight_five_fract_corrected_mean * np.sqrt((self.spike_five_three_err/five_three_fract_corrected_mean)**2 + 
                                                                                                                                (self.spike_eight_three_err/eight_three_fract_corrected_mean)**2))**2)
        #234/238 spike corrected
        four_eight_spike_corrected_mean = ((four_eight_fract_corrected_mean - self.spike_four_three/eight_three_fract_corrected_mean)/(1 - self.spike_eight_three/eight_three_fract_corrected_mean)) * (10**6) #in ppm
        four_eight_spike_corrected_err = np.sqrt(four_eight_fract_corrected_err**2 + (four_eight_fract_corrected_mean * np.sqrt((self.spike_four_three_err/four_three_fract_corrected_mean)**2 + 
                                                                                                                                (self.spike_eight_three_err/eight_three_fract_corrected_mean)**2))**2) * (10**6) #in ppm                                                                   
                                                                                                            
        """
        Chemistry blank corrections
        """
        #234/233 chem blank corrected
        four_three_chem_corrected_mean = four_three_spike_corrected_mean - self.chemblank_four_three
        four_three_chem_corrected_err = np.sqrt(four_three_spike_corrected_err**2 + self.chemblank_four_three_err**2)
        
        #235/233 chem blank corrected
        five_three_chem_corrected_mean = five_three_spike_corrected_mean - self.chemblank_five_three
        five_three_chem_corrected_err = np.sqrt(five_three_spike_corrected_err**2 + self.chemblank_five_three_err**2)
        
        #238/233 chem blank corrected
        eight_three_chem_corrected_mean = eight_three_spike_corrected_mean - self.chemblank_eight_three
        eight_three_chem_corrected_err = np.sqrt(eight_three_spike_corrected_err**2 + self.chemblank_eight_three_err**2)
        
        #238/235 chem blank corrected
        eight_five_chem_corrected_mean = (eight_five_spike_corrected_mean - self.chemblank_five_three/five_three_spike_corrected_mean)/(1 - self.chemblank_eight_three/five_three_spike_corrected_mean)
        eight_five_chem_corrected_err = np.sqrt( eight_five_spike_corrected_err**2 + (np.sqrt((self.chemblank_five_three_err/five_three_spike_corrected_mean)**2 + (self.chemblank_eight_three_err/eight_three_spike_corrected_mean)**2))**2 )
        
        #234/238 chem blank corrected
        four_eight_chem_corrected_mean = (four_eight_spike_corrected_mean/(10.**6) - self.chemblank_four_three/eight_three_spike_corrected_mean)/(1 - self.chemblank_eight_three/eight_three_spike_corrected_mean) * (10.**6) #in ppm
        four_eight_chem_corrected_err = np.sqrt( (four_eight_spike_corrected_err/(10.**6))**2 + ((four_eight_spike_corrected_mean/(10.**6)) *
                                                  np.sqrt((self.chemblank_four_three_err/four_three_spike_corrected_mean)**2 + (self.chemblank_eight_three_err/eight_three_spike_corrected_mean)**2))**2) * (10.**6) #in ppm                                         
        
        
        """
        112A Normalization
        """                                         
        
        four_eight_112A = 52.852
        eight_five_112A = 137.832
    
        four_eight_normalized_mean = four_eight_chem_corrected_mean - (self.spiked_standard - four_eight_112A)
        four_eight_normalized_err = np.sqrt(four_eight_chem_corrected_err**2 + self.spiked_standard_err**2)
        
        lambda_234 = 0.00000282206
        lambda_238 = 0.000000000155125
        
        #d234U in delta units
        d234U = ((four_eight_normalized_mean/(10**6)) * lambda_234/lambda_238 - 1) * 1000
        d234U_err = ((four_eight_normalized_err/(10**6)) * lambda_234/lambda_238) * 1000
        
        #epsilon-238 in epsilon units
        e238U = ((eight_five_chem_corrected_mean/eight_five_112A) - 1) * 10000
        e238U_err = (eight_five_chem_corrected_err/eight_five_112A) * 10000
        
        """
        238U ppb
        """
        pmolg_238 = ((eight_three_chem_corrected_mean * self.spike_wt * self.spike_three/(10.**12))/self.sample_wt) * (10.**12)
        pmolg_238_err = pmolg_238 * np.sqrt((self.spike_three_err/self.spike_three)**2 + (eight_three_chem_corrected_err/eight_three_chem_corrected_mean)**2)
        pmolg_238_wt_err = pmolg_238 * np.sqrt((pmolg_238_err/pmolg_238)**2 + (self.sample_wt_err/self.sample_wt)**2 + (self.spike_wt_err/self.spike_wt)**2)
        
        ppb_238 = (pmolg_238/(10.**12)) * self.wt_238 * (10.**9)
        ppb_238_err = abs(ppb_238 * (pmolg_238_err/pmolg_238))
        ppb_238_wt_err = abs(ppb_238 * (pmolg_238_wt_err/pmolg_238))
        
        lstU_cups = [ppb_238, ppb_238_err, ppb_238_wt_err, d234U, d234U_err, six_three_tail_corrected_mean, six_three_tail_corrected_err, AS_seven_eight, AS_seven_eight_err]
        
        try:
            os.remove("U.xlsx")
        except: pass
    
        try: 
            os.remove("Uwash.xlsx")
        except: pass
    
        return lstU_cups


    def Thsem_calc(self, Th_file, Thwash_file, lst_ThAge):
        """
        Th sem/cups calc
        """
        self.filename_Th = Th_file
        self.filename_Thwash = Thwash_file
        
        self.zero_nine_corrnorm = lst_ThAge[0]
        self.zero_nine_corrnorm_err = lst_ThAge[1]
        self.two_nine_corrnorm = lst_ThAge[2]
        self.two_nine_corrnorm_err = lst_ThAge[3]
        self.nine_mean = lst_ThAge[4]
        self.nine_counts= lst_ThAge[5]
        
        """
        Machine blanks
        """
        
        int_time_230 = 1.049 #230
        int_time_other = 0.262 #229 and 232
        
        #229
        working_nine_mb = isofilter(self.filename_Thwash, "C")
        nine_mb = working_nine_mb.getMean()
        nine_mb_err = max( (2 * working_nine_mb.getStanddev()/np.sqrt(working_nine_mb.getCounts())), 
                           (nine_mb * 2 / np.sqrt(working_nine_mb.getCounts() * int_time_other * nine_mb)) )
        
        #230
        working_zero_mb = isofilter(self.filename_Thwash, "D")
        zero_mb = working_zero_mb.getMean()
        zero_mb_err = max( (2 * working_zero_mb.getStanddev()/np.sqrt(working_zero_mb.getCounts())), 
                           (zero_mb * 2 / np.sqrt(working_zero_mb.getCounts() * int_time_230 * zero_mb)) )
        
        #232
        working_two_mb = isofilter(self.filename_Thwash, "E")
        two_mb = working_two_mb.getMean()
        two_mb_err = max( (2 * working_two_mb.getStanddev()/np.sqrt(working_two_mb.getCounts())), 
                           (two_mb * 2 / np.sqrt(working_two_mb.getCounts() * int_time_other * two_mb)) )
        
        """
        Machine blank correction
        """
        
        #229/229 machine blank errors
        nine_nine_err = nine_mb / self.nine_mean
        
        #230/229 machine blank errors
        zero_nine_err1 = zero_mb/ self.nine_mean
        zero_nine_err2 = zero_nine_err1 * (zero_mb_err/zero_mb)
        
        #232/229 machine blank errors
        two_nine_err1 = two_mb / self.nine_mean
        two_nine_err2 = two_nine_err1 * (two_mb_err/two_mb)
        
        #230/229 machine blank corrected 
        zero_nine_mb_corrected = (self.zero_nine_corrnorm - zero_nine_err1)/(1 - nine_nine_err)
        zero_nine_mb_corrected_err = np.sqrt(self.zero_nine_corrnorm_err**2 + zero_nine_err2**2)
        
        #232/229 machine blank corrected
        two_nine_mb_corrected = (self.two_nine_corrnorm - two_nine_err1)/(1 - nine_nine_err)
        two_nine_mb_corrected_err = np.sqrt(self.two_nine_corrnorm_err**2 + two_nine_err2**2)
        
        """
        Spike correction
        """
        #230/229 spike corrected
        zero_nine_spike_corrected = zero_nine_mb_corrected - self.spike_zero_nine
        zero_nine_spike_corrected_err = np.sqrt(zero_nine_mb_corrected_err**2 + self.spike_zero_nine_err**2)
        
        #232/229 spike corrected
        two_nine_spike_corrected = two_nine_mb_corrected - self.spike_nine_two
        two_nine_spike_corrected_err = np.sqrt(two_nine_mb_corrected_err**2 + self.spike_nine_two_err**2)
        
        """
        Chem blank correction
        """
        
        #230/229 chem blank corrected
        zero_nine_chem_corrected = zero_nine_spike_corrected - self.chemblank_zero_nine
        zero_nine_chem_corrected_err = np.sqrt(zero_nine_spike_corrected_err**2 + self.chemblank_zero_nine_err**2)
        zero_nine_chem_corrected_relerr = zero_nine_chem_corrected_err/zero_nine_chem_corrected
        
        #232/229 chem blank corrected
        two_nine_chem_corrected = two_nine_spike_corrected - self.chemblank_two_nine
        two_nine_chem_corrected_err = np.sqrt(two_nine_spike_corrected_err**2 + self.chemblank_two_nine_err**2)
        two_nine_chem_corrected_relerr = two_nine_chem_corrected_err/two_nine_chem_corrected
        
        #230/232 chem blank corrected
        zero_two_chem_corrected = zero_nine_chem_corrected / two_nine_chem_corrected
        zero_two_chem_corrected_relerr = np.sqrt(zero_nine_chem_corrected_relerr**2 + two_nine_chem_corrected_relerr**2)
        zero_two_chem_corrected_err = zero_two_chem_corrected * zero_two_chem_corrected_relerr
        
        """
        Final Th calculations
        """
        #230 
        zero_fmolg = (zero_nine_chem_corrected * (self.spike_wt * self.spike_nine/(10.**12)) / self.sample_wt)
        zero_fmolg_err = abs(zero_fmolg * np.sqrt(zero_nine_chem_corrected_relerr**2 + (self.spike_nine_err/self.spike_nine)**2))
        zero_fmolg_wt_err = abs(zero_fmolg * np.sqrt((zero_fmolg_err/zero_fmolg) **2 +
                                                  + (self.sample_wt_err/self.sample_wt)**2 + (self.spike_wt_err/self.spike_wt)**2 ))
        zero_ppt = zero_fmolg * self.wt_230 * (10.**12)
        zero_ppt_err = abs(zero_ppt *  (zero_fmolg_err/zero_fmolg))
        zero_ppt_wt_err = abs(zero_ppt * (zero_fmolg_wt_err / zero_fmolg))
        
        #232 ppt
        two_pmolg = (two_nine_chem_corrected * (self.spike_wt * self.spike_nine/(10.**12)) / self.sample_wt)
        two_pmolg_err = abs(two_pmolg * np.sqrt(two_nine_chem_corrected_relerr**2 + (self.spike_nine_err/self.spike_nine)**2))
        two_pmolg_wt_err = abs(two_pmolg * np.sqrt( (two_pmolg_err/two_pmolg)**2 + 
                                                   (self.sample_wt_err/self.sample_wt)**2 + (self.spike_wt_err/self.spike_wt)**2))
        two_ppt = two_pmolg * self.wt_232 * (10.**12)
        two_ppt_err = abs(two_ppt * (two_pmolg_err/two_pmolg))
        two_ppt_wt_err = abs(two_ppt * (two_pmolg_wt_err/two_pmolg))
        
        #230/232 ratio
        zero_two_final = zero_two_chem_corrected
        zero_two_final_err = abs(zero_two_chem_corrected_err)
        
        lstTh_Age = [zero_fmolg, zero_ppt, zero_ppt_err, zero_ppt_wt_err, two_pmolg, two_pmolg_err, two_ppt, two_ppt_err, two_ppt_wt_err, zero_two_final, zero_two_final_err]
        
        try:
            os.remove("Th.xlsx")
        except: pass
    
        try: 
            os.remove("Thwash.xlsx")
        except: pass
        
        return lstTh_Age
        

    def Thcups_calc(self, Th_file, Thwash_file, lstUTh, Thwash_option):
        """
        Th cups calc
        """
        self.filename_Th = Th_file
        self.filename_Thwash = Thwash_file
        
        six_three_corr = lstUTh[0]
        six_three_corr_err = lstUTh[1]
        AS = lstUTh[2]
        AS_err = lstUTh[3]
        
        """
        Machine blanks
        """
        int_time = 1.049
        
        int_time_230 = 1.049 #230
        int_time_other = 0.262 #229 and 232
        
        if Thwash_option == "sem":
        
            #229
            working_nine_mb = isofilter(self.filename_Thwash, "C")
            nine_mb = working_nine_mb.getMean()
            nine_mb_err = max( (2 * working_nine_mb.getStanddev()/np.sqrt(working_nine_mb.getCounts())), 
                               (nine_mb * 2 / np.sqrt(working_nine_mb.getCounts() * int_time_other * nine_mb)) )
            
            #230
            working_zero_mb = isofilter(self.filename_Thwash, "D")
            zero_mb = working_zero_mb.getMean()
            zero_mb_err = max( (2 * working_zero_mb.getStanddev()/np.sqrt(working_zero_mb.getCounts())), 
                               (zero_mb * 2 / np.sqrt(working_zero_mb.getCounts() * int_time_230 * zero_mb)) )
            
            #232
            working_two_mb = isofilter(self.filename_Thwash, "E")
            two_mb = working_two_mb.getMean()
            two_mb_err = max( (2 * working_two_mb.getStanddev()/np.sqrt(working_two_mb.getCounts())), 
                               (two_mb * 2 / np.sqrt(working_two_mb.getCounts() * int_time_other * two_mb)) )
        
        elif Thwash_option == "cups":
            
            #229
            working_nine_mb = isofilter(self.filename_Thwash, "C")
            nine_mb = working_nine_mb.getMean() * 62422000
            nine_mb_err = (2 * working_nine_mb.getStanddev()/np.sqrt(working_nine_mb.getCounts())) * 62422000
            
            #230
            working_zero_mb = isofilter(self.filename_Thwash, "D")
            zero_mb = working_zero_mb.getMean() * 62422000
            zero_mb_err = (2 * working_zero_mb.getStanddev()/np.sqrt(working_zero_mb.getCounts())) * 62422000
            
            #232
            working_two_mb = isofilter(self.filename_Thwash, "E")
            two_mb = working_two_mb.getMean() * 62422000
            two_mb_err = (2 * working_two_mb.getStanddev()/np.sqrt(working_two_mb.getCounts())) * 62422000
            
        
        """
        Measured beam intensities
        """
        #229
        working_nine_beam = isofilter(self.filename_Th,"C")
        nine_beam_volts = working_nine_beam.getMean()
        nine_beam_cps = nine_beam_volts/(10**11)/(1.602/(10**19))
        nine_beam_abs_err = 2 * working_nine_beam.getStanddev()/np.sqrt(working_nine_beam.getCounts())
        nine_beam_rel_err = 2 / np.sqrt(nine_beam_cps * working_nine_beam.getCounts() * int_time) * 1000
        
        #230
        working_zero_beam = isofilter(self.filename_Th,"D")
        zero_beam_volts = working_zero_beam.getMean()
        zero_beam_cps = zero_beam_volts/(10**11)/(1.602/(10**19))
        zero_beam_abs_err = 2 * working_zero_beam.getStanddev()/np.sqrt(working_zero_beam.getCounts())
        zero_beam_rel_err = 2 / np.sqrt(zero_beam_cps * working_zero_beam.getCounts() * int_time) * 1000
        
        #232
        working_two_beam = isofilter(self.filename_Th,"E")
        two_beam_volts = working_two_beam.getMean()
        two_beam_cps = two_beam_volts/(10**11)/(1.602/(10**19))
        two_beam_abs_err = 2 * working_two_beam.getStanddev()/np.sqrt(working_two_beam.getCounts())
        two_beam_rel_err = 2 / np.sqrt(two_beam_cps * working_two_beam.getCounts() * int_time) * 1000
        
        """
        Measured ratios and corrected errors and arrays
        """
        
        #230/229 measured ratio
        zero_nine_working = isofilter(self.filename_Th, "F")
        zero_nine_meas = zero_nine_working.getMean()
        zero_nine_meas_err = max((2 * zero_nine_working.getStanddev() / np.sqrt(zero_nine_working.getCounts())), 
                                          zero_nine_meas * np.sqrt(zero_beam_rel_err**2)/(10**3))
        zero_nine_meas_array = isocorrection().array(self.filename_Th, "F")
        
        #232/229 measured ratio
        two_nine_working = isofilter(self.filename_Th, "G")
        two_nine_meas = two_nine_working.getMean()
        two_nine_meas_err = max((2 * two_nine_working.getStanddev() / np.sqrt(two_nine_working.getCounts())), 
                                          two_nine_meas * np.sqrt(two_beam_rel_err**2)/(10**3))
        two_nine_meas_array = isocorrection().array(self.filename_Th, "G")
        
        #230/232 measured ratio
        zero_two_working = isofilter(self.filename_Th, "H")
        zero_two_meas = zero_two_working.getMean()
        zero_two_meas_err = max((2 * zero_two_working.getStanddev() / np.sqrt(zero_two_working.getCounts())),
                                zero_two_meas * np.sqrt(zero_beam_rel_err**2 + two_beam_rel_err*2)/ (10**3))
        zero_two_meas_array = isocorrection().array(self.filename_Th, "H")
        
        
        """
        Drift correction for 230/229 and 230/232
        """
        
        #230 drift array
    
        zero_array = isocorrection().array(self.filename_Th, "D")
        drift_array = isocorrection().drift_correction_offset(zero_array, zero_nine_meas_array)
        
        #230/229 drift corrected
        nine_array = isocorrection().array(self.filename_Th, "C")
        zero_nine_drift_corrected_array = isocorrection().drift_correction(drift_array, nine_array)
        zero_nine_drift_corrected_err = max( (2 * np.nanstd(a = zero_nine_drift_corrected_array, ddof = 1)/ np.sqrt(len(zero_nine_drift_corrected_array[np.logical_not(np.isnan(zero_nine_drift_corrected_array))]))),
                                             (np.nanmean(zero_nine_drift_corrected_array) * np.sqrt(zero_beam_rel_err**2)/(10**3)) )
        
        
        #230/232 drift corrected
        two_array = isocorrection().array(self.filename_Th, "E")
        zero_two_drift_corrected_array = isocorrection().drift_correction_alt(drift_array, two_array, zero_two_meas_array)
        zero_two_drift_corrected_err = max( (2 * np.nanstd(a = zero_two_drift_corrected_array, ddof = 1)/ np.sqrt(len(zero_two_drift_corrected_array[np.logical_not(np.isnan(zero_two_drift_corrected_array))]))),
                                             (np.nanmean(zero_two_drift_corrected_array) * np.sqrt(zero_beam_rel_err**2 + two_beam_rel_err**2)/(10**3)) )
        
        
        """
        Machine blank correction
        """
        
        #230/229 mb corrected
        zero_nine_mb_corrected_array = isocorrection().machine_blank_correction(zero_nine_drift_corrected_array, 
                                                    nine_beam_volts, nine_mb, zero_mb)
        zero_nine_mb_corrected_err = np.sqrt(zero_nine_drift_corrected_err**2 + (zero_mb_err/nine_beam_cps)**2)
        
        
        #232/229 mb corrected
        two_nine_mb_corrected_array = isocorrection().machine_blank_correction(two_nine_meas_array, 
                                                    nine_beam_volts, nine_mb, two_mb)
        two_nine_mb_corrected_err = np.sqrt(two_nine_meas_err**2 + (two_mb_err/nine_beam_cps)**2)
        
        #230/232 mb corrected
        zero_two_mb_corrected_array = isocorrection().machine_blank_correction_alt(zero_two_drift_corrected_array, 
                                                   nine_beam_cps, two_mb, zero_mb, two_nine_meas)
        zero_two_mb_corrected_err = np.sqrt(zero_two_drift_corrected_err**2 + ((zero_mb_err/nine_beam_cps)/two_nine_meas)**2 
                                            + ((two_mb_err/nine_beam_cps)/two_nine_meas)**2)
        
        
        """
        Tail correction
        """
        #230/229 tail ratios ppm
        zero_nine_tail = AS * (10.**6)
        zero_nine_tail_err = max((zero_nine_tail * (AS_err/AS)), (0.5 * zero_nine_tail))
        
        #232/229 tail ratios ppm
        two_nine_tail = AS * self.five_seven_tail * (10.**6)
        two_nine_tail_err = max((two_nine_tail * np.sqrt((self.five_seven_tail_err/self.five_seven_tail)**2 + (AS_err/AS)**2)), 
                                (0.5 * two_nine_tail))
        
        #230/232 tail ratios ppm
        zero_two_tail = AS * self.six_seven_tail * (10.**6)
        zero_two_tail_err = max((zero_two_tail * np.sqrt((self.six_seven_tail_err/self.six_seven_tail)**2 + (AS_err/AS)**2)), 
                                (0.5 * zero_two_tail))
        
        #229/232 tail ratios ppm
        nine_two_tail = AS * self.five_seven_tail * (10.**6)
        nine_two_tail_err = max((nine_two_tail * np.sqrt((self.six_seven_tail_err/self.six_seven_tail)**2 + (AS_err/AS)**2)),
                                (0.5 * nine_two_tail))
        
        #232/229 machine blank corrected mean
        two_nine_mb_corr = np.nanmean(two_nine_mb_corrected_array)
        
        #230/229 tail corrected
        zero_nine_tail_corrected_array = isocorrection().tail_correction_th(zero_nine_mb_corrected_array, zero_nine_tail, two_nine_tail, 
                                                      zero_two_tail, nine_two_tail, two_nine_mb_corr, '230/229')
        zero_nine_tail_corrected_err = np.sqrt(zero_nine_mb_corrected_err**2 + (zero_nine_tail_err/(10**6))**2 + ((zero_two_tail_err/(10**6)) * two_nine_mb_corr)**2)
        
        #232/229 tail corrected
        two_nine_tail_corrected_array = isocorrection().tail_correction_th(two_nine_mb_corrected_array, zero_nine_tail, two_nine_tail, 
                                                      zero_two_tail, nine_two_tail, two_nine_mb_corr, '232/229')
        two_nine_tail_corrected_err = np.sqrt(two_nine_mb_corrected_err**2 + (two_nine_tail_err/(10**6))**2)
        
        #230/232 tail corrected
        zero_two_tail_corrected_array = isocorrection().tail_correction_th(zero_two_mb_corrected_array, zero_nine_tail, two_nine_tail, 
                                                      zero_two_tail, nine_two_tail, two_nine_mb_corr, '230/232')
        zero_two_tail_corrected_err = np.sqrt(zero_two_mb_corrected_err**2 + ((zero_nine_tail_err/(10**6))/two_nine_mb_corr)**2 + 
                                              (zero_two_tail_err/(10**6))**2 + ((two_nine_tail_err/(10**6))/two_nine_mb_corr)**2)
        
        """
        Fractionation correction
        """
        #230/229 fractionation corrected
        zero_nine_fract_corrected_array = isocorrection().fractionation_correction(zero_nine_tail_corrected_array, 
                                                       six_three_corr, '230', '229', self.spike_six_three)
        zero_nine_fract_corrected_mean = np.nanmean(zero_nine_fract_corrected_array)
        zero_nine_fract_corrected_err = np.sqrt((zero_nine_tail_corrected_err/np.nanmean(zero_nine_tail_corrected_array))**2 + 
                                                (1 * (six_three_corr_err/six_three_corr)/3)**2) * zero_nine_fract_corrected_mean
        
        #232/229 fractionation corrected
        two_nine_fract_corrected_array = isocorrection().fractionation_correction(two_nine_tail_corrected_array, 
                                                       six_three_corr, '232', '229', self.spike_six_three)
        two_nine_fract_corrected_mean = np.nanmean(two_nine_fract_corrected_array)
        two_nine_fract_corrected_err = np.sqrt((two_nine_tail_corrected_err/np.nanmean(two_nine_tail_corrected_array))**2 + 
                                                (3 * (six_three_corr_err/six_three_corr)/3)**2) * two_nine_fract_corrected_mean
                                               
        #230/232 fractionation corrected
        zero_two_fract_corrected_array = isocorrection().fractionation_correction(zero_two_tail_corrected_array, 
                                                       six_three_corr, '230', '232', self.spike_six_three)
        zero_two_fract_corrected_mean = np.nanmean(zero_two_fract_corrected_array)
        zero_two_fract_corrected_err = np.sqrt((zero_two_tail_corrected_err/np.nanmean(zero_two_tail_corrected_array))**2 + 
                                                (2 * (six_three_corr_err/six_three_corr)/3)**2) * zero_two_fract_corrected_mean
        
        """
        Spike correction
        """
        #230/229 spike corrected
        zero_nine_spike_corrected_mean = zero_nine_fract_corrected_mean - self.spike_zero_nine
        zero_nine_spike_corrected_err = np.sqrt(zero_nine_fract_corrected_err**2 + self.spike_zero_nine_err**2)
        
        #232/229 spike corrected
        two_nine_spike_corrected_mean = two_nine_fract_corrected_mean - self.spike_nine_two
        two_nine_spike_corrected_err = np.sqrt(two_nine_fract_corrected_err**2 + self.spike_nine_two_err**2)
        
        #230/232 spike corrected
        zero_two_spike_corrected_mean = (zero_two_fract_corrected_mean - self.spike_zero_nine/two_nine_fract_corrected_mean) / (1 - self.spike_nine_two/two_nine_fract_corrected_mean)
        zero_two_spike_corrected_err = np.sqrt(zero_two_fract_corrected_err**2 + (self.spike_zero_nine_err/two_nine_fract_corrected_mean)**2 
                                               + (self.spike_nine_two_err/two_nine_fract_corrected_mean)**2)
        
        """
        Chemistry blank corrections
        """
        #230/229 chem blank corrected
        zero_nine_chem_corrected_mean = zero_nine_spike_corrected_mean - self.chemblank_zero_nine
        zero_nine_chem_corrected_err = np.sqrt(zero_nine_spike_corrected_err**2 + self.chemblank_zero_nine_err**2)
        zero_nine_chem_corrected_relerr = zero_nine_chem_corrected_err/zero_nine_chem_corrected_mean
        
        #232/229 chem blank corrected
        two_nine_chem_corrected_mean = two_nine_spike_corrected_mean - self.chemblank_two_nine
        two_nine_chem_corrected_err = np.sqrt(two_nine_spike_corrected_err**2 + self.chemblank_two_nine_err**2)
        two_nine_chem_corrected_relerr = two_nine_chem_corrected_err/two_nine_chem_corrected_mean
        
        #230/232 chem blank corrected
        zero_two_chem_corrected_mean = (zero_two_spike_corrected_mean - self.chemblank_zero_nine/two_nine_spike_corrected_mean) / (1 - self.chemblank_two_nine/two_nine_spike_corrected_mean)
        zero_two_chem_corrected_err = np.sqrt(zero_two_spike_corrected_err**2 + (self.chemblank_two_nine_err/two_nine_spike_corrected_mean)**2 
                                              + (self.chemblank_zero_nine_err/two_nine_spike_corrected_mean)**2)
        zero_two_chem_corrected_relerr = zero_two_chem_corrected_err/zero_two_chem_corrected_mean
        
        
        """
        Final Th calculations
        """
        
        #230 ppt
        zero_fmolg = (zero_nine_chem_corrected_mean * (self.spike_wt * self.spike_nine/(10.**12)) / self.sample_wt)
        zero_fmolg_err = abs(zero_fmolg * np.sqrt(zero_nine_chem_corrected_relerr**2 + (self.spike_nine_err/self.spike_nine)**2))
        zero_fmolg_wt_err = abs(zero_fmolg * np.sqrt((zero_fmolg_err/zero_fmolg) **2 +
                                                  + (self.sample_wt_err/self.sample_wt)**2 + (self.spike_wt_err/self.spike_wt)**2 ))
        zero_ppt = zero_fmolg * self.wt_230 * (10.**12)
        zero_ppt_err = abs(zero_ppt *  (zero_fmolg_err/zero_fmolg))
        zero_ppt_wt_err = abs(zero_ppt * (zero_fmolg_wt_err / zero_fmolg))
                          
                       
        
        #232 ppt
        two_pmolg = (two_nine_chem_corrected_mean * (self.spike_wt * self.spike_nine/(10.**12)) / self.sample_wt)
        two_pmolg_err = abs(two_pmolg * np.sqrt(two_nine_chem_corrected_relerr**2 + (self.spike_nine_err/self.spike_nine)**2))
        two_pmolg_wt_err = abs(two_pmolg * np.sqrt( (two_pmolg_err/two_pmolg)**2 + 
                                                   (self.sample_wt_err/self.sample_wt)**2 + (self.spike_wt_err/self.spike_wt)**2))
        two_ppt = two_pmolg * self.wt_232 * (10.**12)
        two_ppt_err = abs(two_ppt * (two_pmolg_err/two_pmolg))
        two_ppt_wt_err = abs(two_ppt * (two_pmolg_wt_err/two_pmolg))
        
        
        #230/232 ratio
        zero_two_final = zero_two_chem_corrected_mean
        zero_two_final_err = abs(zero_two_chem_corrected_err)
        
        lstTh_Age = [zero_fmolg, zero_ppt, zero_ppt_err, zero_ppt_wt_err, two_pmolg, two_pmolg_err, two_ppt, two_ppt_err, two_ppt_wt_err, zero_two_final, zero_two_final_err]
        
        try:
            os.remove("Th.xlsx")
        except: pass
    
        try: 
            os.remove("Thwash.xlsx")
        except: pass
        
        return lstTh_Age

class isocorrection():
    """
    Class for creating numpy arrays of Excel columns and completing element-wise corrections
    """
    
    def __init__(self):
        
        """
        Init def, no inputs needed
        """
        
    def array(self, filename, columnletter):
        """
        Code provides output array for excel row. Includes NaN for non-values
        """
        
        self.column = str(columnletter)+'{}:'+str(columnletter)+'{}'
        self.filename = str(filename)
        self.workbook = openpyxl.load_workbook(self.filename, data_only = True)
        self.ws = self.workbook.worksheets[0]
        
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value:
                    try: 
                        value = float(cell.value)
                        outlist.append(value)
                    except: 
                        outlist.append(np.nan)
                elif cell.value == 0:
                    value= 0.00
                    outlist.append(value)
                else: outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        return outarray
        
    def drift_correction_offset(self, source_array, ratio_array):
        """
        Code provides offset array for given isotope and isotope ratio. Includes NaN for non-values
        """
        int_time = 1.049
        time_constant = 0.65
        
        offset_list = []
        
        for i in range(0, len(source_array) - 1):
            
            if i == 0: offset_list.append(np.nan)
            
            else:
                if np.logical_not(np.isnan(ratio_array[i-1])) and np.logical_not(np.isnan(ratio_array[i])) and np.logical_not(np.isnan(ratio_array[i+1])): 
                    offset = source_array[i] + ((source_array[i+1] - source_array[i-1])/(2*int_time)) * time_constant
                    offset_list.append(offset)  
                else:
                    offset_list.append(np.nan)
            
        offset_array = np.array(offset_list, dtype = np.float)
        return offset_array

    def drift_correction(self, drift_array, source_array):
        """
        Code provides array corrected for drift for a given isotope ratio. Includes NaN for non-values
        """
        driftcorrected_list = []
        
        for i in range(0, len(drift_array)):
            
            if np.logical_not(np.isnan(drift_array[i])):
                drift_corrected = drift_array[i] / source_array[i]
                driftcorrected_list.append(drift_corrected)
            else:
                driftcorrected_list.append(np.nan)
          
        driftcorrected_array = np.array(driftcorrected_list, dtype = np.float)
        return driftcorrected_array
    
    def drift_correction_alt(self, drift_array, source_array_main, source_array_ref):
        """
        Code provides array corrected for drift for a given isotope ratio. Includes NaN for non-values
        """
        driftcorrected_list = []
        
        for i in range(0, len(drift_array)):
            
            if np.logical_not(np.isnan(drift_array[i])) and np.logical_not(np.isnan(source_array_main[i])) and np.logical_not(np.isnan(source_array_ref[i-1])) and np.logical_not(np.isnan(source_array_ref[i])) and np.logical_not(np.isnan(source_array_ref[i+1])):
                drift_corrected = drift_array[i] / source_array_main[i]
                driftcorrected_list.append(drift_corrected)
            else:
                driftcorrected_list.append(np.nan)
          
        driftcorrected_array = np.array(driftcorrected_list, dtype = np.float)
        return driftcorrected_array
            
    def machine_blank_correction(self, source_array, bottom_mean, machine_blank_bottom_mean, machine_blank_top_mean):
        """
        Code provides a machine blank corrected array for a given isotope ratio. Includes NaN for non-values
        """
        mbcorrected_list = []
        
        for i in range(0, len(source_array)):
            
            if np.logical_not(np.isnan(source_array[i])):
                mbcorrected = (source_array[i] - machine_blank_top_mean/(bottom_mean * 62422000.))/(1 - machine_blank_bottom_mean/(bottom_mean * 62422000.))
                mbcorrected_list.append(mbcorrected)
            else: 
                mbcorrected_list.append(np.nan)
                
        mbcorrected_array = np.array(mbcorrected_list, dtype = np.float)
        return mbcorrected_array
    
    def machine_blank_correction_alt(self, source_array, bottom_mean, machine_blank_bottom_mean, machine_blank_top_mean, ratio_array_mean):
        """
        Code provides a machine blank corrected array for a given isotope ratio. Includes NaN for non-values
        """
        mbcorrected_list = []
        
        for i in range(0, len(source_array)):
            
            if np.logical_not(np.isnan(source_array[i])):
                mbcorrected = (source_array[i] - machine_blank_top_mean/(bottom_mean * ratio_array_mean))/(1 - machine_blank_bottom_mean/(bottom_mean * ratio_array_mean))
                mbcorrected_list.append(mbcorrected)
            else: 
                mbcorrected_list.append(np.nan)
                
        mbcorrected_array = np.array(mbcorrected_list, dtype = np.float)
        return mbcorrected_array
    
    def tail_correction(self, source_array, tail_top, tail_bottom, eight_three_mb_corr, option):
        """
        Code provides a tail corrected array for a given isotope ratio. This tail correction is applied to 234/233, 
        235/233, and 236/233, with slight variation for 238/233. Includes NaN for non-values
        """
        
        tailcorrected_list = []
        
        if str(option) == 'norm':
            for i in range(0, len(source_array)):
                
                if np.logical_not(np.isnan(source_array[i])):
                    tailcorrected = (source_array[i] - tail_top/(10**6) * eight_three_mb_corr)/(1 - tail_bottom/(10**6) * eight_three_mb_corr)
                    tailcorrected_list.append(tailcorrected)
                else:
                    tailcorrected_list.append(np.nan)
                    
        elif str(option) == '238/233':
            for i in range(0, len(source_array)):
                
                if np.logical_not(np.isnan(source_array[i])):
                    tailcorrected = source_array[i]/(1 - (tail_bottom/(10**6)) * eight_three_mb_corr)
                    tailcorrected_list.append(tailcorrected)
                else:
                    tailcorrected_list.append(np.nan)
        
        tailcorrected_array = np.array(tailcorrected_list, dtype = np.float)
        return tailcorrected_array

    def tail_correction_alt(self, top_array_corr, bottom_array_corr):
        """
        Code provides a tail corrected array for a given isotope ratio. This tail correction is applied to 238/235
        and 234/238. Includes NaN for non-values
        """
        
        tailcorrected_list = []
        
        for i in range(0, len(top_array_corr)):
            if np.logical_not(np.isnan(top_array_corr[i])) and np.logical_not(np.isnan(bottom_array_corr[i])): 
                tailcorrected = top_array_corr[i] / bottom_array_corr[i]
                tailcorrected_list.append(tailcorrected) 
            else:
                tailcorrected_list.append(np.nan)
        
        tailcorrected_array = np.array(tailcorrected_list, dtype = np.float)
        return tailcorrected_array
    
    def tail_correction_th(self, source_array, zero_nine_tail, two_nine_tail, zero_two_tail, nine_two_tail, ratio_mb, option):
        """
        Code provides a tail corrected array for a given isotope ratio for Th. Includes NaN for non-values
        """
        
        tailcorrected_list = []
        
        if str(option) == '230/229':
            for i in range(0, len(source_array)):
                if np.logical_not(np.isnan(source_array[i])):
                    tailcorrected = (source_array[i] - (zero_nine_tail/(10**6)) - ((zero_two_tail/(10**6)) * ratio_mb)) / (1 - ((nine_two_tail/(10**6)) * ratio_mb))
                    tailcorrected_list.append(tailcorrected)
                else:
                    tailcorrected_list.append(np.nan)
                    
        elif str(option) == '232/229':
            for i in range(0, len(source_array)):
                if np.logical_not(np.isnan(source_array[i])):
                    tailcorrected = (source_array[i] - (two_nine_tail/(10**6))) / (1 - ((nine_two_tail/(10**6)) * ratio_mb))
                    tailcorrected_list.append(tailcorrected)
                else:
                    tailcorrected_list.append(np.nan)
                    
        elif str(option) == '230/232':
            for i in range(0, len(source_array)):
                if np.logical_not(np.isnan(source_array[i])):
                    tailcorrected = (source_array[i] - ((zero_nine_tail/(10**6))/ratio_mb))/(1 - (zero_two_tail/(10**6)))
                    tailcorrected_list.append(tailcorrected)
                else:
                    tailcorrected_list.append(np.nan)
        
        tailcorrected_array = np.array(tailcorrected_list, dtype = np.float)
        return tailcorrected_array
        
    
    def fractionation_correction(self, source_array, six_three_tail_corr, top_mass, bottom_mass, spike_six_three):
        """
        Code provides a fractionation corrected array for a given isotope ratio. Includes NaN for non-values
        """
        
        wt_236 = 236.045563
        wt_233 = 233.039629
        
        if top_mass == '234': wt_top = 234.040947
        elif top_mass == '235': wt_top = 235.043924
        elif top_mass == '236': wt_top = 236.045563
        elif top_mass == '238': wt_top = 238.050785
        elif top_mass == '230': wt_top = 230.033128
        elif top_mass == '232': wt_top = 232.038051
        
        if bottom_mass == '233': wt_bottom = 233.039629
        elif bottom_mass == '235': wt_bottom = 235.043924
        elif bottom_mass == '238': wt_bottom = 238.050785
        elif bottom_mass == '229': wt_bottom = 229.031756
        elif bottom_mass == '232': wt_bottom = 232.038051
            
        fractcorrected_list = []
        
        for i in range(0, len(source_array)):
            if np.logical_not(np.isnan(source_array[i])):
                fractcorrected = source_array[i] * (spike_six_three/six_three_tail_corr)**(np.log(wt_top/wt_bottom)/np.log(wt_236/wt_233))
                fractcorrected_list.append(fractcorrected)
            else: 
                fractcorrected_list.append(np.nan)
        fractcorrected_array = np.array(fractcorrected_list, dtype = np.float)
        return fractcorrected_array


class Ucalculation():
    """
    Class Ucalculation functions as the U sheet in the age calculation spreadsheet. Ucalculation gives outputs for 
    both the Thcalculation function and the Agecalculation function. 
    
    U_normalized_forTh output is a list of the following values: 
        [0]: 236/233 corrected ratio
        [1]: 236/233 corrected ratio error
    
    U_normalized_forAge output is a list of the following values: 
        [0]: 235/233 normalized ratio
        [1]: 235/233 normalized ratio error
        [2]: 235/234 normalized and corrected ratio
        [3]: 235/234 normalized and corrected ratio error
        [4]: Unfiltered 233 counts
        [5]: Filtered 234/235 counts
        [6]: Unfiltered 233 mean
    """   
    def __init__ (self, spike_input, AS_input, filename_input):
        
        self.spike = spike_input
        self.AS = float(AS_input)
        filename = filename_input
        
        #236/233 filtered measured mean and 2s error
        working = isofilter(filename, "G")
        a = working.getMean()
        b = working.getStanddev()
        c = working.getCounts()
        self.six_three_mean_meas = working.Filtered_mean(a,b,c,44)
        self.six_three_err_meas = working.Filtered_err(a,b,c,44)
        
        #235/233 filtered measured mean and 2s error
        working_b = isofilter(filename, "H")
        a = working_b.getMean()
        b = working_b.getStanddev()
        c = working_b.getCounts()
        self.five_three_mean_meas = working_b.Filtered_mean(a,b,c,44)
        self.five_three_err_meas = working_b.Filtered_err(a,b,c,44) 
    
        #234/235 filtered measured mean and 2s error
        working_c = isofilter(filename,"I")
        a = working_c.getMean()
        b = working_c.getStanddev()
        c = working_c.getCounts()
        self.four_five_mean_meas = working_c.Filtered_mean(a,b,c,44)
        self.four_five_err_meas = working_c.Filtered_err(a,b,c,44) 
        self.four_five_counts = working_c.Filtered_counts(a,b,c,44)
        
        #233 unfiltered mean and counts
        working_d = isofilter(filename, "C")
        self.three_mean_meas = working_d.getMean()
        self.three_counts = working_d.getCounts()
        
        #constants to be used throughout the class
        self.wt_235 = 235.043924
        self.wt_233 = 233.039629
        self.wt_236 = 236.045563
        self.wt_234 = 234.040947
        self.eight_five_rat = 137.83
        self.AS_six_eight = self.AS/5
        self.AS_four_eight = self.AS/20
        self.eight_five_rat_err_rel = 0.0003
        
        #remove excel files
        try: os.remove("U.xlsx")
        except: pass
    
    def U_normalization_forTh(self):
        """
        Function outputs the measured 236/233 ratio and error, the 236/233 ratio and error corrected for the 238 tail, 
        and the 235/233 normalized ratio and error using the 235/233 corrected ratio and further correcting
        235/233 for mass fractionation in the ICP-MS. These values are used later in the Th_normalization function.
        """
        
        #corrects 236/233 ratio for 238 tail 
        self.six_three_corr = self.six_three_mean_meas * ( 1 - (self.AS_six_eight * self.five_three_mean_meas 
                                                                * self.eight_five_rat/self.spike) )
        
        #provides the ratio that will be used to correct for mass fractionation
        rat = float(np.log(self.wt_235/self.wt_233)/np.log(self.wt_236/self.wt_233))
        
        #corrects for mass fractionation in the ICP-MS
        self.five_three_norm = self.five_three_mean_meas * (self.spike/self.six_three_corr)**rat
        
        #provides relative error constants to be used in this function
        AS_six_eight_err_rel = 0.3
        five_three_err_rel = self.five_three_err_meas/self.five_three_mean_meas
        six_three_err_rel = self.six_three_err_meas/self.six_three_mean_meas
       
        #calculculates the 236/233 corrected error
        self.six_three_corr_err = self.six_three_corr * np.sqrt( six_three_err_rel**2 + 
                                                                ( (self.AS_six_eight * self.five_three_mean_meas 
                                                                   * self.eight_five_rat)/self.spike  
                                                                   * np.sqrt( AS_six_eight_err_rel**2 + five_three_err_rel ** 2 + self.eight_five_rat_err_rel**2 ) 
                                                                   / (1 - (self.AS_six_eight * self.five_three_mean_meas * self.eight_five_rat)
                                                                   / self.spike) ) ** 2 ) 
        #calculates the 236/233 relative corrected error
        self.six_three_corr_err_rel = self.six_three_corr_err/self.six_three_corr
        
        #calculates the 235/233 normalized error
        self.five_three_norm_err = self.five_three_norm * np.sqrt( five_three_err_rel**2 
                                                                  + (2 * (self.six_three_corr_err_rel/3))**2  ) 
       
        #a list of your outputs is created and returned, to be used in the Th functions
        lstU_Th = [self.six_three_corr,self.six_three_corr_err ]
        
        return lstU_Th
    
    def U_normalization_forAge(self):
        """
        Function outputs the 235/233 normalized ratio and error, the 235/234 normalized and corrected ratio and error, 
        the unfiltered number of cycles for 233 and mean value and the filtered number of cycles 234/235. These values will be used
        later in the Age Calculation.
        """
        #calculates constants that will be used to calculate normalized 234/235
        four_five_err_rel = self.four_five_err_meas / self.four_five_mean_meas
        
        rat = float(np.log(self.wt_234/self.wt_235)/np.log(self.wt_236/self.wt_233))
        
        #normalizes the 234/235 ratio by correcting for mass fractionation and calculates the resulting error
        self.four_five_norm = self.four_five_mean_meas * (self.spike/self.six_three_corr)**rat
        
        self.four_five_norm_err = self.four_five_norm * np.sqrt( four_five_err_rel**2 + 
                                                                (self.six_three_corr_err_rel/3)**2 )
        
        #calculates constants that will be used to calculated corrected 234/235
        AS_four_eight_err_rel = 0.3
        four_five_norm_err_rel = self.four_five_norm_err/self.four_five_norm
        
        #corrects the normalized 234/235 ratio for 238 tail and calculated the resulting error
        self.four_five_normcorr = self.four_five_norm * (1 - ( self.eight_five_rat 
                                                              * self.AS_four_eight/ self.four_five_norm ))

        self.four_five_normcorr_err = self.four_five_normcorr * np.sqrt( four_five_norm_err_rel**2 + 
                                                                        ( (self.eight_five_rat * self.AS_four_eight / self.four_five_norm) *
                                                                         np.sqrt( self.eight_five_rat_err_rel**2 + AS_four_eight_err_rel**2 + four_five_norm_err_rel**2 )
                                                                         / (1 - ( self.eight_five_rat * self.AS_four_eight/ self.four_five_norm)) ) **2 ) 
        
        self.three_mean_meas = int(self.three_mean_meas)
        
        #a list of your outputs is created and returned, to be used in the Age functions
        lstU_Age = [self.five_three_norm, self.five_three_norm_err, self.four_five_normcorr, 
                    self.four_five_normcorr_err, self.three_counts, self.four_five_counts, self.three_mean_meas] 
       
        return lstU_Age

class Thcalculation():
    """
    Class Thcalculation functions as the Th sheet in the age calculation spreadsheet. Thcalculation gives outputs for 
    the Agecalculation function, and needs to be provided inputs from the Ucalculation class U_normalization_forTh function.
    
    Th_normalization_forAge output is a list of the following values: 
        [0]: 230/229 corrected and normalized ratio
        [1]: 230/229 corrected and normalized ratio error
        [2]: 232/229 corrected and normalized ratio
        [3]: 232/229 corrected and normalized ratio error
        [4]: Unfiltered 229 mean
        [5]: Unfiltered 229 counts
        
    """
    
    def __init__ (self, spike_input, AS_input, filename_input, lstU_Th):
        
        self.spike = spike_input
    
        #AS is the abundant sensitivity 237/238, measured through the AS method on the ICP-MS    
        self.AS = float(AS_input)
        
        #uses the filename given for your Th run
        filename = str(filename_input)
        
        #Compiles the values of the lstU_Th provided by your U_normalization_forTh function
        self.six_three_corr = lstU_Th[0]
        self.six_three_corr_err = lstU_Th[1]
        
        #Note: Hai's macro only filters 230/229 column
        
        #230/232 filtered measured mean and 2s error
        working = isofilter(filename,"G")
        self.zero_two_mean_meas = working.getMean()/1.02
        self.zero_two_counts = working.getCounts()
        self.zero_two_standdev_meas = working.getStanddev()
        self.zero_two_rel_err_meas = (2 * self.zero_two_standdev_meas/(self.zero_two_counts**0.5))/self.zero_two_mean_meas
        self.zero_two_rel_err = max(self.zero_two_rel_err_meas, 0.005)
        self.zero_two_err_meas = self.zero_two_mean_meas * self.zero_two_rel_err
        
        #230/229 filtered measured mean and 2s error
        working_b = isofilter(filename, "E")
        a = working_b.getMean()
        b = working_b.getStanddev()
        c = working_b.getCounts()
        self.zero_nine_mean_meas = working_b.Filtered_mean(a,b,c,28)
        self.zero_nine_err_meas = working_b.Filtered_err(a,b,c,28)
        self.zero_nine_counts = c
        
        #232/229 unfiltered measured mean and 2s error
        working_c = isofilter(filename, "F")
        self.nine_two_mean_meas = working_c.getMean()
        self.two_nine_mean_meas = 1 / (self.nine_two_mean_meas/1.02)
        self.two_nine_counts = working.getCounts()
        self.nine_two_standdev_meas = working_c.getStanddev()
        self.nine_two_rel_err_meas = (2 * self.nine_two_standdev_meas/(self.two_nine_counts**0.5))/self.nine_two_mean_meas
        self.two_nine_rel_err = max(self.nine_two_rel_err_meas, 0.005)
        self.two_nine_err_meas = self.two_nine_mean_meas * self.two_nine_rel_err
        
        #229 unfiltered mean and counts
        working_d = isofilter(filename, "C")
        self.nine_mean_meas = working_d.getMean()
        self.nine_counts = working_d.getCounts()
        
        #230 unfiltered mean
        working_e = isofilter(filename, "D")
        self.zero_mean_meas = working_e.getMean()
        
        #constants to be used throughout the class
        self.wt_233 = 233.039629
        self.wt_236 = 236.045563
        self.wt_229 = 229.031756
        self.wt_230 = 230.033128
        self.wt_232 = 232.038051
        self.AS_zero_nine = self.AS
        self.AS_zero_two = self.AS_zero_nine / 5
        self.AS_two_nine = self.AS_zero_two / 3
        self.eight_five_rat = 137.83
        self.eight_five_rat_err_rel = 0.0003
        self.count_time = 1.049
        
        #remove excel files
        try: os.remove("Th.xlsx")
        except: pass
        
    def Th_normalization_forAge(self):
        
        #corrects the 230/229 and 232/229 ratios for both the 232 and 229 tails
        self.zero_nine_corr = self.zero_nine_mean_meas * (1 - self.AS_zero_two/self.zero_two_mean_meas) * (1 - self.AS_zero_nine)
        
        self.two_nine_corr = self.two_nine_mean_meas * (1 / (1 - (self.AS_two_nine * self.two_nine_mean_meas)))
        
        #constants needed for error calculations
        self.zero_nine_rel_err = self.zero_nine_err_meas/self.zero_nine_mean_meas
        self.AS_zero_two_rel_err = 0.3
        self.AS_zero_nine_rel_err = 0.3
        self.AS_two_nine_rel_err = 0.3
        
        #errors for corrected 230/229 and 232/229 ratios
        self.zero_nine_corr_err = self.zero_nine_corr * ( self.zero_nine_rel_err**2 + 
                                                         ( (self.AS_zero_two/self.zero_two_mean_meas) *  
                                                                (self.AS_zero_two_rel_err**2 + self.zero_two_rel_err**2)**0.5
                                                                 / (1 - self.AS_zero_two/self.zero_two_mean_meas))**2 
                                                         + ( self.AS_zero_nine * self.AS_zero_nine_rel_err/(1 - self.AS_zero_nine) )**2) ** 0.5
       
        self.two_nine_corr_err = self.two_nine_corr * ( self.two_nine_rel_err**2 + 
                                                      ( ((self.AS_two_nine_rel_err**2 + self.two_nine_rel_err**2)**0.5)  
                                                       * (self.AS_two_nine * self.two_nine_mean_meas)/
                                                         (1 - (self.AS_two_nine * self.two_nine_mean_meas))**2) ** 2
                                                       ) ** 0.5
        
        #constant needed for normalization
        
        rat_1 = np.log(self.wt_230/self.wt_229) / np.log(self.wt_236/self.wt_233)
        
        rat_2 = np.log(self.wt_232/self.wt_229) / np.log(self.wt_236/self.wt_233)
        

        #normalizes for corrected 230/229 and 232/229 ratios for mass fractionation
        
        self.zero_nine_corrnorm = self.zero_nine_corr * ((self.spike / self.six_three_corr)**rat_1)
        
        self.two_nine_corrnorm = self.two_nine_corr * ((self.spike / self.six_three_corr)**rat_2)
        
        #constants needed for error calculations
        self.zero_nine_corr_rel_err = self.zero_nine_corr_err / self.zero_nine_corr
        self.six_three_corr_rel_err = self.six_three_corr_err / self.six_three_corr
        self.two_nine_corr_rel_err = self.two_nine_corr_err / self.two_nine_corr
        
        #errors for normalized 230/229 and 232/229 ratios 
        self.zero_nine_corrnorm_err = self.zero_nine_corrnorm * ( self.zero_nine_corr_rel_err**2
                                                                 + ( (self.six_three_corr_rel_err/3)**2 )
                                                                 )**0.5
        
        self.two_nine_corrnorm_err = self.two_nine_corrnorm * ( self.two_nine_corr_rel_err**2
                                                               +  self.six_three_corr_rel_err**2  
                                                               )**0.5
        
        self.nine_mean_meas = int(self.nine_mean_meas)
        
        #finalized 230/229 error
        
        counting_error = 2 / ((self.zero_mean_meas * self.zero_nine_counts * self.count_time)**0.5)
        
        self.zero_nine_final_err = max(self.zero_nine_corrnorm_err, (counting_error * self.zero_nine_corrnorm))
        
        #a list of your outputs is created and returned, to be used in the Age functions
        lstTh_age = [self.zero_nine_corrnorm, self.zero_nine_final_err, self.two_nine_corrnorm,
                     self.two_nine_corrnorm_err, self.nine_mean_meas, self.nine_counts]
        
        return lstTh_age

class background_values():
    
    def __init__(self, U_file, Th_file):
        
        #uses the filename given for your U wash
        self.filename_U = str(U_file)
        
        #uses the filename give for your Th wash
        self.filename_Th = str(Th_file)
        
    def U_wash(self):
        """
        U_wash provides a list the following outputs for the Age Calculation: 
            [0]: 233 unfiltered wash in cps
            [1]: 234 unfiltered wash in cps
            [2]: 235 unfiltered wash in cps
            
        """
        
        wb = openpyxl.load_workbook(self.filename_U)
        ws = wb.worksheets[0]
        
        if ws['C1'].value == '1:232Th':
        
            #233 wash value
            working_a = isofilter(self.filename_U,"D")
            self.three_wash = working_a.getMean()
            
            #234 wash value
            working_b = isofilter(self.filename_U,"E")
            self.four_wash = working_b.getMean()
            
            #235 wash value
            working_c = isofilter(self.filename_U,"F")
            self.five_wash = working_c.getMean()
            
        elif ws['C1'].value == '1:233U':
            
            #233 wash value
            working_a = isofilter(self.filename_U,"C")
            self.three_wash = working_a.getMean()
            
            #234 wash value
            working_b = isofilter(self.filename_U,"D")
            self.four_wash = working_b.getMean()
            
            #235 wash value
            working_c = isofilter(self.filename_U,"E")
            self.five_wash = working_c.getMean()
        
        else: print "Error"
        
        #remove excel files
        try: os.remove("Uwash.xlsx")
        except: pass
        
        lstU_wash = [self.three_wash, self.four_wash, self.five_wash]
        
        return lstU_wash
    
    def Th_wash(self):
        """
        Th_wash provides the following outputs for the Age Calculation: 
            230 unfiltered wash in cpm 
                
        """
        #230 wash value
        working_a = isofilter(self.filename_Th, "D")
        self.zero_wash = working_a.getMean()
        
        #remove excel files
        try: os.remove("Thwash.xlsx")
        except: pass
    
        #calculate "darknoise" value for Age calculation
        self.darknoise = self.zero_wash * 60

        return self.darknoise

class isofilter():
    """
    Class for calculating unfiltered and filtered mean, standdev/error, and counts for an Excel column
    """
    def __init__(self, filename, columnletter): # input filename and columnletter as strings
        self.column = str(columnletter)+'{}:'+str(columnletter)+'{}'
        self.filename = str(filename)
        self.workbook= openpyxl.load_workbook(self.filename, data_only = True)
        self.ws = self.workbook.worksheets[0]
        self.totalCounts = 0
        self.mean = 0 
        self.filteredMean = 0
        self.err = 0
        self.criteria = 0
        self.totalCounts_filt = 0
        self.standdev = 0
        
    def array(self):
        """
        Code provides output array for excel row. Includes NaN for non-values
        """
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value: 
                    try:
                        value = float(cell.value)
                        outlist.append(value)
                    except:
                        outlist.append(np.nan)
                elif cell.value == 0:
                    value= 0.00
                    outlist.append(value)
                else: outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        return outarray
    
    def getMean(self):
        """
        Code works row by row through specified Excel column, and calculates total mean
        """
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value: 
                    try:
                        value = float(cell.value)
                        outlist.append(value)
                    except:
                        outlist.append(np.nan)
                elif cell.value == 0:
                    value= 0.00
                    outlist.append(value)
                else: outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        self.mean = np.nanmean(a = outarray)
        return self.mean
    

    def getStanddev(self):
        """
        Code works row by row through specified Excel column, and calculates standard deviation
        """
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value: 
                    try:
                        value = float(cell.value)
                        outlist.append(value)
                    except:
                        outlist.append(np.nan)
                elif cell.value == 0:
                    value= 0.00
                    outlist.append(value)
                else: outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float) 
        self.standdev = np.nanstd(a = outarray, ddof = 1)
        return self.standdev
    
    def getCounts(self):
        """
        Code works row by row through specified Excel Column, and determines total number of values present (i.e. cycles)
        """
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value: 
                    try:
                        value = float(cell.value)
                        outcounts += 1
                    except:
                        outcounts += 0
                elif cell.value == 0:
                    outcounts +=1

        self.totalCounts = outcounts
        return self.totalCounts
        
    def Filtered_mean(self, mean, standdev, counts, filternumber):
        """
        Code works row by row through specified Excel column, deletes entries that are outside of specified range, 
        and calculates resulting mean
        """
        self.filternumber = filternumber
        self.mean = mean
        self.standdev = standdev
        self.totalCounts = counts
        self.standerr = (self.standdev / (self.totalCounts**0.5))
        self.criteria = self.filternumber * self.standerr
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value: 
                    try:
                        value = float(cell.value)
                        if abs(value - self.mean) > self.criteria:
                            outlist.append(np.nan)
                        else:
                            outlist.append(value)
                    except: outlist.append(np.nan)
                elif cell.value == 0:
                    value = 0.00
                    if abs(value - self.mean) > self.criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                else: outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        self.filteredMean = np.nanmean(a = outarray)
        return self.filteredMean
     
    def Filtered_err(self, mean, standdev, counts, filternumber):
        """
        Code works row by row through specified Excel column, deletes entries that are outside of specified range, 
        and calculates resulting 2s counting stantistics error
        """
        self.filternumber = filternumber
        self.mean = mean
        self.standdev = standdev
        self.totalCounts = counts
        self.standerr = (self.standdev / (self.totalCounts**0.5))
        self.criteria = self.filternumber * self.standerr
        outlist = []
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value: 
                    try:
                        value = float(cell.value)
                        if abs(value - self.mean) > self.criteria:
                            outlist.append(np.nan)
                        else:
                            outlist.append(value)
                            outcounts += 1
                    except: outlist.append(np.nan)
                elif cell.value == 0:
                    value = 0.00
                    if abs(value - self.mean) > self.criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                        outcounts += 1 
                else: outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        outstanddev = np.nanstd(a = outarray, ddof = 1)
        self.err = 2 * (outstanddev / (outcounts ** 0.5))
        return self.err
        
    def Filtered_counts(self, mean, standdev, counts, filternumber):
        """
        Code works row by row through specified Excel column, deletes entries that are outside of specified range, 
        and determines total number of values remaining (i.e. filtered cycles)
        """
        self.filternumber = filternumber
        self.mean = mean
        self.standdev = standdev
        self.totalCounts = counts
        self.standerr = (self.standdev / (self.totalCounts**0.5))
        self.criteria = self.filternumber * self.standerr
        outlist = []
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value: 
                    try:
                        value = float(cell.value)
                        if abs(value - self.mean) > self.criteria:
                            outlist.append(np.nan)
                        else:
                            outlist.append(value)
                            outcounts += 1
                    except: outlist.append(np.nan)
                elif cell.value == 0:
                    value = 0.00
                    if abs(value - self.mean) > self.criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                        outcounts += 1
                else: outlist.append(np.nan)
        self.totalCounts_filt = outcounts
        return self.totalCounts_filt

class plot_figure(tk.Tk):
    """
    Provides plot of 234U and 230Th beam stability
    """
    
    def __init__(self, U, Uindex, Th, Thindex):
        """
        Init def
        """
        self.U = U
        self.Uindex = Uindex
        self.Th = Th
        self.Thindex = Thindex
      
    def plot_fig(self): 
        """
        plot of 234U beam
        """
        toplevel = tk.Toplevel()
        toplevel.title("Beam intensity")
        fig = Figure(figsize = (8, 6))
        
        ax1 = fig.add_subplot(2,1,1)
        x1 = self.Uindex
        y1 = self.U
        y1mean = np.nanmean(y1)
        y1standdev = np.nanstd(a = y1, ddof = 1)
        ax1.scatter(x1, y1, color = 'b', marker = 'o')
        ax1.set_xlabel('Cycles', fontsize = 7, labelpad = 0.5)
        ax1.set_ylabel('234U', fontsize = 7, labelpad = 0.5)
        ax1.set_ylim([y1mean - 10*y1standdev, y1mean + 10*y1standdev])
        ax1.tick_params(labelsize = 5)
        ax1.set_title('Beam Intensity 234U' , fontsize = 8)
        
        ax2 = fig.add_subplot(2,1,2)
        x2 = self.Thindex
        y2 = self.Th
        y2mean = np.nanmean(y2)
        y2standdev = np.nanstd(a = y2, ddof = 1)
        ax2.scatter(x2, y2, color = 'c', marker = 'o')
        ax2.set_xlabel('Cycles', fontsize = 7, labelpad = 0.5)
        ax2.set_ylabel('230Th', fontsize = 7, labelpad = 0.5)
        ax2.set_ylim([y2mean - 10*y2standdev, y2mean + 10*y2standdev])
        ax2.tick_params(labelsize = 5)
        ax2.set_title('Beam Intensity 230Th', fontsize = 8)
        
        fig.set_tight_layout(True)
        
        canvas = FigureCanvasTkAgg(fig, master = toplevel)
        canvas.show()
        canvas.get_tk_widget().pack()
        toplevel.mainloop()        


class Application_preset(tk.Toplevel):
    
    def __init__(self, original):
        #tk.Toplevel.__init__(self)
        self.original_frame = original
        self.otherframe = tk.Toplevel()
        self.otherframe.protocol("WM_DELETE_WINDOW", on_closing)
        self.otherframe.title("Preset values for Age Calculation")
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        tk.Label(dialog_frame, text = "Options for changing preset values", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        
        
        self.spike_conc_option()
        
    def spike_conc_option(self):
        
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = "Would you like to change the 233U concentration of your spike?", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        
        self.CheckVar_spike_yes = tk.IntVar()
        self.CheckVar_spike_yes.set(0)
        self.CheckVar_spike_yes = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_spike_yes, command = self.spike_yes).grid(row = 0, column = 1, sticky = 'w')
        
        self.CheckVar_spike_no = tk.IntVar()
        self.CheckVar_spike_no.set(0)
        self.CheckVar_spike_no = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_spike_no, command = self.spike_no).grid(row = 0, column = 2, sticky = 'e')
        
    def spike_yes(self):
        
        self.spike_yes = 1
        
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = "Enter 233U concentration in pmol/g:", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        self.spike_conc_three = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spike_conc_three.grid(row = 0, column = 1, sticky = 'w')
        self.spike_conc_three.focus_set()
        
        tk.Label(dialog_frame, text = "Would you like to change the sample wt error from 0.1 mg? ", font = ('TkDefaultFont', 10)).grid(row = 1, column = 0, sticky = 'w')
        
        self.CheckVar_samplewt_yes = tk.IntVar()
        self.CheckVar_samplewt_yes.set(0)
        self.CheckVar_samplewt_yes = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_samplewt_yes, command = self.samplewt_yes).grid(row = 1, column = 1, sticky = 'w')
        
        self.CheckVar_samplewt_no = tk.IntVar()
        self.CheckVar_samplewt_no.set(0)
        self.CheckVar_samplewt_no = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_samplewt_no, command = self.samplewt_no).grid(row = 1, column = 1, sticky = 'e')
        
    def spike_no(self):
        
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = "Would you like to change the sample wt error from 0.1 mg? ", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        
        self.CheckVar_samplewt_yes = tk.IntVar()
        self.CheckVar_samplewt_yes.set(0)
        self.CheckVar_samplewt_yes = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_samplewt_yes, command = self.samplewt_yes).grid(row = 0, column = 1, sticky = 'w')
        
        self.CheckVar_samplewt_no = tk.IntVar()
        self.CheckVar_samplewt_no.set(0)
        self.CheckVar_samplewt_no = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_samplewt_no, command = self.samplewt_no).grid(row = 0, column = 2, sticky = 'w')
        
    def samplewt_yes(self):
        
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        
        self.samplewt_yes = 1
        
        tk.Label(dialog_frame, text = "Sample wt error in mg: ", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        self.sample_wt_err = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.sample_wt_err.grid(row = 0, column = 1, sticky = 'w')
        self.sample_wt_err.focus_set()
        
        tk.Label(dialog_frame, text = "Would you like to change the spike wt error from 0.3 mg? ", font = ('TkDefaultFont', 10)).grid(row = 1, column = 0, sticky = 'w')
        
        self.CheckVar_spikewt_yes = tk.IntVar()
        self.CheckVar_spikewt_yes.set(0)
        self.CheckVar_spikewt_yes = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_spikewt_yes, command = self.spikewt_yes).grid(row = 1, column = 1, sticky = 'w')
        
        self.CheckVar_spikewt_no = tk.IntVar()
        self.CheckVar_spikewt_no.set(0)
        self.CheckVar_spikewt_no = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_spikewt_no, command = self.spikewt_no).grid(row = 1, column = 1, sticky = 'e')
    
    def samplewt_no(self):
        
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = "Would you like to change the spike wt error from 0.3 mg? ", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        
        self.CheckVar_spikewt_yes = tk.IntVar()
        self.CheckVar_spikewt_yes.set(0)
        self.CheckVar_spikewt_yes = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_spikewt_yes, command = self.spikewt_yes).grid(row = 0, column = 1, sticky = 'w')
        
        self.CheckVar_spikewt_no = tk.IntVar()
        self.CheckVar_spikewt_no.set(0)
        self.CheckVar_spikewt_no = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_spikewt_no, command = self.spikewt_no).grid(row = 0, column = 2, sticky = 'e')
    
    
    def spikewt_yes(self):
        
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        
        self.spikewt_yes = 1
        
        tk.Label(dialog_frame, text = "Spike wt error in mg: ", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        self.spike_wt_err = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.spike_wt_err.grid(row = 0, column = 1, sticky = 'w')
        self.spike_wt_err.focus_set()
        
        tk.Label(dialog_frame, text = "Would you like to alter the 230/232i value?", font = ('TkDefaultFont', 10)).grid(row = 1, column = 0, sticky = 'w')
        
        self.CheckVar_zerotwo_yes = tk.IntVar()
        self.CheckVar_zerotwo_yes.set(0)
        self.option_zerotwo_yes = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_zerotwo_yes, command = self.zerotwo_yes).grid(row = 1, column = 1, sticky = 'w')
        
        self.CheckVar_zerotwo_no = tk.IntVar()
        self.CheckVar_zerotwo_no.set(0)
        self.option_zerotwo_no_cups = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_zerotwo_no, command = self.zerotwo_no).grid(row = 1, column = 1, sticky = 'e')
    
    def spikewt_no(self):
        
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        
        tk.Label(dialog_frame, text = "Would you like to alter the 230/232i value?", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        
        self.CheckVar_zerotwo_yes = tk.IntVar()
        self.CheckVar_zerotwo_yes.set(0)
        self.option_zerotwo_yes = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_zerotwo_yes, command = self.zerotwo_yes).grid(row = 0, column = 1, sticky = 'w')
        
        self.CheckVar_zerotwo_no = tk.IntVar()
        self.CheckVar_zerotwo_no.set(0)
        self.option_zerotwo_no_cups = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_zerotwo_no, command = self.zerotwo_no).grid(row = 0, column = 2, sticky = 'e')
    
    def zerotwo_yes(self):
        
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        
        self.zerotwo_yes = 1
        
        tk.Label(dialog_frame, text = "Enter 230/232i ratio: ", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        self.zerotwo = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.zerotwo.grid(row = 0, column = 1, sticky = 'w')
        self.zerotwo.focus_set
        
        tk.Label(dialog_frame, text = "Enter 230/232i ratio error: ", font = ('TkDefaultFont', 10)).grid(row = 1, column = 0, sticky = 'w')
        self.zerotwo_err = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.zerotwo_err.grid(row = 1, column = 1, sticky = 'w')
        self.zerotwo_err.focus_set
        
        self.submit_button = tk.Button(dialog_frame, text = "Submit", font = ('TkDefaultFont', 10), default = "active", command = self.click_submit).grid(row = 2, column = 0)
        
    def zerotwo_no(self):
        
        dialog_frame = tk.Frame(self.otherframe)
        dialog_frame.pack()
        
        self.submit_button = tk.Button(dialog_frame, text = "Submit", font = ('TkDefaultFont', 10), default = "active", command = self.click_submit).grid(row = 0, column = 0)
        
    def click_submit(self):
        
        if self.spike_yes == 1:
            spike_conc_three = self.spike_conc_three.get()
            spike_conc_nine = 1.0
        
            preset_values[0] = spike_conc_three
            preset_values[1] = spike_conc_nine
        
        if self.samplewt_yes == 1:
            sample_wt_err = self.sample_wt_err.get()
            
            preset_values[2] = float(sample_wt_err) / 1000
        
        if self.spikewt_yes == 1:
            spike_wt_err = self.spike_wt_err.get()
        
            preset_values[3] = float(spike_wt_err) / 1000
            
        if self.zerotwo_yes == 1:
            zerotwo = self.zerotwo.get()
            zerotwo_err = self.zerotwo_err.get()
        
            preset_values[4] = zerotwo
            preset_values[5] = zerotwo_err
        
        self.otherframe.destroy()
        self.original_frame.show()
    
    def on_closing(self):
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            self.otherframe.destroy()
        
    
root = tk.Tk()
    
app = Application(master=root)

def on_closing():
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        root.destroy()
        root.quit()

root.protocol("WM_DELETE_WINDOW", on_closing)

root.mainloop()       
    
    
