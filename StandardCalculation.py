#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Created on Fri May 26 11:37:25 2017

@author: julianissen

Finalized version of standard calculator for use in running standards on MC-ICP-MS
Includes the ability to run standard using both SEM and cups configuration

"""

import sys
import Tkinter as tk
import tkFileDialog as filedialog
import tkMessageBox as messagebox
import openpyxl
import csv
import numpy as np
import os
from scipy.optimize import curve_fit
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from itertools import islice


class Application(tk.Frame):
    """
    Root GUI for working with StandardCalculation. Gives option of either running SEM or Cups standard.
    """
    def __init__(self, master):
        """
        Initiates Tkinter window for importing standard information
        """
        tk.Frame.__init__(self,master)
        self.dialog_frame_top = tk.Frame(self)
        self.dialog_frame_top.pack()
        tk.Label(self.dialog_frame_top, text = "Welcome to the Standard Calculation Program!", font = ('TkDefaultFont', 10) ).grid(row = 0, column = 0, sticky = 'e')
        self.master.title("Standard Calculator")
        self.create_widgets()
        self.pack()
   
    def create_widgets(self):
        """
        Prompts whether running standard on Cups or SEM
        """
        
        self.dialog_frame = tk.Frame(self)
        self.dialog_frame.pack()
        
        tk.Label(self.dialog_frame, text = "Check which method used for standard run: ", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        self.CheckVar_sem = tk.IntVar()
        self.CheckVar_sem.set(0)
        self.checkbutton_sem = tk.Checkbutton(self.dialog_frame, text = 'SEM', font = ('TkDefaultFont', 10), variable = self.CheckVar_sem, command = self.sem_command).grid(row = 0, column = 1, sticky = 'w')
        
        self.CheckVar_cups = tk.IntVar()
        self.CheckVar_cups.set(0)
        self.checkbutton_cups = tk.Checkbutton(self.dialog_frame, text = 'CUPS', font = ('TkDefaultFont', 10), variable = self.CheckVar_cups, command = self.cups_command).grid(row = 0, column = 2, sticky = 'w')
        
    def sem_command(self):
        """
        Runs functions for calculating SEM standard
        """
        Application_sem()
    
    def cups_command(self):
        """    
        Runs functions for calculating Cups standard
        """
        Application_cups()
            
class Application_sem(tk.Frame):
    """
    GUI for working with SEM Standard runs
    """
    
    def __init__(self, master=None):
        """
        Initiates SEM Tkinter options
        """
        tk.Frame.__init__(self)
        self.pack()
        self.create_widgets_sem()
    
    def create_widgets_sem(self):
        """
        Creates manual entry windows for AS, 234U wash, and spike.  
        """
         
        dialog_frame_sem = tk.Frame(self)
        dialog_frame_sem.pack()
        
        tk.Label(dialog_frame_sem, text = "112A Standard Calculator for SEM:", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        
        #abundance sensitivity
        tk.Label(dialog_frame_sem, text = "Enter abundance sensitivity for 237U-238U:  ", font = ('TkDefaultFont', 10)).grid(row  = 1, column = 0, sticky = 'w')
        self.AS = tk.Entry(dialog_frame_sem, background = 'white', width = 12)
        self.AS.grid(row = 1, column = 1, sticky = 'w')
        self.AS.focus_set()
        
        #234U wash in cps
        tk.Label(dialog_frame_sem, text = "Enter 234U wash cps:  ", font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'w')
        self.wash = tk.Entry(dialog_frame_sem, background = 'white', width = 12)
        self.wash.grid(row = 2, column = 1, sticky = 'w')
        self.wash.focus_set()
        
        #spike used
        tk.Label(dialog_frame_sem, text = "Enter standard spike information (choose from: DIII-B, DIII-A, 1I, 1H):  ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0, sticky = 'w')
        self.spikeinput = tk.Entry(dialog_frame_sem, background = 'white', width = 12)
        self.spikeinput.grid(row = 3, column = 1, sticky = 'w')
        self.spikeinput.focus_set()
        
        #option of altering standard method
        tk.Label(dialog_frame_sem, text = "Would you like to alter your standard method file before running?: ", font = ('TkDefaultFont', 10) ).grid(row = 4, column = 0, sticky = 'e')
        
        self.CheckVar_option_yes = tk.IntVar()
        self.CheckVar_option_yes.set(0)
        self.option_checkbutton_yes = tk.Checkbutton(dialog_frame_sem, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_option_yes, command = self.option_yes).grid(row = 4, column = 1, sticky = 'w')
        
        self.CheckVar_option_no = tk.IntVar()
        self.CheckVar_option_no.set(0)
        self.option_checkbutton_no = tk.Checkbutton(dialog_frame_sem, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_option_no, command = self.option_no).grid(row = 4, column = 1, sticky = 'e')
    
    def option_no(self):
        """
        Uploading unaltered 112A standard
        """
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        #uploading 112A standard file
        self.usem_uploadbutton = tk.Button(dialog_frame, text = 'Upload 112A standard file', font = ('TkDefaultFont', 10), command = self.file_usem_upload).grid(row = 0, column = 0, sticky = 'e')
    
        self.CheckVar_usem_upload = tk.IntVar()
        self.CheckVar_usem_upload.set(0)
        self.usem_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_usem_upload).grid(row = 0, column = 1, sticky = 'w')
        
        #run SEM standard calculation
        self.standard = tk.Button(dialog_frame, text = 'Calculate Standard',font = ('TkDefaultFont', 10),  default = 'active', command = self.standard).grid(row = 1, column = 1, sticky = 'w')
        
        #quit
        self.quit = tk.Button(dialog_frame, text="QUIT", font = ('TkDefaultFont', 10), command= self.quit_program).grid(row = 1, column = 2, sticky = 'w')
    
    def option_yes(self):
        """
        Changing 112A standard file by specifying which cycle to end on, uploading altered 112A files
        """
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        #altering 112A standard file
        tk.Label(dialog_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.rowinput.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput.focus_set()
        
        #uploading 112A standard file
        self.usem_uploadbutton = tk.Button(dialog_frame, text = 'Upload 112A standard file', font = ('TkDefaultFont', 10), command = self.file_usem_upload_option).grid(row = 1, column = 0, sticky = 'e')
    
        self.CheckVar_usem_upload = tk.IntVar()
        self.CheckVar_usem_upload.set(0)
        self.usem_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_usem_upload).grid(row = 1, column = 1, sticky = 'w')
    
        #run SEM standard calculation
        self.standard = tk.Button(dialog_frame, text = 'Calculate Standard', font = ('TkDefaultFont', 10), default = 'active', command = self.standard).grid(row = 2, column = 1, sticky = 'w')
    
        #quit
        self.quit = tk.Button(dialog_frame, text="QUIT", font = ('TkDefaultFont', 10), command= self.quit_program).grid(row = 2, column = 2, sticky = 'w')
    
    def quit_program(self):
        """
        Window destroy
        """
        self.master.destroy()
        root.quit()
        
    def file_usem_upload(self):
        """
        Uploads 112A SEM standard file and once loaded marks checkbox
        """
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_u_sem = openpyxl.Workbook()
            ws = filename_u_sem.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_u_sem.save("112A_sem.xlsx")
            self.filename_u_sem = "112A_sem.xlsx"
            self.CheckVar_usem_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_u_sem = filename_raw
            self.CheckVar_usem_upload.set(1)

    def file_usem_upload_option(self):
        """
        Uploads altered 112A SEM standard file and once loaded marks checkbox
        """
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_u_sem = openpyxl.Workbook()
            ws = filename_u_sem.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in islice(reader, 0, int(self.rowinput.get())+10):
                    ws.append(row)
            filename_u_sem.save("112A_sem.xlsx")
            self.filename_u_sem = "112A_sem.xlsx"
            self.CheckVar_usem_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
        
    def standard(self):
        """
        Function for calculating 236/233, 235/233, and d234U of SEM standard 
        """
        
        self.spike_input = self.spikeinput.get()
        spike = self.spike_input
    
        #derives spike value based off dictionary entries
        spike_six_three_dictionary = {"DIII-B":1.008398,"DIII-A": 1.008398,"1I":1.010128,"1H":1.010128}
        spike_six_three_err_dictionary = {"DIII-B": 0.00015, "DIII-A": 0.00015, "1I": 0.00015, "1H": 0.00015}
        spike_three_dictionary = {"DIII-B": 0.78938, "DIII-A": 0.78933, "1I": 0.61351, "1H": 0.78997}
        spike_three_err_dictionary = {"DIII-B": 0.00002, "DIII-A": 0.00002, "1I": 0.00002, "1H": 0.00002}
        spike_nine_dictionary = {"DIII-B": 0.21734, "DIII-A": 0.21705, "1I": 0.177187, "1H": 0.22815}
        spike_nine_err_dictionary = {"DIII-B": 0.00001, "DIII-A": 0.00002, "1I": 0.00001, "1H": 0.00001}
        spike_zero_nine_dictionary = {"DIII-B": 0.0000625, "DIII-A": 0.0000625, "1I": 0.0000402, "1H": 0.0000402}
        spike_zero_nine_err_dictionary = {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.0000011, "1H": 0.0000011}
        spike_nine_two_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.00, "1H": 0.00}
        spike_nine_two_err_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.00, "1H": 0.00}
        spike_four_three_dictionary = {"DIII-B": 0.003195, "DIII-A": 0.003195, "1I":0.003180, "1H": 0.003180}
        spike_four_three_err_dictionary= {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.000003, "1H": 0.000003}
        spike_five_three_dictionary = {"DIII-B": 0.105321, "DIII-A": 0.10532, "1I": 0.10521, "1H":0.10521}
        spike_five_three_err_dictionary = {"DIII-B": 0.00003, "DIII-A": 0.00003, "1I": 0.00003, "1H": 0.00003}
        spike_eight_three_dictionary = {"DIII-B": 0.016802, "DIII-A": 0.01680, "1I": 0.01700, "1H":0.01700 }
        spike_eight_three_err_dictionary = {"DIII-B": 0.00001, "DIII-A": 0.00001,"1I": 0.00001, "1H": 0.00001}

        if spike in spike_six_three_dictionary:
            self.spike = float(spike_six_three_dictionary[spike]) #spike ratio
        else: 
            messagebox.showwarning("Error!", "No valid spike info entered! ")
        
        if spike in spike_six_three_err_dictionary: 
            self.spike_six_three_err = float(spike_six_three_err_dictionary[spike]) #error of spike ratio
            
        if spike in spike_three_dictionary:
            self.spike_three = float(spike_three_dictionary[spike]) #in pmol/g
        else:pass
    
        if spike in spike_three_err_dictionary:
            self.spike_three_err = float(spike_three_err_dictionary[spike]) #in pmol/g
        else:pass
    
        if spike in spike_nine_dictionary:
            self.spike_nine = float(spike_nine_dictionary[spike]) #in pmol/g
        else: pass
    
        if spike in spike_nine_err_dictionary: 
            self.spike_nine_err = float(spike_nine_err_dictionary[spike]) #in pmol/g
        else: pass
    
        if spike in spike_zero_nine_dictionary:
            self.spike_zero_nine = float(spike_zero_nine_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_zero_nine_err_dictionary:
            self.spike_zero_nine_err = float(spike_zero_nine_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_nine_two_dictionary: 
            self.spike_nine_two = float(spike_nine_two_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_nine_two_err_dictionary:
            self.spike_nine_two_err = float(spike_nine_two_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_four_three_dictionary:
            self.spike_four_three = float(spike_four_three_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_four_three_err_dictionary:
            self.spike_four_three_err = float(spike_four_three_err_dictionary[spike]) #error of spike ratio
        else: pass
            
        if spike in spike_five_three_dictionary:
            self.spike_five_three = float(spike_five_three_dictionary[spike]) #spike ratio
        else: pass
        
        if spike in spike_five_three_err_dictionary:
            self.spike_five_three_err = float(spike_five_three_err_dictionary[spike]) #error of spike ratio
        else: pass
        
        if spike in spike_eight_three_dictionary:
            self.spike_eight_three = float(spike_eight_three_dictionary[spike]) #spike ratio
        else: pass
        
        if spike in spike_eight_three_err_dictionary:
            self.spike_eight_three_err = float(spike_eight_three_err_dictionary[spike]) #error of spike ratio
        else: pass
        
        #machine parameters
        self.AS = float(self.AS.get())
        self.wash = float(self.wash.get())
    
        #filename 
        filename = self.filename_u_sem
        
        #constants used
        self.wt_235 = 235.043924
        self.wt_233 = 233.039629
        self.wt_236 = 236.045563
        self.wt_234 = 234.040947
        self.eight_five_rat = 137.83
        self.AS_six_eight = self.AS/5
        self.AS_four_eight = self.AS/20
        self.eight_five_rat_err_rel = 0.0003
        lambda_238 = 0.000000000155125
        lambda_234 = 0.00000282206055
        
        #236/233 filtered measured mean and 2s error
        working = isofilter(filename, "G", 44)
        a = working.getMean()
        b = working.getStanddev()
        c = working.getCounts()
        self.six_three_mean_meas = working.Filtered_mean(a,b,c)
        self.six_three_err_meas = working.Filtered_err(a,b,c)
        
        #235/233 filtered measured mean and 2s error
        working_b = isofilter(filename, "H", 44)
        a = working_b.getMean()
        b = working_b.getStanddev()
        c = working_b.getCounts()
        self.five_three_mean_meas = working_b.Filtered_mean(a,b,c)
        self.five_three_err_meas = working_b.Filtered_err(a,b,c) 
    
        #234/235 filtered measured mean and 2s error
        working_c = isofilter(filename,"I", 44)
        a = working_c.getMean()
        b = working_c.getStanddev()
        c = working_c.getCounts()
        self.four_five_mean_meas = working_c.Filtered_mean(a,b,c)
        self.four_five_err_meas = working_c.Filtered_err(a,b,c) 
        
        #234/233 unfiltered mean 
        working_d = isofilter(filename, "J", 44)
        self.four_three_mean_meas = working_d.getMean()
        
        #233 unfiltered mean and counts
        working_e = isofilter(filename, "C", 44)
        self.three_mean_meas = working_e.getMean()
        self.three_counts = working_e.getCounts()
        
        #corrects 236/233 ratio for 238 tail 
        self.six_three_corr = self.six_three_mean_meas * ( 1 - (self.AS_six_eight * self.five_three_mean_meas 
                                                                * self.eight_five_rat/self.spike) )
        
        #provides the ratio that will be used to correct for mass fractionation
        rat = float(np.log(self.wt_235/self.wt_233)/np.log(self.wt_236/self.wt_233))
        
        #corrects 235/233 for mass fractionation in the ICP-MS
        self.five_three_norm = self.five_three_mean_meas * (self.spike/self.six_three_corr)**rat
        
        #provides relative error constants to be used in this function
        AS_six_eight_err_rel = 0.3
        five_three_err_rel = self.five_three_err_meas/self.five_three_mean_meas
        six_three_err_rel = self.six_three_err_meas/self.six_three_mean_meas
       
        #calculates the 236/233 corrected error
        self.six_three_corr_err = self.six_three_corr * np.sqrt( six_three_err_rel**2 + 
                                                                ( (self.AS_six_eight * self.five_three_mean_meas 
                                                                   * self.eight_five_rat)/self.spike  
                                                                   * np.sqrt( AS_six_eight_err_rel**2 + five_three_err_rel ** 2 + self.eight_five_rat_err_rel**2 ) 
                                                                   / (1 - (self.AS_six_eight * self.five_three_mean_meas * self.eight_five_rat)
                                                                   / self.spike) ) ** 2 ) 
        #calculates the 236/233 relative corrected error
        self.six_three_corr_err_rel = self.six_three_corr_err/self.six_three_corr
        
        #calculates the 235/233 normalized error
        self.five_three_norm_err = self.five_three_norm * np.sqrt( five_three_err_rel**2 
                                                                  + (2 * (self.six_three_corr_err_rel/3))**2  ) 
        
        #calculates constants that will be used to calculate normalized 234/235
        four_five_err_rel = self.four_five_err_meas / self.four_five_mean_meas
        
        rat = float(np.log(self.wt_234/self.wt_235)/np.log(self.wt_236/self.wt_233))
        
        #normalizes the 234/235 ratio by correcting for mass fractionation and calculates the resulting error
        self.four_five_norm = self.four_five_mean_meas * (self.spike/self.six_three_corr)**rat
        
        self.four_five_norm_err = self.four_five_norm * np.sqrt( four_five_err_rel**2 + 
                                                                (self.six_three_corr_err_rel/3)**2 )
        
        #calculates constants that will be used to calculate corrected 234/235
        AS_four_eight_err_rel = 0.3
        four_five_norm_err_rel = self.four_five_norm_err/self.four_five_norm
        
        #corrects the normalized 234/235 ratio for 238 tail and calculates the resulting error
        self.four_five_normcorr = self.four_five_norm * (1 - ( self.eight_five_rat 
                                                              * self.AS_four_eight/ self.four_five_norm ))

        self.four_five_normcorr_err = self.four_five_normcorr * np.sqrt( four_five_norm_err_rel**2 + 
                                                                        ( (self.eight_five_rat * self.AS_four_eight / self.four_five_norm) *
                                                                         np.sqrt( self.eight_five_rat_err_rel**2 + AS_four_eight_err_rel**2 + four_five_norm_err_rel**2 )
                                                                         / (1 - ( self.eight_five_rat * self.AS_four_eight/ self.four_five_norm)) ) **2 ) 
        
        #calculates the constants that will be used to calculate normalized 234/233
        rat = float(np.log(self.wt_234/self.wt_233)/np.log(self.wt_236/self.wt_233))
        
        #normalizes the 234/233 ratio by correcting for mass fracitonation
        self.four_three_norm = self.four_three_mean_meas * (self.spike/self.six_three_corr)**rat
        
        #calculates delta 234 and 2s error
        self.d_234 = (((((( self.four_five_normcorr * (1 - self.wash/(self.three_mean_meas * self.five_three_norm * self.four_five_normcorr)))/ self.eight_five_rat) 
                            * (1 - self.spike_four_three/self.four_three_norm)) / (1 - self.spike_five_three/self.five_three_norm))
                            * (lambda_234/lambda_238)) - 1) * 1000
        
        self.d_234_err = self.four_five_normcorr_err * ( (1/self.eight_five_rat) * (lambda_234/lambda_238) ) * (1 - (self.spike_four_three/self.four_three_norm) ) / (1 - (self.spike_five_three/self.five_three_norm))  * 1000
        
        # message box 
        messagebox.showinfo( "112A STANDARD VALUES: ",
        "\n236/233: " + str("{0:.5f}".format(self.six_three_mean_meas)) + " ± " + str("{0:.6f}".format(self.six_three_err_meas)) +\
        "\n235/233: " + str("{0:.4f}".format(self.five_three_norm)) + " ± " + str("{0:.4f}".format(self.five_three_norm_err)) +\
        "\nd234: " + str("{0:.4f}".format(self.d_234)) + " ± " + str("{0:.2f}".format(self.d_234_err)))
       
        # plot of 234 beam stability
        plot_figure().plot_234(self.filename_u_sem, "A", "D")
        
        # deleting excel files
        os.remove("112A_sem.xlsx")

class Application_cups(tk.Frame):
    
    """
    GUI for working with Cups standard runs
    """
    
    def __init__(self, master=None):
        """
        Initiates Cups Tkinter options
        """
        tk.Frame.__init__(self)
        self.pack()
        self.create_widgets_cups()
    
    def create_widgets_cups(self):
        """
        Creates manual entry windows for spike  
        """
         
        dialog_frame_cups = tk.Frame(self)
        dialog_frame_cups.pack()
        
        tk.Label(dialog_frame_cups, text = "112A Standard Calculator for Cups:", font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'w')
        
        #spike used
        tk.Label(dialog_frame_cups, text = "Enter standard spike information (choose from: DIII-B, DIII-A, 1I, 1H):  ", font = ('TkDefaultFont', 10)).grid(row = 1, column = 0, sticky = 'w')
        self.spikeinput = tk.Entry(dialog_frame_cups, background = 'white', width = 12)
        self.spikeinput.grid(row = 1, column = 1, sticky = 'w')
        self.spikeinput.focus_set()
        
        #option of altering unspiked standard file
        tk.Label(dialog_frame_cups, text = "Would you like to alter your unspiked standard method file before running?: ", font = ('TkDefaultFont', 10)).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_unspiked_yes = tk.IntVar()
        self.CheckVar_unspiked_yes.set(0)
        self.option_unspiked_yes = tk.Checkbutton(dialog_frame_cups, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_unspiked_yes, command = self.unspiked_yes).grid(row = 2, column = 1, sticky = 'w')
        
        self.CheckVar_unspiked_no = tk.IntVar()
        self.CheckVar_unspiked_no.set(0)
        self.option_unspiked_no = tk.Checkbutton(dialog_frame_cups, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_unspiked_no, command = self.unspiked_no).grid(row = 2, column = 1, sticky = 'e')
    
    def unspiked_yes(self):
        """
        Changing unspiked standard file by specifying which cycle to end on, uploading altered unspiked standard files
        """
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
    
        #altering unspiked standard file
        tk.Label(dialog_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10)).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput_unspiked = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.rowinput_unspiked.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_unspiked.focus_set()
        
        #uploading unspiked standard file
        self.unspiked_uploadbutton = tk.Button(dialog_frame, text = 'Upload unspiked standard', font = ('TkDefaultFont', 10), command = self.file_unspiked_upload_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_unspiked_upload = tk.IntVar()
        self.CheckVar_unspiked_upload.set(0)
        self.unspiked_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_unspiked_upload).grid(row = 1, column = 1, sticky = 'w')
        
        #uploading unspiked standard wash file
        self.unspiked_wash_uploadbutton = tk.Button(dialog_frame, text = 'Upload unspiked standard wash', font = ('TkDefaultFont', 10), command = self.file_unspiked_wash_upload).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_unspiked_wash_upload = tk.IntVar()
        self.CheckVar_unspiked_wash_upload.set(0)
        self.unspiked_wash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_unspiked_wash_upload).grid(row = 2, column = 1, sticky = 'w')
        
        #option of altering spiked standard file
        tk.Label(dialog_frame, text = "Would you like to alter your spiked standard method file before running?: ", font = ('TkDefaultFont', 10)).grid(row = 3, column = 0)
        
        self.CheckVar_spiked_yes = tk.IntVar()
        self.CheckVar_spiked_yes.set(0)
        self.option_spiked_yes = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_spiked_yes, command = self.spiked_yes).grid(row = 3, column = 1, sticky = 'w')
        
        self.CheckVar_spiked_no = tk.IntVar()
        self.CheckVar_spiked_no.set(0)
        self.option_spiked_no = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_spiked_no, command = self.spiked_no).grid(row = 3, column = 1)
        
    def unspiked_no(self):
        """
        Uploading unaltered unspiked standard files
        """
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        #uploading unspiked standard file
        self.unspiked_uploadbutton = tk.Button(dialog_frame, text = 'Upload unspiked standard', font = ('TkDefaultFont', 10), command = self.file_unspiked_upload).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_unspiked_upload = tk.IntVar()
        self.CheckVar_unspiked_upload.set(0)
        self.unspiked_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_unspiked_upload).grid(row = 0, column = 1, sticky = 'w')
        
        #uploading unspiked standard wash file
        self.unspiked_wash_uploadbutton = tk.Button(dialog_frame, text = 'Upload unspiked standard wash', font = ('TkDefaultFont', 10), command = self.file_unspiked_wash_upload).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_unspiked_wash_upload = tk.IntVar()
        self.CheckVar_unspiked_wash_upload.set(0)
        self.unspiked_wash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_unspiked_wash_upload).grid(row = 1, column = 1, sticky = 'w')
        
        #option of altering spiked standard file
        tk.Label(dialog_frame, text = "Would you like to alter your spiked standard method file before running?: ", font = ('TkDefaultFont', 10) ).grid(row = 2, column = 0, sticky = 'w')
        
        self.CheckVar_spiked_yes = tk.IntVar()
        self.CheckVar_spiked_yes.set(0)
        self.option_spiked_yes = tk.Checkbutton(dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10), variable = self.CheckVar_spiked_yes, command = self.spiked_yes).grid(row = 2, column = 1, sticky = 'w')
        
        self.CheckVar_spiked_no = tk.IntVar()
        self.CheckVar_spiked_no.set(0)
        self.option_spiked_no = tk.Checkbutton(dialog_frame, text = 'No', font = ('TkDefaultFont', 10), variable = self.CheckVar_spiked_no, command = self.spiked_no).grid(row = 2, column = 2, sticky = 'w')
        
    def spiked_yes(self):
        """
        Changing spiked standard file by specifying which cycle to end on, uploading altered spiked standard files
        """
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        #altering spiked standard file
        tk.Label(dialog_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10) ).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput_spiked = tk.Entry(dialog_frame, background = 'white', width = 12)
        self.rowinput_spiked.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_spiked.focus_set()
        
        #uploading spiked standard file
        self.spiked_uploadbutton = tk.Button(dialog_frame, text = 'Upload concentrated spiked standard', font = ('TkDefaultFont', 10), command = self.file_spiked_upload_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_spiked_upload = tk.IntVar()
        self.CheckVar_spiked_upload.set(0)
        self.spiked_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_spiked_upload).grid(row = 1, column = 1, sticky = 'w')
        
        #uploading spiked standard wash file
        self.spiked_wash_uploadbutton = tk.Button(dialog_frame, text = 'Upload concentrated spiked standard wash', command = self.file_spiked_wash_upload).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_spiked_wash_upload = tk.IntVar()
        self.CheckVar_spiked_wash_upload.set(0)
        self.spiked_wash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_spiked_wash_upload).grid(row = 2, column = 1, sticky = 'w')
        
        #wash method
        tk.Label(dialog_frame, text = "Spiked standard wash run on: ", font = ('TkDefaultFont', 10) ).grid(row = 3, column = 0, sticky = 'w')
        
        self.CheckVar_Uwash_sem = tk.IntVar()
        self.CheckVar_Uwash_sem.set(0)
        self.Uwash_sem_checkbutton= tk.Checkbutton(dialog_frame, text = "SEM", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_sem, command = self.Uwash_sem).grid(row = 3, column = 1, sticky = 'w')
        
        self.CheckVar_Uwash_cups = tk.IntVar()
        self.CheckVar_Uwash_cups.set(0)
        self.Uwash_cups_checkbutton= tk.Checkbutton(dialog_frame, text = "CUPS", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_cups, command = self.Uwash_cups).grid(row = 3, column = 1, sticky = 'e')
        
        #run Cups standard calculation
        self.standard = tk.Button(dialog_frame, text = 'Calculate Standard', default = 'active', font = ('TkDefaultFont', 10), command = self.standard).grid(row = 4, column = 1, sticky = 'w')
        
        #quit
        self.quit = tk.Button(dialog_frame, text="QUIT", font = ('TkDefaultFont', 10), command= self.quit_program).grid(row = 4, column = 2, sticky = 'w')
    
    def spiked_no(self):
        """
        Uploading unaltered spiked standard files
        """
        
        dialog_frame = tk.Frame(self)
        dialog_frame.pack()
        
        #uploading spiked standard file
        self.spiked_uploadbutton = tk.Button(dialog_frame, text = 'Upload concentrated spiked standard', font = ('TkDefaultFont', 10), command = self.file_spiked_upload).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_spiked_upload = tk.IntVar()
        self.CheckVar_spiked_upload.set(0)
        self.spiked_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_spiked_upload).grid(row = 0, column = 1, sticky = 'w')
        
        #uploading spiked standard wash file
        self.spiked_wash_uploadbutton = tk.Button(dialog_frame, text = 'Upload concentrated spiked standard wash', font = ('TkDefaultFont', 10), command = self.file_spiked_wash_upload).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_spiked_wash_upload = tk.IntVar()
        self.CheckVar_spiked_wash_upload.set(0)
        self.spiked_wash_upload_checkbutton = tk.Checkbutton(dialog_frame, text = 'Uploaded', font = ('TkDefaultFont', 10), variable = self.CheckVar_spiked_wash_upload).grid(row = 1, column = 1, sticky = 'w')
        
        #wash method
        tk.Label(dialog_frame, text = "Spiked standard wash run on: ", font = ('TkDefaultFont', 10) ).grid(row = 2, column = 0, sticky = 'w')
        
        self.CheckVar_Uwash_sem = tk.IntVar()
        self.CheckVar_Uwash_sem.set(0)
        self.Uwash_sem_checkbutton= tk.Checkbutton(dialog_frame, text = "SEM", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_sem, command = self.Uwash_sem).grid(row = 2, column = 1, sticky = 'w')
        
        self.CheckVar_Uwash_cups = tk.IntVar()
        self.CheckVar_Uwash_cups.set(0)
        self.Uwash_cups_checkbutton= tk.Checkbutton(dialog_frame, text = "CUPS", font = ('TkDefaultFont', 10), variable = self.CheckVar_Uwash_cups, command = self.Uwash_cups).grid(row = 2, column = 1, sticky = 'e')
        
        #run Cups standard calculation
        self.standard = tk.Button(dialog_frame, text = 'Calculate Standard', font = ('TkDefaultFont', 10), default = 'active', command = self.standard).grid(row = 3, column = 1, sticky = 'w')
        
        #quit
        self.quit = tk.Button(dialog_frame, text="QUIT", font = ('TkDefaultFont', 10), command= self.quit_program).grid(row = 3, column = 2, sticky = 'w')
    
    def quit_program(self):
        """
        Window destroy
        """
        self.master.destroy()
        root.quit()
        
    def Uwash_cups(self):
        """
        Cups U wash method 
        """
        self.Uwash = "cups"
    
    def Uwash_sem(self):
        """
        SEM U wash mthod
        """
        self.Uwash = "sem"
        
    def file_unspiked_upload_option(self):
        """
        Uploads altered unspiked standard file and once loaded marks checkbox
        """
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_unspiked = openpyxl.Workbook()
            ws = filename_unspiked.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in islice(reader, 0, int(self.rowinput_unspiked.get()) + 9):
                    ws.append(row)
            filename_unspiked.save("112A_unspiked.xlsx")
            self.filename_unspiked = "112A_unspiked.xlsx"
            self.CheckVar_unspiked_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:])) 
      
    def file_unspiked_upload(self):
        """
        Uploads unspiked standard file and once loaded marks checkbox
        """
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_unspiked = openpyxl.Workbook()
            ws = filename_unspiked.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_unspiked.save("112A_unspiked.xlsx")
            self.filename_unspiked = "112A_unspiked.xlsx"
            self.CheckVar_unspiked_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_unspiked = filename_raw
            self.CheckVar_unspiked_upload.set(1)
            
    def file_unspiked_wash_upload(self):
        """
        Uploads unspiked standard wash file and once loaded marks checkbox
        """
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_unspiked_wash = openpyxl.Workbook()
            ws = filename_unspiked_wash.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_unspiked_wash.save("112A_unspiked_wash.xlsx")
            self.filename_unspiked_wash = "112A_unspiked_wash.xlsx"
            self.CheckVar_unspiked_wash_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_unspiked_wash = filename_raw
            self.CheckVar_unspiked_wash_upload.set(1)
            
    def file_spiked_upload_option(self):
        """
        Uploads altered spiked standard file and once loaded marks checkbox
        """
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_spiked = openpyxl.Workbook()
            ws = filename_spiked.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in islice(reader, 0, int(self.rowinput_spiked.get()) + 9):
                    ws.append(row)
            filename_spiked.save("112A_spiked.xlsx")
            self.filename_spiked = "112A_spiked.xlsx"
            self.CheckVar_spiked_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:])) 
            
    def file_spiked_upload(self):
        """
        Uploads spiked standard file and once loaded marks checkbox
        """
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_spiked = openpyxl.Workbook()
            ws = filename_spiked.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_spiked.save("112A_spiked.xlsx")
            self.filename_spiked = "112A_spiked.xlsx"
            self.CheckVar_spiked_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_spiked = filename_raw
            self.CheckVar_spiked_upload.set(1)
            
    def file_spiked_wash_upload(self):
        """
        Uploads spiked standard wash file and once loaded marks checkbox
        """
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_spiked_wash = openpyxl.Workbook()
            ws = filename_spiked_wash.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_spiked_wash.save("112A_spiked_wash.xlsx")
            self.filename_spiked_wash = "112A_spiked_wash.xlsx"
            self.CheckVar_spiked_wash_upload.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_spiked_wash = filename_raw
            self.CheckVar_spiked_wash_upload.set(1)
    
    def standard(self):
        """
        Run standard functionfor Cups. Runs first the unspiked standard function followed by the concentrated spike standard run.
        """
        
        self.unspiked_standard()
        self.spiked_standard()
            
    def unspiked_standard(self):
        """
        Function for calculating 237 tail values from unspiked standard run
        """
        
        #constants used
        three_wt = 233.039629
        three_five_wt = 233.5
        four_wt = 234.040947
        four_five_wt = 234.5
        five_wt = 235.043924
        six_wt = 236.045563
        
        """
        UNSPIKED STANDARD
        """
        
        """
        Unspiked standard wash values
        """
        
        #233U wash
        unspiked_three_wash_working = isofilter(self.filename_unspiked_wash,"C", 44)
        unspiked_three_wash = unspiked_three_wash_working.getMean()
        if unspiked_three_wash < 0: unspiked_three_wash = 0.0
        unspiked_three_wash_err = abs( (2 * unspiked_three_wash_working.getStanddev()) / (unspiked_three_wash_working.getCounts())**0.5)
        
        #233.5U wash
        unspiked_three_five_wash_working = isofilter(self.filename_unspiked_wash, "D", 44)
        unspiked_three_five_wash = unspiked_three_five_wash_working.getMean()
        if unspiked_three_five_wash < 0: unspiked_three_five_wash = 0.0
        unspiked_three_five_wash_err = abs( (2 * unspiked_three_five_wash_working.getStanddev()) / (unspiked_three_five_wash_working.getCounts())**0.5)
        
        #234.5U wash
        unspiked_four_five_wash_working = isofilter(self.filename_unspiked_wash, "E", 44)
        unspiked_four_five_wash = unspiked_four_five_wash_working.getMean()
        if unspiked_four_five_wash < 0: unspiked_four_five_wash = 0.0
        unspiked_four_five_wash_err = abs( (2 * unspiked_four_five_wash_working.getStanddev()) / (unspiked_four_five_wash_working.getCounts())**0.5)
        
        #236U wash
        unspiked_six_wash_working = isofilter(self.filename_unspiked_wash, "F", 44)
        unspiked_six_wash = unspiked_six_wash_working.getMean()
        if unspiked_six_wash < 0: unspiked_six_wash = 0.0
        unspiked_six_wash_err = abs( (2 * unspiked_six_wash_working.getStanddev()) / (unspiked_six_wash_working.getCounts())**0.5)
        
        #237U wash
        unspiked_seven_wash_working = isofilter(self.filename_unspiked_wash, "G", 44)
        unspiked_seven_wash = unspiked_seven_wash_working.getMean()
        if unspiked_seven_wash < 0: unspiked_seven_wash = 0
        unspiked_seven_wash_err = abs( (2 * unspiked_six_wash_working.getStanddev()) / (unspiked_six_wash_working.getCounts())**0.5)
        
        """
        Unspiked standard values
        """
        
        #233U
        unspiked_three_working = isofilter(self.filename_unspiked, "C", 44)
        unspiked_three = unspiked_three_working.getMean()
        unspiked_three_err = abs( (2 * unspiked_three_working.getStanddev()) / (unspiked_three_working.getCounts()) ** 0.5 )
        
        #233.5U
        unspiked_three_five_working = isofilter(self.filename_unspiked, "D", 44)
        unspiked_three_five = unspiked_three_five_working.getMean()
        unspiked_three_five_err = abs( (2 * unspiked_three_five_working.getStanddev()) / (unspiked_three_five_working.getCounts()) ** 0.5 )
        
        #234.5U
        unspiked_four_five_working = isofilter(self.filename_unspiked, "E", 44)
        unspiked_four_five = unspiked_four_five_working.getMean()
        unspiked_four_five_err = abs( (2 * unspiked_four_five_working.getStanddev()) / (unspiked_four_five_working.getCounts()) ** 0.5 )
        
        #236U
        unspiked_six_working = isofilter(self.filename_unspiked, "F", 44)
        unspiked_six = unspiked_six_working.getMean()
        unspiked_six_err = abs( (2 * unspiked_six_working.getStanddev()) / (unspiked_six_working.getCounts()) ** 0.5 )
        
        #237U
        unspiked_seven_working = isofilter(self.filename_unspiked, "G", 44)
        unspiked_seven = unspiked_seven_working.getMean()
        unspiked_seven_err = abs( (2 * unspiked_seven_working.getStanddev()) / (unspiked_seven_working.getCounts()) ** 0.5 )
        
        #233/237U
        unspiked_three_seven_working = isofilter(self.filename_unspiked, "H", 44)
        unspiked_three_seven = unspiked_three_seven_working.getMean()
        unspiked_three_seven_err = abs( (2 * unspiked_three_seven_working.getStanddev()) / (unspiked_three_seven_working.getCounts()) ** 0.5 )
        
        #233.5/237U
        unspiked_three_five_seven_working = isofilter(self.filename_unspiked, "I", 44)
        unspiked_three_five_seven = unspiked_three_five_seven_working.getMean()
        unspiked_three_five_seven_err = abs( (2 * unspiked_three_five_seven_working.getStanddev()) / (unspiked_three_five_seven_working.getCounts()) ** 0.5 )
        
        #234.5/237U
        unspiked_four_five_seven_working = isofilter(self.filename_unspiked, "J", 44)
        unspiked_four_five_seven = unspiked_four_five_seven_working.getMean()
        unspiked_four_five_seven_err = abs( (2 * unspiked_four_five_seven_working.getStanddev()) / (unspiked_four_five_seven_working.getCounts()) ** 0.5 )
        
        #236/237U
        unspiked_six_seven_working = isofilter(self.filename_unspiked, "K", 44)
        unspiked_six_seven = unspiked_six_seven_working.getMean()
        unspiked_six_seven_err = abs( (2 * unspiked_six_seven_working.getStanddev()) / (unspiked_six_seven_working.getCounts()) ** 0.5 )
        
        """
        Unspiked machine blank corrected tail values
        """
        #233/237 mb corrected
        unspiked_three_seven_blankcorr = unspiked_three_seven - ((unspiked_three_wash/unspiked_seven)/(1 - (unspiked_seven_wash/unspiked_seven) ))
        
        #233/235 mb corrected
        unspiked_three_five_seven_blankcorr = unspiked_three_five_seven - ((unspiked_three_five_wash/unspiked_seven)/(1 - (unspiked_seven_wash/unspiked_seven) ))
        
        #234/235 mb corrected
        unspiked_four_five_seven_blankcorr = unspiked_four_five_seven - ((unspiked_four_five_wash/unspiked_seven)/(1 - (unspiked_seven_wash/unspiked_seven) ))
        
        #236/237 mb corrected
        unspiked_six_seven_blankcorr = unspiked_six_seven - ((unspiked_six_wash/unspiked_seven)/(1 - (unspiked_seven_wash/unspiked_seven) ))
        
        """
        Measured tail error values
        """
        
        #2s relative errors (int)
        
        three_2s_rel_err = max((unspiked_three_seven_err/unspiked_three_seven), (unspiked_three_wash/unspiked_three))
        
        three_five_2s_rel_err = max((unspiked_three_five_seven_err/unspiked_three_five_seven), (unspiked_three_five_wash/unspiked_three_five))
        
        four_five_2s_rel_err = max((unspiked_four_five_seven_err/unspiked_four_five_seven), (unspiked_four_five_wash/unspiked_four_five))
        
        six_2s_rel_err = max((unspiked_six_seven_err/unspiked_six_seven), (unspiked_six_wash/unspiked_six))
        
        #Max tails
        
        three_max_tail = unspiked_three_seven_blankcorr * (1 + three_2s_rel_err)
         
        three_five_max_tail = unspiked_three_five_seven_blankcorr * (1 + three_five_2s_rel_err)
        
        four_five_max_tail = unspiked_four_five_seven_blankcorr * (1 + four_five_2s_rel_err)
    
        six_max_tail = unspiked_six_seven_blankcorr * (1 + six_2s_rel_err)
        
        #Min tails
        
        three_min_tail = unspiked_three_seven_blankcorr * (1 - three_2s_rel_err)
        
        three_five_min_tail = unspiked_three_five_seven_blankcorr * (1 - three_five_2s_rel_err)
        
        four_five_min_tail = unspiked_four_five_seven_blankcorr * (1 - four_five_2s_rel_err)
        
        six_min_tail = unspiked_six_seven_blankcorr * (1 - six_2s_rel_err)
        
        """
        Corrected tail error values
        """
        
        #calculate log line of best fit
        
        y_log_min = np.array([np.log10(three_min_tail), np.log10(three_five_min_tail), np.log10(four_five_min_tail), np.log10(six_min_tail)])
        
        y_log_max = np.array([np.log10(three_max_tail), np.log10(three_five_max_tail), np.log10(four_five_max_tail), np.log10(six_max_tail)])
        
        x = np.array([three_wt, three_five_wt, four_five_wt, six_wt])
        
        def line(x, a, b):
            return a + b * x
    
        popt_min, pcov_min = curve_fit(line, x, y_log_min)
        
        b_min = 10 ** popt_min[0]
        
        m_min = 10 ** popt_min[1]
    
        popt_max, pcov_max = curve_fit(line, x, y_log_max)
        
        b_max = 10 ** popt_max[0]
        
        m_max = 10 ** popt_max[1]
        
        #Max corrected tails
        
        three_max_corr_tail = b_max * (m_max ** three_wt)
        
        three_five_max_corr_tail = b_max * (m_max ** three_five_wt)
        
        four_max_corr_tail = b_max * (m_max ** four_wt)
        
        four_five_max_corr_tail = b_max * (m_max ** four_five_wt)
        
        five_max_corr_tail = b_max * (m_max ** five_wt)
        
        six_max_corr_tail = b_max * (m_max ** six_wt)
        
        #Min corrected tails
        
        three_min_corr_tail = b_min * (m_min ** three_wt)
        
        three_five_min_corr_tail = b_min * (m_min ** three_five_wt)
        
        four_min_corr_tail = b_min * (m_min ** four_wt)
        
        four_five_min_corr_tail = b_min * (m_min ** four_five_wt)
        
        five_min_corr_tail = b_min * (m_min ** five_wt)
        
        six_min_corr_tail = b_min * (m_min ** six_wt)
        
        #Offset from measurement
        
        three_tail_offset = (((three_max_corr_tail + three_min_corr_tail)/2) - unspiked_three_seven_blankcorr) / unspiked_three_seven_blankcorr
        
        three_five_tail_offset = (((three_five_max_corr_tail + three_five_min_corr_tail)/2) - unspiked_three_five_seven_blankcorr) / unspiked_three_five_seven_blankcorr
        
        four_five_tail_offset = (((four_five_max_corr_tail + four_five_min_corr_tail)/2) - unspiked_four_five_seven_blankcorr) / unspiked_four_five_seven_blankcorr
        
        six_tail_offset = (((six_max_corr_tail + six_min_corr_tail)/2) - unspiked_six_seven_blankcorr) / unspiked_six_seven_blankcorr
        
        four_tail_offset = ((four_five_tail_offset * (four_wt - three_five_wt)) + (three_five_tail_offset * (four_five_wt - four_wt))) / (four_five_wt - three_five_wt)
                            
        five_tail_offset = ((six_tail_offset * (five_wt - four_five_wt)) + (four_five_tail_offset * (six_wt - five_wt))) / (six_wt - four_five_wt)
    
        """
        Finalized tail values
        """
        
        #Tail/237
        
        self.three_seven_tail = unspiked_three_seven_blankcorr
        
        three_five_seven_tail = unspiked_three_five_seven_blankcorr
        
        self.four_seven_tail = np.average([four_max_corr_tail, four_min_corr_tail]) * ( 1 - four_tail_offset)
        
        four_five_seven_tail = unspiked_four_five_seven_blankcorr
        
        self.five_seven_tail = five_max_corr_tail 
        
        self.six_seven_tail = unspiked_six_seven_blankcorr
        
        #2s relative error calculated
        
        three_2s_rel_err_corr = three_2s_rel_err
        
        three_five_2s_rel_err_corr = three_five_2s_rel_err
        
        four_2s_rel_err_corr = max( np.sqrt((three_five_2s_rel_err_corr**2) + (four_five_2s_rel_err**2)), (((four_max_corr_tail-four_min_corr_tail)/2)/self.four_seven_tail)  )
        
        four_five_2s_rel_err_corr = four_five_2s_rel_err
        
        five_2s_rel_err_corr = max( np.sqrt((four_five_2s_rel_err_corr**2) + (six_2s_rel_err**2)), (((five_max_corr_tail-five_min_corr_tail)/2)/self.five_seven_tail)  )
        
        six_2s_rel_err_corr = six_2s_rel_err
        
        #Final tail errors
        
        self.three_seven_err = self.three_seven_tail * max(three_2s_rel_err_corr, 0.05)
        
        self.four_seven_err = self.four_seven_tail * max(four_2s_rel_err_corr, 0.05)
        
        self.five_seven_err = self.five_seven_tail * max(five_2s_rel_err_corr, 0.05)
        
        self.six_seven_err = self.six_seven_tail * max(six_2s_rel_err_corr, 0.05)
        
        """
        # Optional message box for displaying unspiked tail values. Delete quotation marks to run.
        # message box 
        messagebox.showinfo( "UNSPIKED STANDARD TAIL VALUES: ",
        "\n233/237: " + str("{0:.4f}".format(self.three_seven_tail)) + " ± " + str("{0:.4f}".format(self.three_seven_err)) +\
        "\n234/237: " + str("{0:.4f}".format(self.four_seven_tail)) + " ± " + str("{0:.4f}".format(self.four_seven_err)) +\
        "\n235/237: " + str("{0:.4f}".format(self.five_seven_tail)) + " ± " + str("{0:.4f}".format(self.five_seven_err)) +\
        "\n236/237: " + str("{0:.4f}".format(self.six_seven_tail)) + " ± " + str("{0:.4f}".format(self.six_seven_err)))
        """
        # deleting excel files
        try:
            os.remove("112A_unspiked.xlsx")
            os.remove("112A_unspiked_wash.xlsx")
        except: pass
    
    def spiked_standard(self):
        """
        Function for calculating 234/238, 237/238 236/233, 238/235 and d234U of spiked standard
        """
        
        spike_input = self.spikeinput.get()
        spike = spike_input
    
        #derives spike value based off dictionary entries
        spike_six_three_dictionary = {"DIII-B":1.008398,"DIII-A": 1.008398,"1I":1.010128,"1H":1.010128}
        spike_six_three_err_dictionary = {"DIII-B": 0.00015, "DIII-A": 0.00015, "1I": 0.00015, "1H": 0.00015}
        spike_three_dictionary = {"DIII-B": 0.78938, "DIII-A": 0.78933, "1I": 0.61351, "1H": 0.78997}
        spike_three_err_dictionary = {"DIII-B": 0.00002, "DIII-A": 0.00002, "1I": 0.00002, "1H": 0.00002}
        spike_nine_dictionary = {"DIII-B": 0.21734, "DIII-A": 0.21705, "1I": 0.177187, "1H": 0.22815}
        spike_nine_err_dictionary = {"DIII-B": 0.00001, "DIII-A": 0.00002, "1I": 0.00001, "1H": 0.00001}
        spike_zero_nine_dictionary = {"DIII-B": 0.0000625, "DIII-A": 0.0000625, "1I": 0.0000402, "1H": 0.0000402}
        spike_zero_nine_err_dictionary = {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.0000011, "1H": 0.0000011}
        spike_nine_two_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.00, "1H": 0.00}
        spike_nine_two_err_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.00, "1H": 0.00}
        spike_four_three_dictionary = {"DIII-B": 0.003195, "DIII-A": 0.003195, "1I":0.003180, "1H": 0.003180}
        spike_four_three_err_dictionary= {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.000003, "1H": 0.000003}
        spike_five_three_dictionary = {"DIII-B": 0.10532, "DIII-A": 0.10532, "1I": 0.10521, "1H":0.10521}
        spike_five_three_err_dictionary = {"DIII-B": 0.00003, "DIII-A": 0.00003, "1I": 0.00003, "1H": 0.00003}
        spike_eight_three_dictionary = {"DIII-B": 0.01680, "DIII-A": 0.01680, "1I": 0.01700, "1H":0.01700 }
        spike_eight_three_err_dictionary = {"DIII-B": 0.00001, "DIII-A": 0.00001,"1I": 0.00001, "1H": 0.00001}

        if spike in spike_six_three_dictionary:
            spike_six_three = float(spike_six_three_dictionary[spike]) #spike ratio
        else: 
            messagebox.showwarning("Error!", "No valid spike info entered! ")
        
        if spike in spike_six_three_err_dictionary: 
            spike_six_three_err = float(spike_six_three_err_dictionary[spike]) #error of spike ratio
            
        if spike in spike_three_dictionary:
            spike_three = float(spike_three_dictionary[spike]) #in pmol/g
        else:pass
    
        if spike in spike_three_err_dictionary:
            spike_three_err = float(spike_three_err_dictionary[spike]) #in pmol/g
        else:pass
    
        if spike in spike_nine_dictionary:
            spike_nine = float(spike_nine_dictionary[spike]) #in pmol/g
        else: pass
    
        if spike in spike_nine_err_dictionary: 
            spike_nine_err = float(spike_nine_err_dictionary[spike]) #in pmol/g
        else: pass
    
        if spike in spike_zero_nine_dictionary:
            spike_zero_nine = float(spike_zero_nine_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_zero_nine_err_dictionary:
            spike_zero_nine_err = float(spike_zero_nine_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_nine_two_dictionary: 
            spike_nine_two = float(spike_nine_two_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_nine_two_err_dictionary:
            spike_nine_two_err = float(spike_nine_two_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_four_three_dictionary:
            spike_four_three = float(spike_four_three_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_four_three_err_dictionary:
            spike_four_three_err = float(spike_four_three_err_dictionary[spike]) #error of spike ratio
        else: pass
            
        if spike in spike_five_three_dictionary:
            spike_five_three = float(spike_five_three_dictionary[spike]) #spike ratio
        else: pass
        
        if spike in spike_five_three_err_dictionary:
            spike_five_three_err = float(spike_five_three_err_dictionary[spike]) #error of spike ratio
        else: pass
        
        if spike in spike_eight_three_dictionary:
            spike_eight_three = float(spike_eight_three_dictionary[spike]) #spike ratio
        else: pass
        
        if spike in spike_eight_three_err_dictionary:
            spike_eight_three_err = float(spike_eight_three_err_dictionary[spike]) #error of spike ratio
        else: pass
        
        
        
        """"
        SPIKED STANDARD
        """
        
        int_time = 1.049
        
        
        """
        Machine blanks
        """
        int_time_234 = 1.049
        int_time_other = 0.262 #233, 235, 236, 237, 238
        
        #machine blank calculation for Uwash run on SEM
        if self.Uwash == "sem":
        
            #233 
            working_three_mb = isofilter(self.filename_spiked_wash, "D", 44)
            three_mb = working_three_mb.getMean()
            three_mb_err = max( (2 * working_three_mb.getStanddev()/np.sqrt(working_three_mb.getCounts())), 
                               (three_mb * 2 / np.sqrt(working_three_mb.getCounts() * int_time_other * three_mb)) )
            
            #234 
            working_four_mb = isofilter(self.filename_spiked_wash, "E", 44)
            four_mb = working_four_mb.getMean()
            four_mb_err = max( (2 * working_four_mb.getStanddev()/np.sqrt(working_four_mb.getCounts())), 
                               (four_mb * 2 / np.sqrt(working_four_mb.getCounts() * int_time_234 * four_mb)) )
            
            #235
            working_five_mb = isofilter(self.filename_spiked_wash, "F", 44)
            five_mb = working_five_mb.getMean()
            five_mb_err = max( (2 * working_five_mb.getStanddev()/np.sqrt(working_five_mb.getCounts())), 
                               (five_mb * 2 / np.sqrt(working_five_mb.getCounts() * int_time_other * five_mb)) )
            
            #236
            working_six_mb = isofilter(self.filename_spiked_wash, "G", 44)
            six_mb = working_six_mb.getMean()
            six_mb_err = max( (2 * working_six_mb.getStanddev()/np.sqrt(working_six_mb.getCounts())), 
                               (six_mb * 2 / np.sqrt(working_six_mb.getCounts() * int_time_other * six_mb)) )
            
            #237
            working_seven_mb = isofilter(self.filename_spiked_wash, "H", 44)
            seven_mb = working_seven_mb.getMean()
            if seven_mb < 0: seven_mb = 0.0
            seven_mb_err = max( (2 * working_seven_mb.getStanddev()/np.sqrt(working_seven_mb.getCounts())), 
                               (three_mb * 2 / np.sqrt(working_seven_mb.getCounts() * int_time_other * three_mb)) )
            
            #234/238
            working_four_eight_mb = isofilter(self.filename_spiked_wash, "L", 44)
            four_eight_mb = working_four_eight_mb.getMean()
            four_eight_mb_err = 2 * working_four_eight_mb.getStanddev()/np.sqrt(working_four_eight_mb.getCounts())
            
            #238
            eight_mb = four_mb / four_eight_mb
            eight_mb_err = eight_mb * (four_eight_mb_err/four_eight_mb)
        
        #machine blank calculation for Uwash run on Cups
        elif self.Uwash == "cups":
            
            #233 
            working_three_mb = isofilter(self.filename_spiked_wash, "C", 44)
            three_mb = working_three_mb.getMean() * 62422000
            three_mb_err = (2 * working_three_mb.getStanddev()/np.sqrt(working_three_mb.getCounts())) * 6242000
            
            #234
            working_four_mb = isofilter(self.filename_spiked_wash, "D", 44)
            four_mb = working_four_mb.getMean() * 62422000
            four_mb_err = (2 * working_four_mb.getStanddev()/np.sqrt(working_four_mb.getCounts())) * 6242000
            
            #235
            working_five_mb = isofilter(self.filename_spiked_wash, "E", 44)
            five_mb = working_five_mb.getMean() * 62422000
            five_mb_err = (2 * working_five_mb.getStanddev()/np.sqrt(working_five_mb.getCounts())) * 6242000
            
            #236
            working_six_mb = isofilter(self.filename_spiked_wash, "F", 44)
            six_mb = working_six_mb.getMean() * 62422000 
            six_mb_err = (2 * working_six_mb.getStanddev()/np.sqrt(working_six_mb.getCounts())) * 6242000
            
            #237
            working_seven_mb = isofilter(self.filename_spiked_wash, "G", 44)
            seven_mb = working_seven_mb.getMean()
            if seven_mb < 0: seven_mb = 0.0
            seven_mb_err = 2 * working_seven_mb.getStanddev()/np.sqrt(working_seven_mb.getCounts())
            
            #238
            working_eight_mb = isofilter(self.filename_spiked_wash, "H", 44)
            eight_mb = working_eight_mb.getMean() * 62422000
            eight_mb_err = (2 * working_eight_mb.getStanddev()/np.sqrt(working_eight_mb.getCounts())) * 6242000
       
        """
        Measured beam intensities
        """
        #233
        working_three_beam = isofilter(self.filename_spiked,"C", 44)
        three_beam_volts = working_three_beam.getMean()
        three_beam_cps = three_beam_volts/(10**11)/(1.602/(10**19))
        three_beam_abs_err = 2 * working_three_beam.getStanddev()/np.sqrt(working_three_beam.getCounts())
        three_beam_rel_err = 2 / np.sqrt(three_beam_cps * working_three_beam.getCounts() * int_time) * 1000
        
        #234
        working_four_beam = isofilter(self.filename_spiked,"D", 44)
        four_beam_volts = working_four_beam.getMean()
        four_beam_cps = four_beam_volts/(10**11)/(1.602/(10**19))
        four_beam_abs_err = 2 * working_four_beam.getStanddev()/np.sqrt(working_four_beam.getCounts())
        four_beam_rel_err = 2 / np.sqrt(four_beam_cps * working_four_beam.getCounts() * int_time) * 1000
        
        #235
        working_five_beam = isofilter(self.filename_spiked,"E", 44)
        five_beam_volts = working_five_beam.getMean()
        five_beam_cps = five_beam_volts/(10**11)/(1.602/(10**19))
        five_beam_abs_err = 2 * working_five_beam.getStanddev()/np.sqrt(working_five_beam.getCounts())
        five_beam_rel_err = 2 / np.sqrt(five_beam_cps * working_five_beam.getCounts() * int_time) * 1000
        
        #236
        working_six_beam = isofilter(self.filename_spiked,"F", 44)
        six_beam_volts = working_six_beam.getMean()
        six_beam_cps = six_beam_volts/(10**11)/(1.602/(10**19))
        six_beam_abs_err = 2 * working_six_beam.getStanddev()/np.sqrt(working_six_beam.getCounts())
        six_beam_rel_err = 2 / np.sqrt(six_beam_cps * working_six_beam.getCounts() * int_time) * 1000
        
        #237
        working_seven_beam = isofilter(self.filename_spiked,"G", 44)
        seven_beam_cps = working_seven_beam.getMean()
        seven_beam_abs_err = 2 * working_seven_beam.getStanddev()/np.sqrt(working_seven_beam.getCounts())
        seven_beam_rel_err = 2 / np.sqrt(seven_beam_cps * working_seven_beam.getCounts() * int_time) * 1000
        
        #238
        working_eight_beam = isofilter(self.filename_spiked,"H", 44)
        eight_beam_volts = working_eight_beam.getMean()
        eight_beam_cps = eight_beam_volts/(10**11)/(1.602/(10**19))
        eight_beam_abs_err = 2 * working_eight_beam.getStanddev()/np.sqrt(working_eight_beam.getCounts())
        eight_beam_rel_err = 2 / np.sqrt(eight_beam_cps * working_eight_beam.getCounts() * int_time) * 1000
        
        """
        Measured ratios, errors, and arrays
        """
        
        #234/233 measured ratio
        four_three_working = isofilter(self.filename_spiked, "I", 44)
        spiked_four_three_meas = four_three_working.getMean()
        spiked_four_three_meas_err = max((2 * four_three_working.getStanddev() / np.sqrt(four_three_working.getCounts())), 
                                          spiked_four_three_meas * np.sqrt(four_beam_rel_err**2)/(10**3))
        spiked_four_three_meas_array = isocorrection().array(self.filename_spiked, "I")
        
        
        #235/233 measured ratio
        five_three_working = isofilter(self.filename_spiked, "J", 44)
        spiked_five_three_meas = five_three_working.getMean()
        spiked_five_three_meas_err = max((2 * five_three_working.getStanddev() / np.sqrt(five_three_working.getCounts())),
                                         spiked_five_three_meas * np.sqrt(five_beam_rel_err**2)/(10**3))
        spiked_five_three_meas_array = isocorrection().array(self.filename_spiked, "I")
        
        #236/233 measured ratio
        six_three_working = isofilter(self.filename_spiked, "K", 44)
        spiked_six_three_meas = six_three_working.getMean()
        spiked_six_three_meas_err = max((2 * six_three_working.getStanddev() / np.sqrt(six_three_working.getCounts())),
                                         spiked_six_three_meas * np.sqrt(six_beam_rel_err**2 + three_beam_rel_err**2)/(10**3))
        spiked_six_three_meas_array = isocorrection().array(self.filename_spiked, "K")
        
        #238/233 measured ratio
        eight_three_working = isofilter(self.filename_spiked, "L", 44)
        spiked_eight_three_meas = eight_three_working.getMean()
        spiked_eight_three_meas_err = max((2 * eight_three_working.getStanddev() / np.sqrt(eight_three_working.getCounts())),
                                         spiked_eight_three_meas * np.sqrt(eight_beam_rel_err**2)/(10**3))
        spiked_eight_three_meas_array = isocorrection().array(self.filename_spiked, "L")
        
        #237/238 measured ratio
        seven_eight_working = isofilter(self.filename_spiked, "M", 44)
        spiked_seven_eight_meas = seven_eight_working.getMean()
        spiked_seven_eight_meas_err = max((2 * seven_eight_working.getStanddev() / np.sqrt(seven_eight_working.getCounts())), 
                                          spiked_seven_eight_meas * np.sqrt(eight_beam_rel_err**2 + seven_beam_rel_err**2)/(10**3))
        spiked_seven_eight_meas_array = isocorrection().array(self.filename_spiked, "M")
        
        """
        Drift correction for 234/233
        """
        
        #234/233 drift correction
    
        four_array = isocorrection().array(self.filename_spiked, "D")
        three_array = isocorrection().array(self.filename_spiked, "C")
        four_three_array = isocorrection().array(self.filename_spiked, "I")
        drift_array = isocorrection().drift_correction_offset(four_array, four_three_array)
        four_three_drift_corrected_array = isocorrection().drift_correction(drift_array, three_array)
        
        four_three_drift_corrected_err = max( (2 * np.nanstd(a = four_three_drift_corrected_array, ddof = 1)/ np.sqrt(len(four_three_drift_corrected_array[np.logical_not(np.isnan(four_three_drift_corrected_array))]))),
                                             (np.nanmean(four_three_drift_corrected_array) * np.sqrt(four_beam_rel_err**2)/(10**3)) )
        
        
        """
        Machine blank correction
        """
        
        #234/233 mb corrected
        four_three_mb_corrected_array = isocorrection().machine_blank_correction(four_three_drift_corrected_array, 
                                                     three_beam_volts, three_mb, four_mb)
        four_three_mb_corrected_err = np.sqrt(four_three_drift_corrected_err**2 + (four_mb_err/three_beam_cps)**2)
        
        #235/233 mb corrected
        five_three_mb_corrected_array = isocorrection().machine_blank_correction(isocorrection().array(self.filename_spiked, "J"), 
                                                      three_beam_volts, three_mb, five_mb)
        five_three_mb_corrected_err = np.sqrt(spiked_five_three_meas_err**2 + (five_mb_err/three_beam_cps)**2)
        
        #236/233 mb corrected
        six_three_mb_corrected_array = isocorrection().machine_blank_correction(isocorrection().array(self.filename_spiked, "K"), 
                                                     three_beam_volts, three_mb, six_mb)
        six_three_mb_corrected_err = np.sqrt(spiked_six_three_meas_err**2 + (six_mb_err/three_beam_cps)**2)
        
        #238/233 mb corrected
        eight_three_mb_corrected_array = isocorrection().machine_blank_correction(isocorrection().array(self.filename_spiked, "L"),
                                                       three_beam_volts, three_mb, eight_mb)
        eight_three_mb_corrected_err = np.sqrt(spiked_eight_three_meas_err**2 + (eight_mb_err/three_beam_cps)**2)
        
        #237/238 mb corrected
        seven_eight_mb_corrected_array = isocorrection().machine_blank_correction(isocorrection().array(self.filename_spiked, "M"), 
                                                       eight_beam_volts, eight_mb, seven_mb)
        
        #AS for standard run
        AS_seven_eight = np.nanmean(seven_eight_mb_corrected_array)
        AS_seven_eight_err = np.sqrt(spiked_seven_eight_meas_err**2 + (seven_mb_err/three_beam_cps/(np.nanmean(eight_three_mb_corrected_array)))**2)
        
        
        """
        Tail correction
        """
        
        #233/238 tail ratios ppm
        three_eight_tail = self.three_seven_tail * AS_seven_eight * (10**6)
        three_eight_tail_err = three_eight_tail * np.sqrt( (self.three_seven_err/self.three_seven_tail)**2 + (AS_seven_eight_err/AS_seven_eight)**2 ) 
        
        #234/238 tail ratios ppm
        four_eight_tail = self.four_seven_tail * AS_seven_eight * (10**6)
        four_eight_tail_err = four_eight_tail * np.sqrt( (self.four_seven_err/self.four_seven_tail)**2 + (AS_seven_eight_err/AS_seven_eight)**2 )
        
        #235/238 tail ratios ppm
        five_eight_tail = self.five_seven_tail * AS_seven_eight * (10**6)
        five_eight_tail_err = five_eight_tail * np.sqrt( (self.five_seven_err/self.five_seven_tail)**2 + (AS_seven_eight_err/AS_seven_eight)**2 )
        
        #236/238 tail ratios ppm
        six_eight_tail = self.six_seven_tail * AS_seven_eight * (10**6)
        six_eight_tail_err = six_eight_tail * np.sqrt( (self.six_seven_err/self.six_seven_tail)**2 + (AS_seven_eight_err/AS_seven_eight)**2 )
        
        #238/233 machine blank corrected mean
        eight_three_mb_corr = np.nanmean(eight_three_mb_corrected_array)
        
        #234/233 tail corrected
        four_three_tail_corrected_array = isocorrection().tail_correction(four_three_mb_corrected_array, four_eight_tail, three_eight_tail, eight_three_mb_corr, option = 'norm')
        four_three_tail_corrected_err = np.sqrt(four_three_mb_corrected_err**2 + (four_eight_tail_err/(10**6)*(np.nanmean(eight_three_mb_corrected_array)))**2)
        
        #235/233 tail corrected
        five_three_tail_corrected_array = isocorrection().tail_correction(five_three_mb_corrected_array, five_eight_tail, three_eight_tail, eight_three_mb_corr, option = 'norm')
        five_three_tail_corrected_err = np.sqrt(five_three_mb_corrected_err**2 + (five_eight_tail_err/(10**6) * (np.nanmean(eight_three_mb_corrected_array)))**2)
        
        #236/233 tail corrected
        six_three_tail_corrected_array = isocorrection().tail_correction(six_three_mb_corrected_array, six_eight_tail, three_eight_tail, eight_three_mb_corr, option = 'norm')
        six_three_tail_corrected_mean = np.nanmean(six_three_tail_corrected_array)
        six_three_tail_corrected_err = np.sqrt(six_three_mb_corrected_err**2 + (six_eight_tail_err/(10**6) * (np.nanmean(eight_three_mb_corrected_array)))**2)
        
        #238/233 tail corrected
        eight_three_tail_corrected_array = isocorrection().tail_correction(eight_three_mb_corrected_array, three_eight_tail, three_eight_tail, eight_three_mb_corr, option = '238/233')
        eight_three_tail_corrected_err = eight_three_mb_corrected_err
        
        #238/235 tail corrected
        eight_five_tail_corrected_array = isocorrection().tail_correction_alt(eight_three_tail_corrected_array, five_three_tail_corrected_array)
        eight_five_tail_corrected_err = np.sqrt((2 * np.nanstd(a = eight_five_tail_corrected_array, ddof = 1)/np.sqrt( len(eight_five_tail_corrected_array[np.logical_not(np.isnan(eight_five_tail_corrected_array))])))**2 + 
                                                 ((five_eight_tail_err/(10**6) * (eight_three_mb_corr/(np.nanmean(five_three_mb_corrected_array)))) * np.nanmean(eight_five_tail_corrected_array))**2 )
        
        #234/238 tail corrected
        four_eight_tail_corrected_array = isocorrection().tail_correction_alt(four_three_tail_corrected_array, eight_three_tail_corrected_array)
        four_eight_tail_corrected_err = np.sqrt((2 * np.nanstd(a = four_eight_tail_corrected_array, ddof = 1)/np.sqrt( len(four_eight_tail_corrected_array[np.logical_not(np.isnan(four_eight_tail_corrected_array))])))**2 + 
                                                 (four_eight_tail_err/(10**6))**2)
        

        
        """
        Fractionation correction
        """
        
        #234/233 fractionation corrected
        four_three_fract_corrected_array = isocorrection().fractionation_correction(four_three_tail_corrected_array,    
                                                        six_three_tail_corrected_mean, '234', '233', spike_six_three)
        four_three_fract_corrected_mean = np.nanmean(four_three_fract_corrected_array)
        four_three_fract_corrected_err = np.sqrt((four_three_tail_corrected_err/np.nanmean(four_three_tail_corrected_array))**2 + 
                                                 (1 * (six_three_tail_corrected_err/six_three_tail_corrected_mean)/3)**2) * four_three_fract_corrected_mean
        
        #235/233 fractionation corrected                                         
        five_three_fract_corrected_array = isocorrection().fractionation_correction(five_three_tail_corrected_array, 
                                                        six_three_tail_corrected_mean, '235', '233', spike_six_three)
        five_three_fract_corrected_mean = np.nanmean(five_three_fract_corrected_array)
        five_three_fract_corrected_err = np.sqrt((five_three_tail_corrected_err/np.nanmean(five_three_tail_corrected_array))**2 + 
                                                 (2 * (six_three_tail_corrected_err/six_three_tail_corrected_mean)/3)**2) * five_three_fract_corrected_mean
        
        #238/233 fractionation corrected                                         
        eight_three_fract_corrected_array = isocorrection().fractionation_correction(eight_three_tail_corrected_array,  
                                                         six_three_tail_corrected_mean, '238', '233', spike_six_three)
        eight_three_fract_corrected_mean = np.nanmean(eight_three_fract_corrected_array)
        eight_three_fract_corrected_err = np.sqrt((eight_three_tail_corrected_err/np.nanmean(eight_three_tail_corrected_array))**2 + 
                                                  (5 * (six_three_tail_corrected_err/six_three_tail_corrected_mean)/3)**2) * eight_three_fract_corrected_mean
        
        #238/235 fractionation corrected                                          
        eight_five_fract_corrected_array = isocorrection().fractionation_correction(eight_five_tail_corrected_array,  
                                                        six_three_tail_corrected_mean, '238', '235', spike_six_three)
        eight_five_fract_corrected_mean = np.nanmean(eight_five_fract_corrected_array)
        eight_five_fract_corrected_err = np.sqrt((eight_five_tail_corrected_err/np.nanmean(eight_five_tail_corrected_array))**2 + 
                                                 (3 * (six_three_tail_corrected_err/six_three_tail_corrected_mean)/3)**2) * eight_five_fract_corrected_mean
        
        #238/234 fractionation corrected                                         
        four_eight_fract_corrected_array = isocorrection().fractionation_correction(four_eight_tail_corrected_array, 
                                                        six_three_tail_corrected_mean, '234', '238', spike_six_three)
        four_eight_fract_corrected_mean = np.nanmean(four_eight_fract_corrected_array)
        four_eight_fract_corrected_err = np.sqrt((four_eight_tail_corrected_err/np.nanmean(four_eight_tail_corrected_array))**2 + 
                                                 (4 * (six_three_tail_corrected_err/six_three_tail_corrected_mean)/3)**2) * four_eight_fract_corrected_mean
        
                                                                                     
        """
        Spike correction
        """
        
        #234/233 spike corrected
        four_three_spike_corrected_mean = four_three_fract_corrected_mean  - spike_four_three
        four_three_spike_corrected_err = np.sqrt(four_three_fract_corrected_err**2 + spike_four_three_err**2)
        
        #235/233 spike corrected
        five_three_spike_corrected_mean = five_three_fract_corrected_mean - spike_five_three
        five_three_spike_corrected_err = np.sqrt(five_three_fract_corrected_err**2 + spike_five_three_err**2)
        
        #238/233 spike corrected
        eight_three_spike_corrected_mean = eight_three_fract_corrected_mean - spike_eight_three
        eight_three_spike_corrected_err = np.sqrt(eight_three_fract_corrected_err**2 + spike_eight_three_err**2)
        
        #238/235 spike corrected
        eight_five_spike_corrected_mean = (eight_five_fract_corrected_mean - spike_eight_three/five_three_fract_corrected_mean)/(1 - spike_five_three/five_three_fract_corrected_mean)
        eight_five_spike_corrected_err = np.sqrt(eight_five_fract_corrected_err**2 + (eight_five_fract_corrected_mean * np.sqrt((spike_five_three_err/five_three_fract_corrected_mean)**2 + 
                                                                                                                                (spike_eight_three_err/eight_three_fract_corrected_mean)**2))**2)
        #234/238 spike corrected
        four_eight_spike_corrected_mean = ((four_eight_fract_corrected_mean - spike_four_three/eight_three_fract_corrected_mean)/(1 - spike_eight_three/eight_three_fract_corrected_mean)) * (10**6) #in ppm
        four_eight_spike_corrected_err = np.sqrt(four_eight_fract_corrected_err**2 + (four_eight_fract_corrected_mean * np.sqrt((spike_four_three_err/four_three_fract_corrected_mean)**2 + 
                                                                                                                                (spike_eight_three_err/eight_three_fract_corrected_mean)**2))**2) * (10**6) #in ppm                                                                   
                                                                                                                                
        """
        d234U calculation
        """
        lambda_234 = 0.00000282206
        lambda_238 = 0.000000000155125
        
        d234U = ((four_eight_spike_corrected_mean/(10**6)) * lambda_234/lambda_238 - 1) * 1000
        d234U_err = ((four_eight_spike_corrected_err/(10**6)) * lambda_234/lambda_238) * 1000
        
        
        """
        Message box
        """
        messagebox.showinfo("Standard values for Cups: ",
                            "\n234/238 standard value: " + str("{0:.3f}".format(four_eight_spike_corrected_mean)) + " ± " + str("{0:.3f}".format(four_eight_spike_corrected_err)) + " ppm" +\
                            "\n237/238 AS: " + str("{0:.2f}".format(AS_seven_eight * (10**6))) + " ppm" +\
                            "\n236/233: " + str("{0:.5f}".format(six_three_tail_corrected_mean)) + " ± " + str("{0:.5f}".format(six_three_tail_corrected_err)) +\
                            "\n238/235: " + str("{0:.3f}".format(eight_five_spike_corrected_mean)) + " ± " + str("{0:.3f}".format(eight_five_spike_corrected_err)) +\
                            "\nd234U: " + str("{0:.3f}".format(d234U)) + " ± " + str("{0:.3f}".format(d234U_err)))
        
        #plotting 234U beam stability
        plot_figure().plot_234(self.filename_spiked, "A", "D")
        
        # deleting excel files
        try:
            os.remove("112A_spiked.xlsx")
            os.remove("112A_spiked_wash.xlsx")
        except: pass
        
        
class plot_figure(tk.Tk):
    """
    Provides plot of 234U beam stability
    """
    
    def __init__(self):
        """
        Init def, no inputs needed
        """
      
    def plot_234(self, filename, column1, column2): 
        """
        plot of 234U beam
        """
        toplevel = tk.Toplevel()
        toplevel.title("234U beam")
        fig = Figure(figsize = (8,4))
        ax = fig.add_subplot(111)
        x = isocorrection().array(filename, column1)
        y = isocorrection().array(filename, column2)
        ymean = np.nanmean(y)
        ystanddev = np.nanstd(a = y, ddof = 1)
        ax.scatter(x, y, color = 'b', marker = 'o')
        ax.set_xlabel('Cycles', fontsize = 7, labelpad = 0.5)
        ax.set_ylabel('234U', fontsize = 7, labelpad = 0.5)
        ax.set_ylim([ymean - 10*ystanddev, ymean + 10*ystanddev])
        ax.set_title('Beam Intensity 234U' , fontsize = 10)
        
        canvas = FigureCanvasTkAgg(fig, master = toplevel)
        canvas.show()
        canvas.get_tk_widget().pack()
        toplevel.mainloop()
       
        
class isofilter():
    """
    Class for calculating unfiltered and filtered mean, standdev/error, and counts for an Excel column
    """
    def __init__(self, filename, columnletter, filternumber): 
        """
        Loads specified Excel column
        """
        self.column = str(columnletter)+'{}:'+str(columnletter)+'{}'
        self.filename = str(filename)
        self.filternumber = int(filternumber)
        self.workbook = openpyxl.load_workbook(self.filename, data_only = True)
        self.ws = self.workbook.worksheets[0]
        self.totalCounts = 0
        self.mean = 0 
        self.filteredMean = 0
        self.err = 0
        self.criteria = 0
        self.totalCounts_filt = 0
        self.standdev = 0
    
    def getMean(self):
        """
        Code works row by row through specified Excel column, and calculates total mean
        """
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value: 
                    try:
                        value = float(cell.value)
                        outlist.append(value)
                    except:
                        outlist.append(np.nan)
                elif cell.value == 0:
                    value= 0.00
                    outlist.append(value)
                else: outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        self.mean = np.nanmean(a = outarray)
        return self.mean
    

    def getStanddev(self):
        """
        Code works row by row through specified Excel column, and calculates standard deviation
        """
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value: 
                    try:
                        value = float(cell.value)
                        outlist.append(value)
                    except:
                        outlist.append(np.nan)
                elif cell.value == 0:
                    value= 0.00
                    outlist.append(value)
                else: outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float) 
        self.standdev = np.nanstd(a = outarray, ddof = 1)
        return self.standdev
    
    def getCounts(self):
        """
        Code works row by row through specified Excel Column, and determines total number of values present (i.e. cycles)
        """
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value: 
                    try:
                        value = float(cell.value)
                        outcounts += 1
                    except:
                        outcounts += 0
                elif cell.value == 0:
                    outcounts +=1

        self.totalCounts = outcounts
        return self.totalCounts
        
    def Filtered_mean(self, mean, standdev, counts):
        """
        Code works row by row through specified Excel column, deletes entries that are outside of specified range, 
        and calculates resulting mean
        """
        self.mean = mean
        self.standdev = standdev
        self.totalCounts = counts
        self.standerr = (self.standdev / (self.totalCounts**0.5))
        self.criteria = self.filternumber * self.standerr
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value: 
                    try:
                        value = float(cell.value)
                        if abs(value - self.mean) > self.criteria:
                            outlist.append(np.nan)
                        else:
                            outlist.append(value)
                    except: outlist.append(np.nan)
                elif cell.value == 0:
                    value = 0.00
                    if abs(value - self.mean) > self.criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                else: outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        self.filteredMean = np.nanmean(a = outarray)
        return self.filteredMean
     
    def Filtered_err(self, mean, standdev, counts):
        """
        Code works row by row through specified Excel column, deletes entries that are outside of specified range, 
        and calculates resulting 2s counting stantistics error
        """
        self.mean = mean
        self.standdev = standdev
        self.totalCounts = counts
        self.standerr = (self.standdev / (self.totalCounts**0.5))
        self.criteria = self.filternumber * self.standerr
        outlist = []
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value: 
                    try:
                        value = float(cell.value)
                        if abs(value - self.mean) > self.criteria:
                            outlist.append(np.nan)
                        else:
                            outlist.append(value)
                            outcounts += 1
                    except: outlist.append(np.nan)
                elif cell.value == 0:
                    value = 0.00
                    if abs(value - self.mean) > self.criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                        outcounts += 1 
                else: outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        outstanddev = np.nanstd(a = outarray, ddof = 1)
        self.err = 2 * (outstanddev / (outcounts ** 0.5))
        return self.err
        
    def Filtered_counts(self, mean, standdev, counts):
        """
        Code works row by row through specified Excel column, deletes entries that are outside of specified range, 
        and determines total number of values remaining (i.e. filtered cycles)
        """
        self.mean = mean
        self.standdev = standdev
        self.totalCounts = counts
        self.standerr = (self.standdev / (self.totalCounts**0.5))
        self.criteria = self.filternumber * self.standerr
        outlist = []
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value: 
                    try:
                        value = float(cell.value)
                        if abs(value - self.mean) > self.criteria:
                            outlist.append(np.nan)
                        else:
                            outlist.append(value)
                            outcounts += 1
                    except: outlist.append(np.nan)
                elif cell.value == 0:
                    value = 0.00
                    if abs(value - self.mean) > self.criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                        outcounts += 1
                else: outlist.append(np.nan)
        self.totalCounts_filt = outcounts
        return self.totalCounts_filt         
    

class isocorrection():
    """
    Class for creating numpy arrays of Excel columns and completing element-wise corrections
    """
    
    def __init__(self):
        
        """
        Init def, no inputs needed
        """
        
    def array(self, filename, columnletter):
        """
        Code provides output array for excel row. Includes NaN for non-values
        """
        self.column = str(columnletter)+'{}:'+str(columnletter)+'{}'
        self.filename = str(filename)
        self.workbook = openpyxl.load_workbook(self.filename, data_only = True)
        self.ws = self.workbook.worksheets[0]
        
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 9)):
            for cell in row:
                if cell.value:
                    try: 
                        value = float(cell.value)
                        outlist.append(value)
                    except: 
                        outlist.append(np.nan)
                elif cell.value == 0:
                    value= 0.00
                    outlist.append(value)
                else: outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        return outarray
        
    def drift_correction_offset(self, source_array, ratio_array):
        """
        Code provides offset array for given isotope and isotope ratio. Includes NaN for non-values
        """
        int_time = 1.049
        time_constant = 0.65
        
        offset_list = []
        
        for i in range(0, len(source_array) - 1):
            
            if i == 0: offset_list.append(np.nan)
            
            else:
                if np.logical_not(np.isnan(ratio_array[i-1])) and np.logical_not(np.isnan(ratio_array[i])) and np.logical_not(np.isnan(ratio_array[i+1])): 
                    offset = source_array[i] + ((source_array[i+1] - source_array[i-1])/(2*int_time)) * time_constant
                    offset_list.append(offset)  
                else:
                    offset_list.append(np.nan)
            
        offset_array = np.array(offset_list, dtype = np.float)
        return offset_array

    def drift_correction(self, drift_array, source_array):
        """
        Code provides array corrected for drift for a given isotope ratio. Includes NaN for non-values
        """
        driftcorrected_list = []
        
        for i in range(0, len(drift_array)):
            
            if np.logical_not(np.isnan(drift_array[i])):
                drift_corrected = drift_array[i] / source_array[i]
                driftcorrected_list.append(drift_corrected)
            else:
                driftcorrected_list.append(np.nan)
          
        driftcorrected_array = np.array(driftcorrected_list, dtype = np.float)
        return driftcorrected_array
            
    def machine_blank_correction(self, source_array, bottom_mean, machine_blank_bottom_mean, machine_blank_top_mean):
        """
        Code provides a machine blank corrected array for a given isotope ratio. Includes NaN for non-values
        """
        mbcorrected_list = []
        
        for i in range(0, len(source_array)):
            
            if np.logical_not(np.isnan(source_array[i])):
                mbcorrected = (source_array[i] - machine_blank_top_mean/(bottom_mean * 62422000.))/(1 - machine_blank_bottom_mean/(bottom_mean * 62422000.))
                mbcorrected_list.append(mbcorrected)
            else: 
                mbcorrected_list.append(np.nan)
                
        mbcorrected_array = np.array(mbcorrected_list, dtype = np.float)
        return mbcorrected_array
    
    def tail_correction(self, source_array, tail_top, tail_bottom, eight_three_mb_corr, option):
        """
        Code provides a tail corrected array for a given isotope ratio. This tail correction is applied to 234/233, 
        235/233, and 236/233, with slight variation for 238/233. Includes NaN for non-values
        """
        
        tailcorrected_list = []
        
        if str(option) == 'norm':
            for i in range(0, len(source_array)):
                
                if np.logical_not(np.isnan(source_array[i])):
                    tailcorrected = (source_array[i] - tail_top/(10**6) * eight_three_mb_corr)/(1 - tail_bottom/(10**6) * eight_three_mb_corr)
                    tailcorrected_list.append(tailcorrected)
                else:
                    tailcorrected_list.append(np.nan)
                    
        elif str(option) == '238/233':
            for i in range(0, len(source_array)):
                
                if np.logical_not(np.isnan(source_array[i])):
                    tailcorrected = source_array[i]/(1 - (tail_bottom/(10**6)) * eight_three_mb_corr)
                    tailcorrected_list.append(tailcorrected)
                else:
                    tailcorrected_list.append(np.nan)
        
        tailcorrected_array = np.array(tailcorrected_list, dtype = np.float)
        return tailcorrected_array

    def tail_correction_alt(self, top_array_corr, bottom_array_corr):
        """
        Code provides a tail corrected array for a given isotope ratio. This tail correction is applied to 238/235
        and 234/238. Includes NaN for non-values
        """
        
        tailcorrected_list = []
        
        for i in range(0, len(top_array_corr)):
            if np.logical_not(np.isnan(top_array_corr[i])) and np.logical_not(np.isnan(bottom_array_corr[i])): 
                tailcorrected = top_array_corr[i] / bottom_array_corr[i]
                tailcorrected_list.append(tailcorrected) 
            else:
                tailcorrected_list.append(np.nan)
        
        tailcorrected_array = np.array(tailcorrected_list, dtype = np.float)
        return tailcorrected_array
    
    def fractionation_correction(self, source_array, six_three_tail_corr, top_mass, bottom_mass, spike_six_three):
        """
        Code provides a fractionation corrected array for a given isotope ratio.Includes NaN for non-values
        """
        
        wt_236 = 236.045563
        wt_233 = 233.039629
        
        if top_mass == '234': wt_top = 234.040947
        elif top_mass == '235': wt_top = 235.043924
        elif top_mass == '236': wt_top = 236.045563
        elif top_mass == '238': wt_top = 238.050785
        
        if bottom_mass == '233': wt_bottom = 233.039629
        elif bottom_mass == '235': wt_bottom = 235.043924
        elif bottom_mass == '238': wt_bottom = 238.050785
            
        fractcorrected_list = []
        
        for i in range(0, len(source_array)):
            if np.logical_not(np.isnan(source_array[i])):
                fractcorrected = source_array[i] * (spike_six_three/six_three_tail_corr)**(np.log(wt_top/wt_bottom)/np.log(wt_236/wt_233))
                fractcorrected_list.append(fractcorrected)
            else: 
                fractcorrected_list.append(np.nan)
        fractcorrected_array = np.array(fractcorrected_list, dtype = np.float)
        return fractcorrected_array

            
        
root = tk.Tk()
    
app = Application(master=root)

def on_closing():
    """
    If window is X'ed out of, prompts you to quit 
    """
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        root.destroy()
        root.quit()

root.protocol("WM_DELETE_WINDOW", on_closing)

root.mainloop()   

