#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Thu Jul 20 11:28:27 2017

@author: julianissen

Finalized version of reagent blank calculator. Exports Excel document with reagent blank values.  

"""


import sys
import Tkinter as tk
import tkFileDialog as filedialog
import tkMessageBox as messagebox
import numpy as np
import pandas as pd
import openpyxl
import csv
import os
from itertools import islice

class Application(tk.Frame):
    """
    GUI for working with ReagentBlankCalculation
    """
    
    def __init__(self, master):
        """
        Initiates Tkinter window for importing reagent blank information
        """
        tk.Frame.__init__(self,master)
        self.dialog_frame_top = tk.Frame(self)
        self.dialog_frame_top.pack()
        tk.Label(self.dialog_frame_top, text = "Welcome to the Reagent Blank Calculation Program!", font = ('TkDefaultFont', 10)  ).grid(row = 0, column = 0, sticky = 'e')
        self.master.title("ReagentBlank Calculator")
        self.create_widgets()
        self.pack()
    
    def create_widgets(self):
        """
        Creates manual entry windows for blank name, solution weight, uptake rate, ionization efficiency
        and reagent blank export file name
        """
        self.dialog_frame = tk.Frame(self)
        self.dialog_frame.pack()
        
        #chemblank name 
        tk.Label(self.dialog_frame, text = "Enter blank name:  ", font = ('TkDefaultFont', 10) ).grid(row = 0, column = 0, sticky = 'w')
        self.blankname = tk.Entry(self.dialog_frame, background = 'white', width = 12)
        self.blankname.grid(row = 0, column = 1, sticky = 'w')
        self.blankname.focus_set()
        
        #solution weight
        tk.Label(self.dialog_frame, text = "Enter solution weight (g):  ", font = ('TkDefaultFont', 10) ).grid(row = 1, column = 0, sticky = 'w')
        self.sln_wt = tk.Entry(self.dialog_frame, background = 'white', width = 12)
        self.sln_wt.grid(row = 1, column = 1, sticky = 'w')
        self.sln_wt.focus_set()
        
        #uptake rate
        tk.Label(self.dialog_frame, text = "Enter uptake rate:  ", font = ('TkDefaultFont', 10) ).grid(row = 2, column = 0, sticky = 'w')
        self.uptake_rate = tk.Entry(self.dialog_frame, background = 'white', width = 12)
        self.uptake_rate.grid(row = 2, column = 1, sticky = 'w')
        self.uptake_rate.focus_set()
        
        #ionization efficiency
        tk.Label(self.dialog_frame, text = "Enter ionization efficiency:  ", font = ('TkDefaultFont', 10) ).grid(row = 3, column = 0, sticky = 'w')
        self.IE = tk.Entry(self.dialog_frame, background = 'white', width = 12)
        self.IE.grid(row = 3, column = 1, sticky = 'w')
        self.IE.focus_set()
        
        #reagent blank filename
        tk.Label(self.dialog_frame, text = "Enter reagent blank export file name (include .xlsx):  ", font = ('TkDefaultFont', 10) ).grid(row = 4, column = 0, sticky = 'w')
        self.filename = tk.Entry(self.dialog_frame, background = 'white', width = 12)
        self.filename.grid(row = 4, column = 1, sticky = 'w')
        self.filename.focus_set()
        
        #option of altering Th method
        tk.Label(self.dialog_frame, text = 'Would you like to alter your Th method file before running?: ', font = ('TkDefaultFont', 10) ).grid(row = 5, column = 0, sticky = 'w')
        
        self.CheckVar_th_yes = tk.IntVar()
        self.CheckVar_th_yes.set(0)
        self.th_yes = tk.Checkbutton(self.dialog_frame, text = 'Yes', font = ('TkDefaultFont', 10) , variable = self.CheckVar_th_yes, command = self.th_yes).grid(row = 5, column = 1, sticky = 'w')
        
        self.CheckVar_th_no = tk.IntVar()
        self.CheckVar_th_no.set(0)
        self.th_no = tk.Checkbutton(self.dialog_frame, text = 'No', font = ('TkDefaultFont', 10) , variable = self.CheckVar_th_no, command = self.th_no).grid(row = 5, column = 1, sticky = 'e')
        
    def th_yes(self):
        """
        Changing Th file by specifying which cycle to end on, uploading altered Th files
        """
        
        checkbutton_frame = tk.Frame(self)
        checkbutton_frame.pack()
        
        #altering Th file
        tk.Label(checkbutton_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10) ).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput_th = tk.Entry(checkbutton_frame, background = 'white', width = 12)
        self.rowinput_th.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_th.focus_set()
      
        #uploading Th file
        self.th_regblank_upload = tk.Button(checkbutton_frame, text = 'Upload Th reagent blank file', font = ('TkDefaultFont', 10) , command = self.file_upload_th_regblank_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_th_regblank = tk.IntVar()
        self.CheckVar_th_regblank.set(0)
        self.th_regblank_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_th_regblank).grid(row = 1, column = 1, sticky = 'w')
        
        #uploading Th wash file
        self.th_regblankwash_upload = tk.Button(checkbutton_frame, text = 'Upload Th reagent blank wash file', font = ('TkDefaultFont', 10) , command = self.file_upload_th_regblankwash).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_th_regblankwash = tk.IntVar()
        self.CheckVar_th_regblankwash.set(0)
        self.th_regblankwash_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_th_regblankwash).grid(row = 2, column = 1, sticky = 'w')
        
        #option of altering U method
        tk.Label(checkbutton_frame, text = 'Would you like to alter your U method file before running?: ', font = ('TkDefaultFont', 10) ).grid(row = 3, column = 0, sticky = 'w')
        
        self.CheckVar_u_yes = tk.IntVar()
        self.CheckVar_u_yes.set(0)
        self.u_yes = tk.Checkbutton(checkbutton_frame, text = 'Yes', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_yes, command = self.u_yes).grid(row = 3, column = 1, sticky = 'w')
        
        self.CheckVar_u_no = tk.IntVar()
        self.CheckVar_u_no.set(0)
        self.u_no = tk.Checkbutton(checkbutton_frame, text = 'No', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_no, command = self.u_no).grid(row = 3, column = 1, sticky = 'e')
    
    def th_no(self):
        """
        Uploading unaltered Th files
        """
        
        checkbutton_frame = tk.Frame(self)
        checkbutton_frame.pack()
        
        #uploading Th file
        self.th_regblank_upload = tk.Button(checkbutton_frame, text = 'Upload Th reagent blank file', font = ('TkDefaultFont', 10) , command = self.file_upload_th_regblank).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_th_regblank = tk.IntVar()
        self.CheckVar_th_regblank.set(0)
        self.th_regblank_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_th_regblank).grid(row = 0, column = 1, sticky = 'w')
        
        self.th_regblankwash_upload = tk.Button(checkbutton_frame, text = 'Upload Th reagent blank wash file', font = ('TkDefaultFont', 10) , command = self.file_upload_th_regblankwash).grid(row = 1, column = 0, sticky = 'e')
        
        #uploading Th wash file
        self.CheckVar_th_regblankwash = tk.IntVar()
        self.CheckVar_th_regblankwash.set(0)
        self.th_regblankwash_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_th_regblankwash).grid(row = 1, column = 1, sticky = 'w')
        
        #option of altering U method
        tk.Label(checkbutton_frame, text = 'Would you like to alter your U method file before running?: ', font = ('TkDefaultFont', 10) ).grid(row = 2, column = 0, sticky = 'w')
        
        self.CheckVar_u_yes = tk.IntVar()
        self.CheckVar_u_yes.set(0)
        self.u_yes = tk.Checkbutton(checkbutton_frame, text = 'Yes', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_yes, command = self.u_yes).grid(row = 2, column = 1, sticky = 'w')
        
        self.CheckVar_u_no = tk.IntVar()
        self.CheckVar_u_no.set(0)
        self.u_no = tk.Checkbutton(checkbutton_frame, text = 'No', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_no, command = self.u_no).grid(row = 2, column = 2, sticky = 'w')       
    
    def u_yes(self):
        """
        Changing U file by specifying which cycle to end on, uploading altered U files
        """
        
        checkbutton_frame = tk.Frame(self)
        checkbutton_frame.pack()
        
        #alteirng U file
        tk.Label(checkbutton_frame, text = 'Enter last cycle # to be analyzed: ', font = ('TkDefaultFont', 10) ).grid(row = 0, column = 0, sticky = 'e')
        self.rowinput_u = tk.Entry(checkbutton_frame, background = 'white', width = 12)
        self.rowinput_u.grid(row = 0, column = 1, sticky = 'w')
        self.rowinput_u.focus_set()
        
        #uploading U file
        self.u_regblank_upload = tk.Button(checkbutton_frame, text = 'Upload U reagent blank file', font = ('TkDefaultFont', 10) , command = self.file_upload_u_regblank_option).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_u_regblank = tk.IntVar()
        self.CheckVar_u_regblank.set(0)
        self.u_regblank_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_regblank).grid(row = 1, column = 1, sticky = 'w')
        
        #uploading U wash file
        self.u_regblankwash_upload = tk.Button(checkbutton_frame, text = 'Upload U reagent blank wash file', font = ('TkDefaultFont', 10) , command = self.file_upload_u_regblankwash).grid(row = 2, column = 0, sticky = 'e')
        
        self.CheckVar_u_regblankwash = tk.IntVar()
        self.CheckVar_u_regblankwash.set(0)
        self.u_regblankwash_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_regblankwash).grid(row = 2, column = 1, sticky = 'w')
        
        #run reagent blank calculation        
        self.chemblank = tk.Button(checkbutton_frame, text = 'Calculate reagent blank and export data', font = ('TkDefaultFont', 10) , command = self.blank_calculate, default = 'active').grid(row = 3, column = 1, sticky = 'w')
        
        #quit
        self.quit = tk.Button(checkbutton_frame, text="QUIT", font = ('TkDefaultFont', 10) , command= self.quit_program).grid(row = 3, column = 2, sticky = 'w')
    
    def u_no(self):
        """
        Uploading unaltered U files
        """
        
        checkbutton_frame = tk.Frame(self)
        checkbutton_frame.pack()     
        
        #uploading U file
        self.u_regblank_upload = tk.Button(checkbutton_frame, text = 'Upload U reagent blank file', font = ('TkDefaultFont', 10) , command = self.file_upload_u_regblank).grid(row = 0, column = 0, sticky = 'e')
        
        self.CheckVar_u_regblank = tk.IntVar()
        self.CheckVar_u_regblank.set(0)
        self.u_regblank_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_regblank).grid(row = 0, column = 1, sticky = 'w')
        
        #uploading U wash file
        self.u_regblankwash_upload = tk.Button(checkbutton_frame, text = 'Upload U reagent blank wash file', font = ('TkDefaultFont', 10) , command = self.file_upload_u_regblankwash).grid(row = 1, column = 0, sticky = 'e')
        
        self.CheckVar_u_regblankwash = tk.IntVar()
        self.CheckVar_u_regblankwash.set(0)
        self.u_regblankwash_checkbutton = tk.Checkbutton(checkbutton_frame, text = 'Uploaded', font = ('TkDefaultFont', 10) , variable = self.CheckVar_u_regblankwash).grid(row = 1, column = 1, sticky = 'w')
        
        #run reagent blank calculation        
        self.chemblank = tk.Button(checkbutton_frame, text = 'Calculate chemblank and export data', font = ('TkDefaultFont', 10) , command = self.blank_calculate, default = 'active').grid(row = 2, column = 1, sticky = 'w')
        
        #quit
        self.quit = tk.Button(checkbutton_frame, text="QUIT", font = ('TkDefaultFont', 10) , command= self.quit_program).grid(row = 2, column = 2, sticky = 'w')
    
    def quit_program(self):
        """
        Window destroy
        """
        self.master.destroy()
        root.quit()
    
    def file_upload_th_regblank(self):
        """
        Uploads Th reagent blank file and once loaded marks checkbox
        """
    
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_th_regblank = openpyxl.Workbook()
            ws = filename_th_regblank.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_th_regblank.save("regblankth.xlsx")
            self.filename_th_regblank = "regblankth.xlsx"
            self.CheckVar_th_regblank.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_th_regblank = filename_raw
            self.CheckVar_th_regblank.set(1)
            
    def file_upload_th_regblank_option(self):
        """
        Uploads altered Th chemblank file and once loaded marks checkbox
        """
    
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_th_regblank = openpyxl.Workbook()
            ws = filename_th_regblank.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in islice(reader, 0, int(self.rowinput_th.get()) + 9):
                    ws.append(row)
            filename_th_regblank.save("regblankth.xlsx")
            self.filename_th_regblank = "regblankth.xlsx"
            self.CheckVar_th_regblank.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:])) 
    
    def file_upload_th_regblankwash(self):
        """
        Uploads Th reagent blank wash file and once loaded marks checkbox
        """
    
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_th_regblankwash = openpyxl.Workbook()
            ws = filename_th_regblankwash.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_th_regblankwash.save("regblankth_wash.xlsx")
            self.filename_th_regblankwash = "regblankth_wash.xlsx"
            self.CheckVar_th_regblankwash.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_th_regblankwash = filename_raw
            self.CheckVar_th_regblankwash.set(1)
    
    def file_upload_u_regblank(self):
        """
        Uploads U reagent blank file and once loaded marks checkbox
        """
    
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_u_regblank = openpyxl.Workbook()
            ws = filename_u_regblank.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_u_regblank.save("regblanku.xlsx")
            self.filename_u_regblank = "regblanku.xlsx"
            self.CheckVar_u_regblank.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_u_regblank = filename_raw
            self.CheckVar_u_regblank.set(1)

    def file_upload_u_regblank_option(self):
        """
        Uploads altered U reagent blank file and once loaded marks checkbox
        """
    
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_u_regblank = openpyxl.Workbook()
            ws = filename_u_regblank.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in islice(reader, 0, int(self.rowinput_u.get()) + 9):
                    ws.append(row)
            filename_u_regblank.save("regblanku.xlsx")
            self.filename_u_regblank = "regblanku.xlsx"
            self.CheckVar_u_regblank.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:])) 
    
    
    def file_upload_u_regblankwash(self):
        """
        Uploads U reagent blank wash file and once loaded marks checkbox
        """
    
        filename_raw = filedialog.askopenfilename(parent=self)
        try:
            filename_u_regblankwash = openpyxl.Workbook()
            ws = filename_u_regblankwash.worksheets[0]
            with open(filename_raw, 'rU') as data:
                reader = csv.reader(data, delimiter = '\t')
                for row in reader:
                    ws.append(row)
            filename_u_regblankwash.save("regblanku_wash.xlsx")
            self.filename_u_regblankwash = "regblanku_wash.xlsx"
            self.CheckVar_u_regblankwash.set(1)
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            self.filename_u_regblankwash = filename_raw
            self.CheckVar_u_regblankwash.set(1)
    
    
    def blank_calculate(self):
        """
        Calculates wash and reagent blank values for all isotopes. Exports an excel file with isotope data.
        """
        
        #sample information
        self.blank_name = self.blankname.get()
        self.sln_wt = float(self.sln_wt.get())
        self.uptake_rate = float(self.uptake_rate.get())
        self.IE = (float(self.IE.get())/100) 
        self.filename = self.filename.get()
        
        #constants to be used throughout the class
        self.wt_229 = 229.031756
        self.wt_230 = 230.033128
        self.wt_232 = 232.038051
        self.wt_233 = 233.039629
        self.wt_234 = 234.040947
        self.wt_235 = 235.043924
        self.wt_236 = 236.045563
        self.wt_238 = 238.050785
        
        """
        Th wash and chem blank values
        """
        #wash 229Th
        working_a = chem_blank(self.filename_th_regblankwash, "C", "229")
        nine_wash = working_a.calc()
        
        
        #chem blank 229Th
        working_b = chem_blank(self.filename_th_regblank, "C", "229")
        nine = working_b.calc()
        
        
        #wash 230Th
        working_c = chem_blank(self.filename_th_regblankwash, "D", "230")
        zero_wash = working_c.calc()
        
        
        #chem blank 230Th
        working_d = chem_blank(self.filename_th_regblank, "D", "230")
        zero = working_d.calc()
        
        
        #wash 232Th
        working_e = chem_blank(self.filename_th_regblankwash, "E", "232")
        two_wash = working_e.calc()
        
        
        #chem blank232Th
        working_f = chem_blank(self.filename_th_regblank, "E", "232")
        two = working_f.calc()
        
        
        """
        U wash and chem blank values
        """
        
        #wash 233U
        working_g = chem_blank(self.filename_u_regblankwash, "D", "233")
        three_wash = working_g.calc()
        
        
        #chem blank 233U
        working_h = chem_blank(self.filename_u_regblank, "D", "233")
        three = working_h.calc()
        
        
        #wash 234U
        working_i = chem_blank(self.filename_u_regblankwash, "E", "234")
        four_wash = working_i.calc()
        
        
        #chem blank 234U
        working_j = chem_blank(self.filename_u_regblank, "E", "234")
        four = working_j.calc()
        
        
        #wash 235U
        working_k = chem_blank(self.filename_u_regblankwash, "F", "235")
        five_wash = working_k.calc()
        
        
        #chem blank 235U
        working_l = chem_blank(self.filename_u_regblank, "F", "235")
        five = working_l.calc()
        
        
        #wash 236U
        working_m = chem_blank(self.filename_u_regblankwash, "G", "236")
        six_wash = working_m.calc()
        
        
        #chem blank 236U
        working_n = chem_blank(self.filename_u_regblank, "G", "236")
        six = working_n.calc()
        
        
        #wash 238U
        working_o = chem_blank(self.filename_u_regblankwash, "H", "238")
        eight_wash = working_o.calc()
        
        
        #chem blank 238U
        working_p = chem_blank(self.filename_u_regblank, "H", "238")
        eight = working_p.calc()
        
        #deleting excel files
        try:
            os.remove("regblankth_wash.xlsx")
            os.remove("regblankth.xlsx")
            os.remove("regblanku_wash.xlsx")
            os.remove("regblanku.xlsx")
        except: pass
    
        """
        Calculates the reagent blank contribution to total signal and 2s error
        
        Note: [0]: mean, [1]: counts, [2] = 2s rel error
        """
        
        #229 reagent blank cps
        nine_blank = nine[0] - nine_wash[0]
        nine_blank_rel_err = np.sqrt((nine_wash[2]*nine_wash[0])**2 + (nine[2]*nine[0])**2)/abs(nine_blank)
        
        #230 reagent blank cps
        zero_blank = zero[0] - zero_wash[0]
        zero_blank_rel_err = np.sqrt((zero_wash[2]*zero_wash[0])**2 + (zero[2]*zero[0])**2)/abs(zero_blank)
        
        #232 reagent blank cps
        two_blank = two[0] - two_wash[0]
        two_blank_rel_err = np.sqrt((two_wash[2]*two_wash[0])**2 + (two[2]*two[0])**2)/abs(two_blank)
        
        #233 reagent blank cps
        three_blank = three[0] - three_wash[0]
        three_blank_rel_err = np.sqrt((three_wash[2]*three_wash[0])**2 + (three[2]*three[0])**2)/abs(three_blank)
        
        #234 reagent blank cps
        four_blank = three[0] - four_wash[0]
        four_blank_rel_err = np.sqrt((four_wash[2]*four_wash[0])**2 + (four[2]*four[0])**2)/abs(four_blank)
        
        #235 reagent blank cps
        five_blank = five[0] - five_wash[0]
        five_blank_rel_err = np.sqrt((five_wash[2]*five_wash[0])**2 + (five[2]*five[0])**2)/abs(five_blank)
        
        #236 reagent blank cps
        six_blank = six[0] - six_wash[0]
        six_blank_rel_err = np.sqrt((six_wash[2]*six_wash[0])**2 + (six[2]*six[0])**2)/abs(six_blank)
        
        #238 reagent blank cps
        eight_blank = eight[0] - eight_wash[0]
        eight_blank_rel_err = np.sqrt((eight_wash[2]*eight_wash[0])**2 + (eight[2]*eight[0])**2)/abs(eight_blank)
        
        """
        Reagent blank and error in grams
        """
        
        #229 reagent blank ag
        nine_regblank = (self.sln_wt * nine_blank * (10.**3))/(self.IE * self.uptake_rate * (6.022E23)) * self.wt_229 * (10.**18)
        nine_regblank_err = abs(nine_blank_rel_err * nine_regblank)
        
        #230 reagent blank ag
        zero_regblank = (self.sln_wt * zero_blank * (10.**3))/(self.IE * self.uptake_rate * (6.022E23)) * self.wt_230 * (10.**18)
        zero_regblank_err = abs(zero_blank_rel_err * zero_regblank)
        
        #232 reagent blank fg
        two_regblank = (self.sln_wt * two_blank * (10.**3))/(self.IE * self.uptake_rate * (6.022E23)) * self.wt_232 * (10.**15)
        two_regblank_err = abs(two_blank_rel_err * two_regblank)
        
        #233 reagent blank ag
        three_regblank = (self.sln_wt * three_blank * (10.**3))/(self.IE * self.uptake_rate * (6.022E23)) * self.wt_233 * (10.**18)
        three_regblank_err = abs(three_blank_rel_err * three_regblank)
        
        #234 reagent blank ag
        four_regblank = (self.sln_wt * four_blank * (10.**3))/(self.IE * self.uptake_rate * (6.022E23)) * self.wt_234 * (10.**18)
        four_regblank_err = abs(four_blank_rel_err * four_regblank)
        
        #235 reagent blank fg
        five_regblank = (self.sln_wt * five_blank * (10.**3))/(self.IE * self.uptake_rate * (6.022E23)) * self.wt_235 * (10.**15)
        five_regblank_err = abs(five_blank_rel_err * five_regblank)
        
        #236 reagent blank ag
        six_regblank = (self.sln_wt * six_blank * (10.**3))/(self.IE * self.uptake_rate * (6.022E23)) * self.wt_236 * (10.**18)
        six_regblank_err = abs(six_blank_rel_err * six_regblank)
        
        #238 reagent blank fg
        eight_regblank = (self.sln_wt * eight_blank * (10.**3))/(self.IE * self.uptake_rate * (6.022E23)) * self.wt_238 * (10.**15)
        eight_regblank_err = abs(eight_blank_rel_err * eight_regblank)
        
        """
        Export to Excel
        """
        
        data = {'1_Reagent_blank': pd.Series([self.blank_name, "Run info"],index = ['1_fileinfo', '5_param']),
                '229Th': pd.Series([nine_regblank, nine_regblank_err, 'ag', 'Sln wt', self.sln_wt, 'g'], index = ['2_regblank', '3_2s err', '4_units', '5_param', '6_param', '7_param']),
                '230Th': pd.Series([zero_regblank, zero_regblank_err, 'ag', 'U.R.', self.uptake_rate, 'mg/sec'], index = ['2_regblank', '3_2s err', '4_units','5_param', '6_param', '7_param']),
                '232Th': pd.Series([two_regblank, two_regblank_err, 'fg', 'I.E', (self.IE * 100), '%'], index = ['2_regblank', '3_2s err', '4_units', '5_param', '6_param', '7_param']),
                '233U': pd.Series([three_regblank, three_regblank_err, 'ag'], index = ['2_regblank', '3_2s err', '4_units']),
                '234U': pd.Series([four_regblank, four_regblank_err, 'ag'], index = ['2_regblank', '3_2s err', '4_units']),
                '235U': pd.Series([five_regblank, five_regblank_err, 'fg'], index = ['2_regblank', '3_2s err', '4_units']),
                '236U': pd.Series([six_regblank, six_regblank_err, 'ag'], index = ['2_regblank', '3_2s err', '4_units']),
                '238U': pd.Series([eight_regblank, eight_regblank_err, 'fg'], index = ['2_regblank', '3_2s err', '4_units'])}
                
        df = pd.DataFrame(data)
        
        writer = pd.ExcelWriter(self.filename, engine = 'openpyxl')
        
        df.to_excel(writer)
        
        writer.save()
        
        messagebox.showinfo("Reagent blank data file saved ! ", "Reagent blank data file name: "+ str(self.filename))

class chem_blank():
        """
        Class for analyzing ICP-MS files
        """
    
        def __init__(self,filename, columnletter, int_time):
            """
            Uploads column from specified file 
            """
            self.column = str(columnletter)+'{}:'+str(columnletter)+'{}'
            self.filename = str(filename)
            self.workbook = openpyxl.load_workbook(self.filename)
            self.ws = self.workbook.active
            
            int_time = str(int_time)
            
            int_dictionary = {"229":0.131, "230":1.049, "232":0.262, "233":0.131, "234":1.049,
                              "235": 0.262, "236":0.131, "238": 0.262}
            
            if int_time in int_dictionary:
                self.inttime = int_dictionary[int_time]
            else: print "Int_time not available"
                      
        def calc(self):
            """
            Code calculates the mean, total cycles, and 2s counting statistics error of Excel column
            """
            outlist = []
            outcounts = 0
            for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
                for cell in row:
                    if cell.value: 
                        try:
                            value = float(cell.value)
                            outlist.append(value)
                            outcounts += 1
                        except:
                            outlist.append(np.nan)
                    elif cell.value == 0:
                        value= 0.00
                        outlist.append(value)
                        outcounts += 1
                    else: outlist.append(np.nan)
            outarray = np.array(outlist, dtype = np.float)
            self.mean = np.nanmean(a = outarray)
            standdev = np.nanstd(a = outarray, ddof = 1)
            self.counts = outcounts
            err_abs =  2 * standdev/((self.counts)**0.5)
            err_rel_option1 = err_abs/self.mean
            err_rel_option2 = 2/((self.mean * self.counts*self.inttime)**0.5)
            self.err_rel = max(err_rel_option1, err_rel_option2)
            
            lst_blank = [self.mean, self.counts, self.err_rel]
        
            #returns list of mean, counts, and 2s counting error 
            return lst_blank
       
            
root = tk.Tk()
    
app = Application(master=root)

def on_closing():
    """
    If window is X'ed out of, prompts you to quit 
    """
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        root.destroy()
        root.quit()

root.protocol("WM_DELETE_WINDOW", on_closing)

root.mainloop()    
        